from tools import *
import xlsxwriter
import xlrd
from openpyxl import Workbook
from openpyxl import load_workbook
import matplotlib
from matplotlib import pyplot as plt
from pylab import *
from openpyxl.utils import get_column_letter, column_index_from_string
from xlrd import open_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
import numpy as np
from pylab import plot, show, savefig, xlim, figure,hold, ylim, legend, boxplot, setp, axes
import matplotlib.gridspec as gridspect





def get_EMG_evolution(EMG_array):
    file_excel_over30 = 'C:/Users/Rita/PycharmProjects/tonusMuscular/Excel_database/database_posturography_over30 _final.xlsx'
    file_excel_20 = 'C:/Users/Rita/PycharmProjects/tonusMuscular/Excel_database/database_posturography_20-30_final.xlsx'
    file_excel_EA = 'C:/Users/Rita/PycharmProjects/tonusMuscular/Excel_database/database_posturography_EA1_final.xlsx'

    wb_over30 = load_workbook(file_excel_over30)
    wb_20 = load_workbook(file_excel_20)
    wb_EA = load_workbook(file_excel_EA)

    count = 430
    value_final_over30 = {}
    value_final_male = {}
    value_final_female = {}
    value_final_EA_over30 = {}
    value_final_EA_evoluted = {}

    for task in EMG_array:
        over30 = {}
        female = {}
        male = {}
        EA_over30 = {}
        EA_evoluted = {}

        col_value = 18


        for index in range(0, 8):
            EMG_evolution_over30 = []
            EMG_evolution_male = []
            EMG_evolution_female = []
            EMG_evolution_EA_over30 = []
            EMG_evolution_EA_evol = []
            l = count


            ws_over30 = wb_over30.get_sheet_by_name('Statistical Analysis over 30')
            ws_20_male = wb_20.get_sheet_by_name('Statistical Analysis_20_Male')
            ws_20_female = wb_20.get_sheet_by_name('Statistical Analysis_20_Female')
            ws_EA_over30 = wb_EA.get_sheet_by_name('Statistical Analysis - EA over3')
            ws_EA_evoluted = wb_EA.get_sheet_by_name('Statistical Analysis - EA evol')


            for row in range(0,12):

                if ws_over30.cell(row= l, column= col_value).value != None:
                    EMG_evolution_over30.append(ws_over30.cell(row= l, column= col_value).value)
                if ws_20_male.cell(row= l, column= col_value).value != None:
                    EMG_evolution_male.append(ws_20_male.cell(row= l, column= col_value).value)
                if ws_20_female.cell(row= l, column= col_value).value != None:
                    EMG_evolution_female.append(ws_20_female.cell(row= l, column= col_value).value)
                if ws_EA_over30.cell(row= l, column= col_value).value != None:
                    EMG_evolution_EA_over30.append(ws_EA_over30.cell(row= l, column= col_value).value)
                if ws_EA_evoluted.cell(row= l, column= col_value).value != None:
                    EMG_evolution_EA_evol.append(ws_EA_evoluted.cell(row= l, column= col_value).value)
                l = l + 1

            if index == 0:
                male["Rectus_L"]        = EMG_evolution_male
                female["Rectus_L"]      = EMG_evolution_female
                over30["Rectus_L"]      = EMG_evolution_over30
                EA_over30["Rectus_L"]   = EMG_evolution_EA_over30
                EA_evoluted["Rectus_L"] = EMG_evolution_EA_evol

            if index == 1:
                male["Rectus_R"] = EMG_evolution_male
                female["Rectus_R"] = EMG_evolution_female
                over30["Rectus_R"] = EMG_evolution_over30
                EA_over30["Rectus_R"] = EMG_evolution_EA_over30
                EA_evoluted["Rectus_R"] = EMG_evolution_EA_evol

            if index == 2:
                male["Obliques_L"] = EMG_evolution_male
                female["Obliques_L"] = EMG_evolution_female
                over30["Obliques_L"] = EMG_evolution_over30
                EA_over30["Obliques_L"] = EMG_evolution_EA_over30
                EA_evoluted["Obliques_L"] = EMG_evolution_EA_evol

            if index == 3:
                male["Obliques_R"] = EMG_evolution_male
                female["Obliques_R"] = EMG_evolution_female
                over30["Obliques_R"] = EMG_evolution_over30
                EA_over30["Obliques_R"] = EMG_evolution_EA_over30
                EA_evoluted["Obliques_R"] = EMG_evolution_EA_evol

            if index == 4:
                male["Ilicostalis_L"] = EMG_evolution_male
                female["Ilicostalis_L"] = EMG_evolution_female
                over30["Ilicostalis_L"] = EMG_evolution_over30
                EA_over30["Ilicostalis_L"] = EMG_evolution_EA_over30
                EA_evoluted["Ilicostalis_L"] = EMG_evolution_EA_evol

            if index == 5:
                male["Ilicostalis_R"] = EMG_evolution_male
                female["Ilicostalis_R"] = EMG_evolution_female
                over30["Ilicostalis_R"] = EMG_evolution_over30
                EA_over30["Ilicostalis_R"] = EMG_evolution_EA_over30
                EA_evoluted["Ilicostalis_R"] = EMG_evolution_EA_evol

            if index == 6:
                male["Multi_L"] = EMG_evolution_male
                female["Multi_L"] = EMG_evolution_female
                over30["Multi_L"] = EMG_evolution_over30
                EA_over30["Multi_L"] = EMG_evolution_EA_over30
                EA_evoluted["Multi_L"] = EMG_evolution_EA_evol

            if index == 7:
                male["Multi_R"] = EMG_evolution_male
                female["Multi_R"] = EMG_evolution_female
                over30["Multi_R"] = EMG_evolution_over30
                EA_over30["Multi_R"] = EMG_evolution_EA_over30
                EA_evoluted["Multi_R"] = EMG_evolution_EA_evol

            col_value = col_value + 2

        value_final_over30[task] = over30
        value_final_male[task] = male
        value_final_female[task] = female
        value_final_EA_over30[task] = EA_over30
        value_final_EA_evoluted[task] = EA_evoluted
        count = count + 23

    return value_final_over30, value_final_male, value_final_female, value_final_EA_over30, value_final_EA_evoluted


def get_COP_evolution(EMG_array):
    file_excel_over30 = 'C:/Users/Rita/PycharmProjects/tonusMuscular/Excel_database/database_posturography_over30 _final.xlsx'
    file_excel_20 = 'C:/Users/Rita/PycharmProjects/tonusMuscular/Excel_database/database_posturography_20-30_final.xlsx'
    file_excel_EA = 'C:/Users/Rita/PycharmProjects/tonusMuscular/Excel_database/database_posturography_EA1_final.xlsx'

    wb_over30 = load_workbook(file_excel_over30)
    wb_20 = load_workbook(file_excel_20)
    wb_EA = load_workbook(file_excel_EA)

    count = 431
    value_final_over30 = {}
    value_final_male = {}
    value_final_female = {}
    value_final_EA_over30 = {}
    value_final_EA_evoluted = {}

    for task in EMG_array:
        over30 = {}
        female = {}
        male = {}
        EA_over30 = {}
        EA_evoluted = {}

        col_value = 2

        for index in range(0, 4):
            EMG_evolution_over30 = []
            EMG_evolution_male = []
            EMG_evolution_female = []
            EMG_evolution_EA_over30 = []
            EMG_evolution_EA_evol = []
            l = count

            ws_over30 = wb_over30.get_sheet_by_name('Statistical Analysis over 30')
            ws_20_male = wb_20.get_sheet_by_name('Statistical Analysis_20_Male')
            ws_20_female = wb_20.get_sheet_by_name('Statistical Analysis_20_Female')
            ws_EA_over30 = wb_EA.get_sheet_by_name('Statistical Analysis - EA over3')
            ws_EA_evoluted = wb_EA.get_sheet_by_name('Statistical Analysis - EA evol')

            for row in range(0, 12):

                if ws_over30.cell(row=l, column=col_value).value != None:
                    EMG_evolution_over30.append(ws_over30.cell(row=l, column=col_value).value)
                if ws_20_male.cell(row=l, column=col_value).value != None:
                    EMG_evolution_male.append(ws_20_male.cell(row=l, column=col_value).value)
                if ws_20_female.cell(row=l, column=col_value).value != None:
                    EMG_evolution_female.append(ws_20_female.cell(row=l, column=col_value).value)
                if ws_EA_over30.cell(row=l, column=col_value).value != None:
                    EMG_evolution_EA_over30.append(ws_EA_over30.cell(row=l, column=col_value).value)
                if ws_EA_evoluted.cell(row=l, column=col_value).value != None:
                    EMG_evolution_EA_evol.append(ws_EA_evoluted.cell(row=l, column=col_value).value)
                l = l + 1


            if index == 0:
                male["STD_X"] = EMG_evolution_male
                female["STD_X"] = EMG_evolution_female
                over30["STD_X"] = EMG_evolution_over30
                EA_over30["STD_X"] = EMG_evolution_EA_over30
                EA_evoluted["STD_X"] = EMG_evolution_EA_evol

            if index == 1:
                male["Vel_X"] = EMG_evolution_male
                female["Vel_X"] = EMG_evolution_female
                over30["Vel_X"] = EMG_evolution_over30
                EA_over30["Vel_X"] = EMG_evolution_EA_over30
                EA_evoluted["Vel_X"] = EMG_evolution_EA_evol

            if index == 2:
                male["STD_Y"] = EMG_evolution_male
                female["STD_Y"] = EMG_evolution_female
                over30["STD_Y"] = EMG_evolution_over30
                EA_over30["STD_Y"] = EMG_evolution_EA_over30
                EA_evoluted["STD_Y"] = EMG_evolution_EA_evol

            if index == 3:
                male["Vel_Y"] = EMG_evolution_male
                female["Vel_Y"] = EMG_evolution_female
                over30["Vel_Y"] = EMG_evolution_over30
                EA_over30["Vel_Y"] = EMG_evolution_EA_over30
                EA_evoluted["Vel_Y"] = EMG_evolution_EA_evol

            col_value = col_value + 2

        value_final_over30[task] = over30
        value_final_male[task] = male
        value_final_female[task] = female
        value_final_EA_over30[task] = EA_over30
        value_final_EA_evoluted[task] = EA_evoluted
        count = count + 23

    return value_final_over30, value_final_male, value_final_female, value_final_EA_over30, value_final_EA_evoluted


def get_evolution_area(area_array):
    file_excel_over30 = 'C:/Users/Rita/PycharmProjects/tonusMuscular/Excel_database/database_posturography_over30 _final.xlsx'
    file_excel_20 = 'C:/Users/Rita/PycharmProjects/tonusMuscular/Excel_database/database_posturography_20-30_final.xlsx'
    file_excel_EA = 'C:/Users/Rita/PycharmProjects/tonusMuscular/Excel_database/database_posturography_EA1_final.xlsx'

    wb_over30 = load_workbook(file_excel_over30)
    wb_20 = load_workbook(file_excel_20)
    wb_EA = load_workbook(file_excel_EA)

    count = 430
    value_final_over30 = {}
    value_final_male = {}
    value_final_female = {}
    value_final_EA_over30 = {}
    value_final_EA_evoluted = {}

    for task in area_array:


        col_value = 13


        area_evolution_over30 = []
        area_evolution_male = []
        area_evolution_female = []
        area_evolution_EA_over30 = []
        area_evolution_EA_evol = []
        l = count

        ws_over30 = wb_over30.get_sheet_by_name('Statistical Analysis over 30')
        ws_20_male = wb_20.get_sheet_by_name('Statistical Analysis_20_Male')
        ws_20_female = wb_20.get_sheet_by_name('Statistical Analysis_20_Female')
        ws_EA_over30 = wb_EA.get_sheet_by_name('Statistical Analysis - EA over3')
        ws_EA_evoluted = wb_EA.get_sheet_by_name('Statistical Analysis - EA evol')

        for row in range(0, 12):

            if ws_over30.cell(row=l, column=col_value).value != None:
                area_evolution_over30.append(ws_over30.cell(row=l, column=col_value).value)
            if ws_20_male.cell(row=l, column=col_value).value != None:
                area_evolution_male.append(ws_20_male.cell(row=l, column=col_value).value)
            if ws_20_female.cell(row=l, column=col_value).value != None:
                area_evolution_female.append(ws_20_female.cell(row=l, column=col_value).value)
            if ws_EA_over30.cell(row=l, column=col_value).value != None:
                area_evolution_EA_over30.append(ws_EA_over30.cell(row=l, column=col_value).value)
            if ws_EA_evoluted.cell(row=l, column=col_value).value != None:
                area_evolution_EA_evol.append(ws_EA_evoluted.cell(row=l, column=col_value).value)
            l = l + 1

        value_final_over30[task] = area_evolution_over30
        value_final_male[task] = area_evolution_male
        value_final_female[task] = area_evolution_female
        value_final_EA_over30[task] = area_evolution_EA_over30
        value_final_EA_evoluted[task] = area_evolution_EA_evol
        count = count + 23

    return value_final_over30, value_final_male, value_final_female, value_final_EA_over30, value_final_EA_evoluted

def plot_evalution_graphs(EMG_over30, EMG_male, EMG_female, EMG_EA_over, EMG_EA_evol,
                          COP_over30, COP_male, COP_female, COP_EA_over, COP_EA_evol,
                          area_over30, area_male, area_female, area_EA_over, area_EA_evol):
    l = 0
    pp = PdfPages('Evalution_EMG_COP.pdf')

    for i in EMG_over30:
        fig = plt.figure(l)
        l = l + 1
        fig.suptitle(str(i) + " - Evalution arrays", fontsize=21)

        plot1 = plt.subplot2grid((4, 4), (0, 0))

        plt.plot(EMG_over30[i]["Rectus_L"], color='blue', label="Over 30")
        plt.plot(EMG_male[i]["Rectus_L"], color='yellow', label="Male 20-30")
        plt.plot(EMG_female[i]["Rectus_L"], color='green', label="Female 20-30")
        plt.plot(EMG_EA_over[i]["Rectus_L"], color='red', label="EA Over 30")
        plt.plot(EMG_EA_evol[i]["Rectus_L"], color='magenta', label="EA evolved")

        if i != "Reach_C" and i != "Reach_L" and i != "Reach_R":
            plt.xticks([0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11], ("1st", "2nd", "3rd", "4th",
                                                                "5th", "6th", "7th",
                                                                "8th",
                                                                "9th", "10th", "11th",
                                                                "12th"), fontsize=8)
        if i == "Reach_C" or i == "Reach_L" or i == "Reach_R":
            plt.xticks([0, 1, 2, 3, 4, 5, 6, 7], ("1st", "2nd", "3rd", "4th", "5th",
                                                  "6th", "7th"), fontsize=8)

        plt.ylabel('EMG Values', fontsize=10)
        #plt.legend(bbox_to_anchor=(1.0, 1.0), loc=2, borderaxespad=0., fontsize=10)
        plot1.set_title("Rectus Left", fontsize=12, bbox=dict(boxstyle="round, pad=0.3", fc="w", color ='royalblue', linewidth = 1))



        plot2 = plt.subplot2grid((4, 4), (0, 1))

        plt.plot(EMG_over30[i]["Obliques_L"], color='blue', label="Over 30")
        plt.plot(EMG_male[i]["Obliques_L"], color='yellow', label="Male 20-30")
        plt.plot(EMG_female[i]["Obliques_L"], color='green', label="Female 20-30")
        plt.plot(EMG_EA_over[i]["Obliques_L"], color='red', label="EA Over 30")
        plt.plot(EMG_EA_evol[i]["Obliques_L"], color='magenta', label="EA evolved")

        if i != "Reach_C" and i != "Reach_L" and i != "Reach_R":
            plt.xticks([0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11], ("1st", "2nd", "3rd", "4th",
                                                                "5th", "6th", "7th",
                                                                "8th",
                                                                "9th", "10th", "11th",
                                                                "12th"), fontsize=8)
        if i == "Reach_C" or i == "Reach_L" or i == "Reach_R":
            plt.xticks([0, 1, 2, 3, 4, 5, 6, 7], ("1st", "2nd", "3rd", "4th", "5th",
                                                  "6th", "7th"), fontsize=8)

        plt.ylabel('EMG Values', fontsize=10)
        #plt.legend(bbox_to_anchor=(1.0, 1.0), loc=2, borderaxespad=0., fontsize=10)
        plot2.set_title("Obliques Left",fontsize=12, bbox=dict(boxstyle="round, pad=0.3", fc="w", color ='royalblue', linewidth = 1))

        plot3 = plt.subplot2grid((4, 4), (0, 2))

        plt.plot(EMG_over30[i]["Rectus_R"], color='blue', label="Over 30")
        plt.plot(EMG_male[i]["Rectus_R"], color='yellow', label="Male 20-30")
        plt.plot(EMG_female[i]["Rectus_R"], color='green', label="Female 20-30")
        plt.plot(EMG_EA_over[i]["Rectus_R"], color='red', label="EA Over 30")
        plt.plot(EMG_EA_evol[i]["Rectus_R"], color='magenta', label="EA evolved")

        if i != "Reach_C" and i != "Reach_L" and i != "Reach_R":
            plt.xticks([0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11], ("1st", "2nd", "3rd", "4th",
                                                                "5th", "6th", "7th",
                                                                "8th",
                                                                "9th", "10th", "11th",
                                                                "12th"), fontsize=8)
        if i == "Reach_C" or i == "Reach_L" or i == "Reach_R":
            plt.xticks([0, 1, 2, 3, 4, 5, 6, 7], ("1st", "2nd", "3rd", "4th", "5th",
                                                  "6th", "7th"), fontsize=8)

        plt.ylabel('EMG Values', fontsize=10)
        #plt.legend(bbox_to_anchor=(1.0, 1.0), loc=2, borderaxespad=0., fontsize=10)
        plot3.set_title("Rectus Right", fontsize=12, bbox=dict(boxstyle="round, pad=0.3", fc="w", color ='royalblue', linewidth = 1))

        plot4 = plt.subplot2grid((4, 4), (0, 3))

        plt.plot(EMG_over30[i]["Obliques_R"], color='blue', label="Over 30")
        plt.plot(EMG_male[i]["Obliques_R"], color='yellow', label="Male 20-30")
        plt.plot(EMG_female[i]["Obliques_R"], color='green', label="Female 20-30")
        plt.plot(EMG_EA_over[i]["Obliques_R"], color='red', label="EA Over 30")
        plt.plot(EMG_EA_evol[i]["Obliques_R"], color='magenta', label="EA evolved")

        if i != "Reach_C" and i != "Reach_L" and i != "Reach_R":
            plt.xticks([0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11], ("1st", "2nd", "3rd", "4th",
                                                                "5th", "6th", "7th",
                                                                "8th",
                                                                "9th", "10th", "11th",
                                                                "12th"), fontsize=8)
        if i == "Reach_C" or i == "Reach_L" or i == "Reach_R":
            plt.xticks([0, 1, 2, 3, 4, 5, 6, 7], ("1st", "2nd", "3rd", "4th", "5th",
                                                  "6th", "7th"), fontsize=8)

        plt.ylabel('EMG Values', fontsize=10)
        #plt.legend(bbox_to_anchor=(1.0, 1.0), loc=2, borderaxespad=0., fontsize=10)
        plot4.set_title("Obliques Right",fontsize=12, bbox=dict(boxstyle="round, pad=0.3", fc="w", color ='royalblue', linewidth = 1))

        plot5 = plt.subplot2grid((4, 4), (1, 0))

        plt.plot(EMG_over30[i]["Ilicostalis_L"], color='blue', label="Over 30")
        plt.plot(EMG_male[i]["Ilicostalis_L"], color='yellow', label="Male 20-30")
        plt.plot(EMG_female[i]["Ilicostalis_L"], color='green', label="Female 20-30")
        plt.plot(EMG_EA_over[i]["Ilicostalis_L"], color='red', label="EA Over 30")
        plt.plot(EMG_EA_evol[i]["Ilicostalis_L"], color='magenta', label="EA evolved")

        if i != "Reach_C" and i != "Reach_L" and i != "Reach_R":
            plt.xticks([0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11], ("1st", "2nd", "3rd", "4th",
                                                                "5th", "6th", "7th",
                                                                "8th",
                                                                "9th", "10th", "11th",
                                                                "12th"), fontsize=8)
        if i == "Reach_C" or i == "Reach_L" or i == "Reach_R":
            plt.xticks([0, 1, 2, 3, 4, 5, 6, 7], ("1st", "2nd", "3rd", "4th", "5th",
                                                  "6th", "7th"), fontsize=8)

        plt.ylabel('EMG Values', fontsize=10)
        #plt.legend(bbox_to_anchor=(1.0, 1.0), loc=2, borderaxespad=0., fontsize=10)
        plot5.set_title("Ilicostalis Left", fontsize=12, bbox=dict(boxstyle="round, pad=0.3", fc="w", color ='royalblue', linewidth = 1))

        plot6 = plt.subplot2grid((4, 4), (1, 1))

        plt.plot(EMG_over30[i]["Multi_L"], color='blue', label="Over 30")
        plt.plot(EMG_male[i]["Multi_L"], color='yellow', label="Male 20-30")
        plt.plot(EMG_female[i]["Multi_L"], color='green', label="Female 20-30")
        plt.plot(EMG_EA_over[i]["Multi_L"], color='red', label="EA Over 30")
        plt.plot(EMG_EA_evol[i]["Multi_L"], color='magenta', label="EA evolved")

        if i != "Reach_C" and i != "Reach_L" and i != "Reach_R":
            plt.xticks([0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11], ("1st", "2nd", "3rd", "4th",
                                                                "5th", "6th", "7th",
                                                                "8th",
                                                                "9th", "10th", "11th",
                                                                "12th"), fontsize=8)
        if i == "Reach_C" or i == "Reach_L" or i == "Reach_R":
            plt.xticks([0, 1, 2, 3, 4, 5, 6, 7], ("1st", "2nd", "3rd", "4th", "5th",
                                                  "6th", "7th"), fontsize=8)

        plt.ylabel('EMG Values', fontsize=10)
        #plt.legend(bbox_to_anchor=(1.0, 1.0), loc=2, borderaxespad=0., fontsize=10)
        plot6.set_title("Multifidus Left",fontsize=12, bbox=dict(boxstyle="round, pad=0.3", fc="w", color ='royalblue', linewidth = 1))

        plot7 = plt.subplot2grid((4, 4), (1, 2))

        plt.plot(EMG_over30[i]["Ilicostalis_R"], color='blue', label="Over 30")
        plt.plot(EMG_male[i]["Ilicostalis_R"], color='yellow', label="Male 20-30")
        plt.plot(EMG_female[i]["Ilicostalis_R"], color='green', label="Female 20-30")
        plt.plot(EMG_EA_over[i]["Ilicostalis_R"], color='red', label="EA Over 30")
        plt.plot(EMG_EA_evol[i]["Ilicostalis_R"], color='magenta', label="EA evolved")

        if i != "Reach_C" and i != "Reach_L" and i != "Reach_R":
            plt.xticks([0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11], ("1st", "2nd", "3rd", "4th",
                                                                "5th", "6th", "7th",
                                                                "8th",
                                                                "9th", "10th", "11th",
                                                                "12th"), fontsize=8)
        if i == "Reach_C" or i == "Reach_L" or i == "Reach_R":
            plt.xticks([0, 1, 2, 3, 4, 5, 6, 7], ("1st", "2nd", "3rd", "4th", "5th",
                                                  "6th", "7th"), fontsize=8)

        plt.ylabel('EMG Values', fontsize=10)
        #plt.legend(bbox_to_anchor=(1.0, 1.0), loc=2, borderaxespad=0., fontsize=10)
        plot7.set_title("Ilicostalis Right",fontsize=12, bbox=dict(boxstyle="round, pad=0.3", fc="w", color ='royalblue', linewidth = 1))

        plot8 = plt.subplot2grid((4, 4), (1, 3))

        plt.plot(EMG_over30[i]["Multi_R"], color='blue', label="Over 30")
        plt.plot(EMG_male[i]["Multi_R"], color='yellow', label="Male 20-30")
        plt.plot(EMG_female[i]["Multi_R"], color='green', label="Female 20-30")
        plt.plot(EMG_EA_over[i]["Multi_R"], color='red', label="EA Over 30")
        plt.plot(EMG_EA_evol[i]["Multi_R"], color='magenta', label="EA evolved")

        if i != "Reach_C" and i != "Reach_L" and i != "Reach_R":
            plt.xticks([0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11], ("1st", "2nd", "3rd", "4th",
                                                                "5th", "6th", "7th",
                                                                "8th",
                                                                "9th", "10th", "11th",
                                                                "12th"), fontsize=8)
        if i == "Reach_C" or i == "Reach_L" or i == "Reach_R":
            plt.xticks([0, 1, 2, 3, 4, 5, 6, 7], ("1st", "2nd", "3rd", "4th", "5th",
                                                  "6th", "7th"), fontsize=8)

        plt.ylabel('EMG Values', fontsize=10)
        #plt.legend(bbox_to_anchor=(1.0, 1.0), loc=2, borderaxespad=0., fontsize=10)
        plot8.set_title("Multifidus Right",fontsize=12, bbox=dict(boxstyle="round, pad=0.3", fc="w", color ='royalblue', linewidth = 1))

        plot9 = plt.subplot2grid((4, 4), (2, 0))

        plt.plot(COP_over30[i]["STD_X"], color='blue', label="Over 30")
        plt.plot(COP_male[i]["STD_X"], color='yellow', label="Male 20-30")
        plt.plot(COP_female[i]["STD_X"], color='green', label="Female 20-30")
        plt.plot(COP_EA_over[i]["STD_X"], color='red', label="EA Over 30")
        plt.plot(COP_EA_evol[i]["STD_X"], color='magenta', label="EA evolved")

        if i != "Reach_C" and i != "Reach_L" and i != "Reach_R":
            plt.xticks([0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11], ("1st", "2nd", "3rd", "4th",
                                                                "5th", "6th", "7th",
                                                                "8th",
                                                                "9th", "10th", "11th",
                                                                "12th"), fontsize=8)
        if i == "Reach_C" or i == "Reach_L" or i == "Reach_R":
            plt.xticks([0, 1, 2, 3, 4, 5, 6, 7], ("1st", "2nd", "3rd", "4th", "5th",
                                                  "6th", "7th"), fontsize=8)

        plt.ylabel('STD COP X Values', fontsize=10)
        #plt.legend(bbox_to_anchor=(1.0, 1.0), loc=2, borderaxespad=0., fontsize=10)
        plot9.set_title("STD X(mm)",fontsize=12, bbox=dict(boxstyle="round, pad=0.3", fc="w", color ='mediumaquamarine', linewidth = 1))

        plot10 = plt.subplot2grid((4, 4), (2, 1))

        plt.plot(COP_over30[i]["STD_Y"], color='blue', label="Over 30")
        plt.plot(COP_male[i]["STD_Y"], color='yellow', label="Male 20-30")
        plt.plot(COP_female[i]["STD_Y"], color='green', label="Female 20-30")
        plt.plot(COP_EA_over[i]["STD_Y"], color='red', label="EA Over 30")
        plt.plot(COP_EA_evol[i]["STD_Y"], color='magenta', label="EA evolved")

        if i != "Reach_C" and i != "Reach_L" and i != "Reach_R":
            plt.xticks([0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11], ("1st", "2nd", "3rd", "4th",
                                                                "5th", "6th", "7th",
                                                                "8th",
                                                                "9th", "10th", "11th",
                                                                "12th"), fontsize=8)
        if i == "Reach_C" or i == "Reach_L" or i == "Reach_R":
            plt.xticks([0, 1, 2, 3, 4, 5, 6, 7], ("1st", "2nd", "3rd", "4th", "5th",
                                                  "6th", "7th"), fontsize=8)

        plt.ylabel('STD COP Y Values', fontsize=10)
        #plt.legend(bbox_to_anchor=(1.0, 1.0), loc=2, borderaxespad=0., fontsize=10)
        plot10.set_title("STD Y (mm)", fontsize=12, bbox=dict(boxstyle="round, pad=0.3", fc="w", color ='mediumaquamarine', linewidth = 1))

        plot11 = plt.subplot2grid((4, 4), (2, 2))

        plt.plot(COP_over30[i]["Vel_X"], color='blue', label="Over 30")
        plt.plot(COP_male[i]["Vel_X"], color='yellow', label="Male 20-30")
        plt.plot(COP_female[i]["Vel_X"], color='green', label="Female 20-30")
        plt.plot(COP_EA_over[i]["Vel_X"], color='red', label="EA Over 30")
        plt.plot(COP_EA_evol[i]["Vel_X"], color='magenta', label="EA evolved")

        if i != "Reach_C" and i != "Reach_L" and i != "Reach_R":
            plt.xticks([0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11], ("1st", "2nd", "3rd", "4th",
                                                                "5th", "6th", "7th",
                                                                "8th",
                                                                "9th", "10th", "11th",
                                                                "12th"), fontsize=8)
        if i == "Reach_C" or i == "Reach_L" or i == "Reach_R":
            plt.xticks([0, 1, 2, 3, 4, 5, 6, 7], ("1st", "2nd", "3rd", "4th", "5th",
                                                  "6th", "7th"), fontsize=8)
        plt.ylabel('Velocity COP X Values', fontsize=10)
        #plt.legend(bbox_to_anchor=(1.0, 1.0), loc=2, borderaxespad=0., fontsize=10)
        plot11.set_title("Velocity X (mm/s)",fontsize=12, bbox=dict(boxstyle="round, pad=0.3", fc="w", color ='mediumaquamarine', linewidth = 1))

        plot12 = plt.subplot2grid((4, 4), (2, 3))

        plt.plot(COP_over30[i]["Vel_Y"], color='blue', label="Over 30")
        plt.plot(COP_male[i]["Vel_Y"], color='yellow', label="Male 20-30")
        plt.plot(COP_female[i]["Vel_Y"], color='green', label="Female 20-30")
        plt.plot(COP_EA_over[i]["Vel_Y"], color='red', label="EA Over 30")
        plt.plot(COP_EA_evol[i]["Vel_Y"], color='magenta', label="EA evolved")

        if i != "Reach_C" and i != "Reach_L" and i != "Reach_R":
            plt.xticks([0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11], ("1st", "2nd", "3rd", "4th",
                                                                "5th", "6th", "7th",
                                                                "8th",
                                                                "9th", "10th", "11th",
                                                                "12th"), fontsize=8)
        if i == "Reach_C" or i == "Reach_L" or i == "Reach_R":
            plt.xticks([0, 1, 2, 3, 4, 5, 6, 7], ("1st", "2nd", "3rd", "4th", "5th",
                                                  "6th", "7th"), fontsize=8)

        plt.ylabel('Velocity COP Y Values', fontsize=10)
        #plt.legend(bbox_to_anchor=(1.0, 1.0), loc=2, borderaxespad=0., fontsize=10)
        plot12.set_title("Velocity Y (mm/s)",fontsize=12, bbox=dict(boxstyle="round, pad=0.3", fc="w", color ='mediumaquamarine', linewidth = 1))

        plot13 = plt.subplot2grid((4, 4), (3, 0), colspan=2, rowspan=1)

        plt.plot(area_over30[i], color='blue', label="Over 30")
        plt.plot(area_male[i], color='yellow', label="Male 20-30")
        plt.plot(area_female[i], color='green', label="Female 20-30")
        plt.plot(area_EA_over[i], color='red', label="EA Over 30")
        plt.plot(area_EA_evol[i], color='magenta', label="EA evolved")

        if i != "Reach_C" and i != "Reach_L" and i != "Reach_R":
            plt.xticks([0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11], ("1st", "2nd", "3rd", "4th",
                                                                "5th", "6th", "7th",
                                                                "8th",
                                                                "9th", "10th", "11th",
                                                                "12th"), fontsize=8)
        if i == "Reach_C" or i == "Reach_L" or i == "Reach_R":
            plt.xticks([0, 1, 2, 3, 4, 5, 6, 7], ("1st", "2nd", "3rd", "4th", "5th",
                                                  "6th", "7th"), fontsize=8)

        plt.ylabel('Area Values', fontsize=10)
        plt.legend(bbox_to_anchor=(1.1, 1.0), loc=2, borderaxespad=0., fontsize=10)
        plot13.set_title("Area (mm2)", fontsize=12, bbox=dict(boxstyle="round, pad=1", fc="w", color ='mediumaquamarine', linewidth = 1))

        plot14 = plt.subplot2grid((4, 4), (3, 2), colspan=2, rowspan=1)
        plt.axis('off')

        if i != "Reach_C" and i != "Reach_L" and i != "Reach_R":
            plot14.annotate("Legend x axis: \n"
                            "\n"
                            "1st - [0 - 2,5]s       7th - [15 - 17,5]s \n"
                            "2nd - [2,5 - 5]s       8th - [17,5 - 20]s \n"
                            "3rd - [5 - 7,5]s       9th - [20 - 22,5]s \n"
                            "4th - [7,5 - 10]s      10th - [22,5 - 25]s \n"
                            "5th - [10 - 12,5]s     11th - [25 - 27,5]s \n"
                            "6th - [12,5 - 15]s     12th - [27,5 - 30]s \n", (0.5, 0.5),
                     xycoords="axes fraction", va="center", ha="left",
                     bbox=dict(boxstyle="round, pad=1", fc="w"))
        if i == "Reach_C" or i == "Reach_L" or i == "Reach_R":
            plot14.annotate("Legend x axis: \n"
                            "\n"
                            "1st - [0 - 1]s         5th - [4 - 5]s \n"
                            "2nd - [1 - 2]s         6th - [5 - 6]s \n"
                            "3rd - [2 - 3]s         7th - [6 - 7]s \n", (0.5, 0.5),
                            xycoords="axes fraction", va="center", ha="left",
                            bbox=dict(boxstyle="round, pad=1", fc="w"))



        plt.subplots_adjust(top=0.90, bottom=0.10, left=0.07, right=0.86, wspace=0.37, hspace=0.68)
        plt.show()
        pp.savefig(fig)
    pp.close()