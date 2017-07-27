from tools import *
import xlsxwriter
import xlrd
from openpyxl import Workbook
from openpyxl import load_workbook
from xlrd import open_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
import numpy as np
from pylab import plot, show, savefig, xlim, figure,hold, ylim, legend, boxplot, setp, axes
import matplotlib.gridspec as gridspec
import matplotlib.pyplot as plt

file_excel_over30 = 'C:/Users/Rita/PycharmProjects/tonusMuscular/Excel_database/database_posturography_over30 _obliques.xlsx'
file_excel_20 = 'C:/Users/Rita/PycharmProjects/tonusMuscular/Excel_database/database_posturography_20-30_obliques.xlsx'
file_excel_EA = 'C:/Users/Rita/PycharmProjects/tonusMuscular/Excel_database/database_posturography_EA1_obliques.xlsx'



def get_value_tonus(EMG_array):

    wb_over30 = load_workbook(file_excel_over30)
    wb_20 = load_workbook(file_excel_20)
    wb_EA = load_workbook(file_excel_EA)


    count = 8
    value_final_over30 = {}
    value_final_male = {}
    value_final_female = {}
    value_final_EA_over30 = {}
    value_final_EA_evoluted = {}

    for task in EMG_array:

        value_rectus_l_over30 = []
        value_rectus_r_over30 = []
        value_obliques_l_over30 = []
        value_obliques_r_over30 = []
        value_ilicostalis_l_over30 = []
        value_ilicostalis_r_over30 = []
        value_multi_l_over30 = []
        value_multi_r_over30 = []

        value_rectus_l_female = []
        value_rectus_r_female = []
        value_obliques_l_female = []
        value_obliques_r_female = []
        value_ilicostalis_l_female = []
        value_ilicostalis_r_female = []
        value_multi_l_female = []
        value_multi_r_female = []

        value_rectus_l_male = []
        value_rectus_r_male = []
        value_obliques_l_male = []
        value_obliques_r_male = []
        value_ilicostalis_l_male = []
        value_ilicostalis_r_male = []
        value_multi_l_male = []
        value_multi_r_male = []

        value_rectus_l_EA_over30 = []
        value_rectus_r_EA_over30 = []
        value_obliques_l_EA_over30 = []
        value_obliques_r_EA_over30 = []
        value_ilicostalis_l_EA_over30 = []
        value_ilicostalis_r_EA_over30 = []
        value_multi_l_EA_over30 = []
        value_multi_r_EA_over30 = []

        value_rectus_l_EA_evoluted = []
        value_rectus_r_EA_evoluted = []
        value_obliques_l_EA_evoluted = []
        value_obliques_r_EA_evoluted = []
        value_ilicostalis_l_EA_evoluted = []
        value_ilicostalis_r_EA_evoluted = []
        value_multi_l_EA_evoluted = []
        value_multi_r_EA_evoluted = []

        for i in wb_over30.sheetnames:
            if i != 'Sheet1' and i != 'Statistical Analysis_20_Female' and i!= 'Statistical Analysis_20_Male'\
                    and i != 'Statistical Analysis over 30' and i != 'Folha1':

                ws_over30 = wb_over30.get_sheet_by_name(i)
                value_rectus_l_over30.append(ws_over30['B' + str(count)].value)
                value_rectus_r_over30.append(ws_over30['C' + str(count)].value)
                value_obliques_l_over30.append(ws_over30['B' + str(count+1)].value)
                value_obliques_r_over30.append(ws_over30['C' + str(count+1)].value)
                value_ilicostalis_l_over30.append(ws_over30['B' + str(count+2)].value)
                value_ilicostalis_r_over30.append(ws_over30['C' + str(count+2)].value)
                value_multi_l_over30.append(ws_over30['B' + str(count+3)].value)
                value_multi_r_over30.append(ws_over30['C' + str(count+3)].value)

        for i in wb_20.sheetnames:
            if i != 'Sheet1' and i != 'Statistical Analysis_20_Female' and i!= 'Statistical Analysis_20_Male'\
                    and i != 'Statistical Analysis over 30' and i != 'Folha1' and i != 'Patient5_Healthy'\
                    and i != 'Patient8_Healthy' and i != 'Patient10_Healthy' and i != 'Patient13_Healthy'\
                    and i != 'Patient17_Healthy' and i != 'Patient18_Healthy' and i != 'Patient23_Healthy'\
                    and i != 'Patient25_Healthy' and i != 'Patient27_Healthy' and i != 'Patient30_Healthy'\
                    and i != 'Patient37_Healthy' and i != 'Patient38_Healthy':

                ws_female = wb_20.get_sheet_by_name(i)
                value_rectus_l_female.append(ws_female['B' + str(count)].value)
                value_rectus_r_female.append(ws_female['C' + str(count)].value)
                value_obliques_l_female.append(ws_female['B' + str(count+1)].value)
                value_obliques_r_female.append(ws_female['C' + str(count+1)].value)
                value_ilicostalis_l_female.append(ws_female['B' + str(count+2)].value)
                value_ilicostalis_r_female.append(ws_female['C' + str(count+2)].value)
                value_multi_l_female.append(ws_female['B' + str(count+3)].value)
                value_multi_r_female.append(ws_female['C' + str(count+3)].value)

            if i == 'Patient8_Healthy' or i == 'Patient10_Healthy' or i == 'Patient13_Healthy'\
                    or i == 'Patient17_Healthy' or i == 'Patient18_Healthy' or i == 'Patient23_Healthy'\
                    or i == 'Patient25_Healthy' or i == 'Patient27_Healthy' or i == 'Patient30_Healthy' \
                    or i == 'Patient5_Healthy' or i == 'Patient37_Healthy' or i == 'Patient38_Healthy':

                ws_male = wb_20.get_sheet_by_name(i)
                value_rectus_l_male.append(ws_male['B' + str(count)].value)
                value_rectus_r_male.append(ws_male['C' + str(count)].value)
                value_obliques_l_male.append(ws_male['B' + str(count + 1)].value)
                value_obliques_r_male.append(ws_male['C' + str(count + 1)].value)
                value_ilicostalis_l_male.append(ws_male['B' + str(count + 2)].value)
                value_ilicostalis_r_male.append(ws_male['C' + str(count + 2)].value)
                value_multi_l_male.append(ws_male['B' + str(count + 3)].value)
                value_multi_r_male.append(ws_male['C' + str(count + 3)].value)


        for i in wb_EA.sheetnames:
            if i != 'Sheet1' and i != 'Patient4_EA' and i != 'Patient8_EA' and i != 'Patient11_EA'\
                    and i != "Statistical Analysis - EA evol" and i != "Statistical Analysis - EA over3":

                ws_EA_over30 = wb_EA.get_sheet_by_name(i)
                value_rectus_l_EA_over30.append(ws_EA_over30['B' + str(count)].value)
                value_rectus_r_EA_over30.append(ws_EA_over30['C' + str(count)].value)
                value_obliques_l_EA_over30.append(ws_EA_over30['B' + str(count+1)].value)
                value_obliques_r_EA_over30.append(ws_EA_over30['C' + str(count+1)].value)
                value_ilicostalis_l_EA_over30.append(ws_EA_over30['B' + str(count+2)].value)
                value_ilicostalis_r_EA_over30.append(ws_EA_over30['C' + str(count+2)].value)
                value_multi_l_EA_over30.append(ws_EA_over30['B' + str(count+3)].value)
                value_multi_r_EA_over30.append(ws_EA_over30['C' + str(count+3)].value)

            if i == 'Patient4_EA' or i == 'Patient8_EA' or i == 'Patient11_EA':

                ws_EA_evoluted = wb_EA.get_sheet_by_name(i)
                value_rectus_l_EA_evoluted.append(ws_EA_evoluted['B' + str(count)].value)
                value_rectus_r_EA_evoluted.append(ws_EA_evoluted['C' + str(count)].value)
                value_obliques_l_EA_evoluted.append(ws_EA_evoluted['B' + str(count+1)].value)
                value_obliques_r_EA_evoluted.append(ws_EA_evoluted['C' + str(count+1)].value)
                value_ilicostalis_l_EA_evoluted.append(ws_EA_evoluted['B' + str(count+2)].value)
                value_ilicostalis_r_EA_evoluted.append(ws_EA_evoluted['C' + str(count+2)].value)
                value_multi_l_EA_evoluted.append(ws_EA_evoluted['B' + str(count+3)].value)
                value_multi_r_EA_evoluted.append(ws_EA_evoluted['C' + str(count+3)].value)

        count = count + 13

        value_final_over30[task] = {"Rectus_L": value_rectus_l_over30,"Rectus_R": value_rectus_r_over30, "Obliques_L": value_obliques_l_over30,
                             "Obliques_R": value_obliques_r_over30, "Ilicostalis_L": value_ilicostalis_l_over30, "Ilicostalis_R": value_ilicostalis_r_over30,
                             "Multifidus_L": value_multi_l_over30, "Multifidus_R": value_multi_r_over30}

        value_final_male[task] = {"Rectus_L": value_rectus_l_male, "Rectus_R": value_rectus_r_male,
                                    "Obliques_L": value_obliques_l_male,
                                    "Obliques_R": value_obliques_r_male, "Ilicostalis_L": value_ilicostalis_l_male,
                                    "Ilicostalis_R": value_ilicostalis_r_male,
                                    "Multifidus_L": value_multi_l_male, "Multifidus_R": value_multi_r_male}

        value_final_female[task] = {"Rectus_L": value_rectus_l_female, "Rectus_R": value_rectus_r_female,
                                    "Obliques_L": value_obliques_l_female,
                                    "Obliques_R": value_obliques_r_female, "Ilicostalis_L": value_ilicostalis_l_female,
                                    "Ilicostalis_R": value_ilicostalis_r_female,
                                    "Multifidus_L": value_multi_l_female, "Multifidus_R": value_multi_r_female}

        value_final_EA_over30[task] = {"Rectus_L": value_rectus_l_EA_over30, "Rectus_R": value_rectus_r_EA_over30,
                                    "Obliques_L": value_obliques_l_EA_over30,
                                    "Obliques_R": value_obliques_r_EA_over30, "Ilicostalis_L": value_ilicostalis_l_EA_over30,
                                    "Ilicostalis_R": value_ilicostalis_r_EA_over30,
                                    "Multifidus_L": value_multi_l_EA_over30, "Multifidus_R": value_multi_r_EA_over30}

        value_final_EA_evoluted[task] = {"Rectus_L": value_rectus_l_EA_evoluted, "Rectus_R": value_rectus_r_EA_evoluted,
                                "Obliques_L": value_obliques_l_EA_evoluted,
                                "Obliques_R": value_obliques_r_EA_evoluted, "Ilicostalis_L": value_ilicostalis_l_EA_evoluted,
                                "Ilicostalis_R": value_ilicostalis_r_EA_evoluted,
                                "Multifidus_L": value_multi_l_EA_evoluted, "Multifidus_R": value_multi_r_EA_evoluted}

    return value_final_over30, value_final_male, value_final_female, value_final_EA_over30, value_final_EA_evoluted




def get_value_freq(freq_array):

    wb_over30 = load_workbook(file_excel_over30)
    wb_20 = load_workbook(file_excel_20)
    wb_EA = load_workbook(file_excel_EA)


    count = 260
    value_final_over30 = {}
    value_final_male = {}
    value_final_female = {}
    value_final_EA_over30 = {}
    value_final_EA_evoluted = {}

    for task in freq_array:
        over30 = {}
        female = {}
        male = {}
        EA_over30 = {}
        EA_evoluted = {}
        l = count

        for index in range(0, 8):
            peak_30 = []
            mean_30 = []
            median_30 = []
            eighty_value_30 = []

            peak_male = []
            mean_male = []
            median_male = []
            eighty_value_male = []

            peak_female = []
            mean_female = []
            median_female = []
            eighty_value_female = []

            peak_EA_over30 = []
            mean_EA_over30 = []
            median_EA_over30 = []
            eighty_value_EA_over30 = []

            peak_EA_evoluted = []
            mean_EA_evoluted = []
            median_EA_evoluted = []
            eighty_value_EA_evoluted = []

            for i in wb_over30.sheetnames:
                if i != 'Sheet1' and i != 'Statistical Analysis_20_Female' and i!= 'Statistical Analysis_20_Male'\
                        and i != 'Statistical Analysis over 30' and i != 'Folha1':

                    ws_over30 = wb_over30.get_sheet_by_name(i)

                    peak_30.append(ws_over30['B' + str(l)].value)
                    mean_30.append(ws_over30['C' + str(l)].value)
                    median_30.append(ws_over30['D' + str(l)].value)
                    eighty_value_30.append(ws_over30['E' + str(l)].value)

            for i in wb_20.sheetnames:
                if i != 'Sheet1' and i != 'Statistical Analysis_20_Female' and i != 'Statistical Analysis_20_Male' \
                        and i != 'Statistical Analysis over 30' and i != 'Folha1' and i != 'Patient5_Healthy' \
                        and i != 'Patient8_Healthy' and i != 'Patient10_Healthy' and i != 'Patient13_Healthy' \
                        and i != 'Patient17_Healthy' and i != 'Patient18_Healthy' and i != 'Patient23_Healthy' \
                        and i != 'Patient25_Healthy' and i != 'Patient27_Healthy' and i != 'Patient30_Healthy' \
                        and i != 'Patient37_Healthy' and i != 'Patient38_Healthy':
                    ws_20 = wb_20.get_sheet_by_name(i)

                    peak_female.append(ws_20['B' + str(l)].value)
                    mean_female.append(ws_20['C' + str(l)].value)
                    median_female.append(ws_20['D' + str(l)].value)
                    eighty_value_female.append(ws_20['E' + str(l)].value)

                if i == 'Patient8_Healthy' or i == 'Patient10_Healthy' or i == 'Patient13_Healthy' \
                        or i == 'Patient17_Healthy' or i == 'Patient18_Healthy' or i == 'Patient23_Healthy' \
                        or i == 'Patient25_Healthy' or i == 'Patient27_Healthy' or i == 'Patient30_Healthy' \
                        or i == 'Patient5_Healthy' or i == 'Patient37_Healthy' or i == 'Patient38_Healthy':

                    ws_20 = wb_20.get_sheet_by_name(i)

                    peak_male.append(ws_20['B' + str(l)].value)
                    mean_male.append(ws_20['C' + str(l)].value)
                    median_male.append(ws_20['D' + str(l)].value)
                    eighty_value_male.append(ws_20['E' + str(l)].value)

            for i in wb_EA.sheetnames:
                if i != 'Sheet1' and i != 'Patient4_EA' and i != 'Patient8_EA' and i !='Patient11_EA' \
                        and i != "Statistical Analysis - EA evol" and i != "Statistical Analysis - EA over3":
                    ws_EA_over30 = wb_EA.get_sheet_by_name(i)

                    peak_EA_over30.append(ws_EA_over30['B' + str(l)].value)
                    mean_EA_over30.append(ws_EA_over30['C' + str(l)].value)
                    median_EA_over30.append(ws_EA_over30['D' + str(l)].value)
                    eighty_value_EA_over30.append(ws_EA_over30['E' + str(l)].value)

                if i == 'Patient4_EA' or i == 'Patient8_EA' or i =='Patient11_EA':
                    ws_EA_evoluted = wb_EA.get_sheet_by_name(i)
                    peak_EA_evoluted.append(ws_EA_evoluted['B' + str(l)].value)
                    mean_EA_evoluted.append(ws_EA_evoluted['C' + str(l)].value)
                    median_EA_evoluted.append(ws_EA_evoluted['D' + str(l)].value)
                    eighty_value_EA_evoluted.append(ws_EA_evoluted['E' + str(l)].value)

            if index == 0:
                male["Rectus_L"] = {"Peak": peak_male, "Mean": mean_male, "Median": median_male,
                                          "80_freq": eighty_value_male}
                female["Rectus_L"] = {"Peak": peak_female, "Mean": mean_female, "Median": median_female,
                                          "80_freq": eighty_value_female}
                over30["Rectus_L"] = {"Peak": peak_30, "Mean": mean_30, "Median": median_30,
                                          "80_freq": eighty_value_30}
                EA_over30["Rectus_L"] = {"Peak": peak_EA_over30, "Mean": mean_EA_over30, "Median": median_EA_over30,
                                      "80_freq": eighty_value_EA_over30}
                EA_evoluted["Rectus_L"] = {"Peak": peak_EA_evoluted, "Mean": mean_EA_evoluted, "Median": median_EA_evoluted,
                                  "80_freq": eighty_value_EA_evoluted}
            if index == 1:
                male["Rectus_R"] = {"Peak": peak_male, "Mean": mean_male, "Median": median_male,
                                          "80_freq": eighty_value_male}
                female["Rectus_R"] = {"Peak": peak_female, "Mean": mean_female, "Median": median_female,
                                          "80_freq": eighty_value_female}
                over30["Rectus_R"] = {"Peak": peak_30, "Mean": mean_30, "Median": median_30,
                                          "80_freq": eighty_value_30}
                EA_over30["Rectus_R"] = {"Peak": peak_EA_over30, "Mean": mean_EA_over30, "Median": median_EA_over30,
                                      "80_freq": eighty_value_EA_over30}
                EA_evoluted["Rectus_R"] = {"Peak": peak_EA_evoluted, "Mean": mean_EA_evoluted, "Median": median_EA_evoluted,
                                  "80_freq": eighty_value_EA_evoluted}
            if index == 2:
                male["Obliques_L"] = {"Peak": peak_male, "Mean": mean_male, "Median": median_male,
                                            "80_freq": eighty_value_male}
                female["Obliques_L"] = {"Peak": peak_female, "Mean": mean_female, "Median": median_female,
                                            "80_freq": eighty_value_female}
                over30["Obliques_L"] = {"Peak": peak_30, "Mean": mean_30, "Median": median_30,
                                            "80_freq": eighty_value_30}
                EA_over30["Obliques_L"] = {"Peak": peak_EA_over30, "Mean": mean_EA_over30, "Median": median_EA_over30,
                                        "80_freq": eighty_value_EA_over30}
                EA_evoluted["Obliques_L"] = {"Peak": peak_EA_evoluted, "Mean": mean_EA_evoluted, "Median": median_EA_evoluted,
                                    "80_freq": eighty_value_EA_evoluted}
            if index == 3:
                male["Obliques_R"] = {"Peak": peak_male, "Mean": mean_male, "Median": median_male,
                                            "80_freq": eighty_value_male}
                female["Obliques_R"] = {"Peak": peak_female, "Mean": mean_female, "Median": median_female,
                                          "80_freq": eighty_value_female}
                over30["Obliques_R"] = {"Peak": peak_30, "Mean": mean_30, "Median": median_30,
                                            "80_freq": eighty_value_30}
                EA_over30["Obliques_R"] = {"Peak": peak_EA_over30, "Mean": mean_EA_over30, "Median": median_EA_over30,
                                           "80_freq": eighty_value_EA_over30}
                EA_evoluted["Obliques_R"] = {"Peak": peak_EA_evoluted, "Mean": mean_EA_evoluted,
                                             "Median": median_EA_evoluted,
                                             "80_freq": eighty_value_EA_evoluted}
            if index == 4:
                male["Ilicostalis_L"] = {"Peak": peak_male, "Mean": mean_male, "Median": median_male,
                                               "80_freq": eighty_value_male}
                female["Ilicostalis_L"] = {"Peak": peak_female, "Mean": mean_female, "Median": median_female,
                                             "80_freq": eighty_value_female}
                over30["Ilicostalis_L"] = {"Peak": peak_30, "Mean": mean_30, "Median": median_30,
                                               "80_freq": eighty_value_30}
                EA_over30["Ilicostalis_L"] = {"Peak": peak_EA_over30, "Mean": mean_EA_over30, "Median": median_EA_over30,
                                           "80_freq": eighty_value_EA_over30}
                EA_evoluted["Ilicostalis_L"] = {"Peak": peak_EA_evoluted, "Mean": mean_EA_evoluted,
                                             "Median": median_EA_evoluted,
                                             "80_freq": eighty_value_EA_evoluted}
            if index == 5:
                male["Ilicostalis_R"] = {"Peak": peak_male, "Mean": mean_male, "Median": median_male,
                                               "80_freq": eighty_value_male}
                female["Ilicostalis_R"] = {"Peak": peak_female, "Mean": mean_female, "Median": median_female,
                                             "80_freq": eighty_value_female}
                over30["Ilicostalis_R"] = {"Peak": peak_30, "Mean": mean_30, "Median": median_30,
                                               "80_freq": eighty_value_30}
                EA_over30["Ilicostalis_R"] = {"Peak": peak_EA_over30, "Mean": mean_EA_over30,
                                              "Median": median_EA_over30,
                                              "80_freq": eighty_value_EA_over30}
                EA_evoluted["Ilicostalis_R"] = {"Peak": peak_EA_evoluted, "Mean": mean_EA_evoluted,
                                                "Median": median_EA_evoluted,
                                                "80_freq": eighty_value_EA_evoluted}
            if index == 6:
                male["Multi_L"] = {"Peak": peak_male, "Mean": mean_male, "Median": median_male,
                                         "80_freq": eighty_value_male}
                female["Multi_L"] = {"Peak": peak_female, "Mean": mean_female, "Median": median_female,
                                       "80_freq": eighty_value_female}
                over30["Multi_L"] = {"Peak": peak_30, "Mean": mean_30, "Median": median_30,
                                         "80_freq": eighty_value_30}
                EA_over30["Multi_L"] = {"Peak": peak_EA_over30, "Mean": mean_EA_over30,
                                              "Median": median_EA_over30,
                                              "80_freq": eighty_value_EA_over30}
                EA_evoluted["Multi_L"] = {"Peak": peak_EA_evoluted, "Mean": mean_EA_evoluted,
                                                "Median": median_EA_evoluted,
                                                "80_freq": eighty_value_EA_evoluted}
            if index == 7:
                male["Multi_R"] = {"Peak": peak_male, "Mean": mean_male, "Median": median_male,
                                         "80_freq": eighty_value_male}
                female["Multi_R"] = {"Peak": peak_female, "Mean": mean_female, "Median": median_female,
                                       "80_freq": eighty_value_female}
                over30["Multi_R"] = {"Peak": peak_30, "Mean": mean_30, "Median": median_30,
                                         "80_freq": eighty_value_30}
                EA_over30["Multi_R"] = {"Peak": peak_EA_over30, "Mean": mean_EA_over30,
                                        "Median": median_EA_over30,
                                        "80_freq": eighty_value_EA_over30}
                EA_evoluted["Multi_R"] = {"Peak": peak_EA_evoluted, "Mean": mean_EA_evoluted,
                                          "Median": median_EA_evoluted,
                                          "80_freq": eighty_value_EA_evoluted}

            l = l + 1

        value_final_over30[task] = over30
        value_final_male[task] = male
        value_final_female[task] = female
        value_final_EA_over30[task] = EA_over30
        value_final_EA_evoluted[task] = EA_evoluted
        count = count + 12

    return value_final_over30, value_final_male, value_final_female, value_final_EA_over30, value_final_EA_evoluted


def get_values_COP(mean_velocity):

    wb_over30 = load_workbook(file_excel_over30)
    wb_20 = load_workbook(file_excel_20)
    wb_EA = load_workbook(file_excel_EA)

    count = 8

    cop_parameters_over30 = {}
    cop_parameters_male = {}
    cop_parameters_female = {}
    cop_parameters_EA_over30 = {}
    cop_parameters_EA_evoluted = {}

    for task in mean_velocity:

        mean_v_x_over30 = []
        mean_v_y_over30 = []
        std_x_over30 = []
        std_y_over30 = []
        amp_x_over30 = []
        amp_y_over30 = []
        area_value_over30 = []

        mean_v_x_male = []
        mean_v_y_male = []
        std_x_male = []
        std_y_male = []
        amp_x_male = []
        amp_y_male = []
        area_value_male = []

        mean_v_x_female = []
        mean_v_y_female = []
        std_x_female = []
        std_y_female = []
        amp_x_female = []
        amp_y_female = []
        area_value_female = []

        mean_v_x_EA_over30 = []
        mean_v_y_EA_over30 = []
        std_x_EA_over30 = []
        std_y_EA_over30 = []
        amp_x_EA_over30= []
        amp_y_EA_over30 = []
        area_value_EA_over30 = []

        mean_v_x_EA_evaluted = []
        mean_v_y_EA_evaluted = []
        std_x_EA_evaluted = []
        std_y_EA_evaluted = []
        amp_x_EA_evaluted = []
        amp_y_EA_evaluted = []
        area_value_EA_evaluted = []

        for i in wb_over30.sheetnames:
            if i != 'Sheet1' and i != 'Statistical Analysis_20_Female' and i != 'Statistical Analysis_20_Male' \
                    and i != 'Statistical Analysis over 30' and i != 'Folha1':
                ws_over30 = wb_over30.get_sheet_by_name(i)

                mean_v_x_over30.append(ws_over30['Q' + str(count)].value)
                mean_v_y_over30.append(ws_over30['R' + str(count)].value)
                std_x_over30.append(ws_over30['Q' + str(count + 1)].value)
                std_y_over30.append(ws_over30['R' + str(count + 1)].value)
                amp_x_over30.append(ws_over30['Q' + str(count + 2)].value)
                amp_y_over30.append(ws_over30['R' + str(count + 2)].value)
                area_value_over30.append(ws_over30['Q' + str(count + 5)].value)

        for i in wb_20.sheetnames:
            if i != 'Sheet1' and i != 'Statistical Analysis_20_Female' and i != 'Statistical Analysis_20_Male' \
                    and i != 'Statistical Analysis over 30' and i != 'Folha1' and i != 'Patient5_Healthy' \
                    and i != 'Patient8_Healthy' and i != 'Patient10_Healthy' and i != 'Patient13_Healthy' \
                    and i != 'Patient17_Healthy' and i != 'Patient18_Healthy' and i != 'Patient23_Healthy' \
                    and i != 'Patient25_Healthy' and i != 'Patient27_Healthy' and i != 'Patient30_Healthy' \
                    and i != 'Patient37_Healthy' and i != 'Patient38_Healthy':
                ws_20 = wb_20.get_sheet_by_name(i)

                mean_v_x_female.append(ws_20['Q' + str(count)].value)
                mean_v_y_female.append(ws_20['R' + str(count)].value)
                std_x_female.append(ws_20['Q' + str(count + 1)].value)
                std_y_female.append(ws_20['R' + str(count + 1)].value)
                amp_x_female.append(ws_20['Q' + str(count + 2)].value)
                amp_y_female.append(ws_20['R' + str(count + 2)].value)
                area_value_female.append(ws_20['Q' + str(count + 5)].value)

            if i == 'Patient8_Healthy' or i == 'Patient10_Healthy' or i == 'Patient13_Healthy' \
                    or i == 'Patient17_Healthy' or i == 'Patient18_Healthy' or i == 'Patient23_Healthy' \
                    or i == 'Patient25_Healthy' or i == 'Patient27_Healthy' or i == 'Patient30_Healthy' \
                    or i == 'Patient5_Healthy' or i == 'Patient37_Healthy' or i == 'Patient38_Healthy':
                ws_20 = wb_20.get_sheet_by_name(i)

                mean_v_x_male.append(ws_20['Q' + str(count)].value)
                mean_v_y_male.append(ws_20['R' + str(count)].value)
                std_x_male.append(ws_20['Q' + str(count + 1)].value)
                std_y_male.append(ws_20['R' + str(count + 1)].value)
                amp_x_male.append(ws_20['Q' + str(count + 2)].value)
                amp_y_male.append(ws_20['R' + str(count + 2)].value)
                area_value_male.append(ws_20['Q' + str(count + 5)].value)

        for i in wb_EA.sheetnames:
            if i != 'Sheet1' and i != 'Patient4_EA' and i != 'Patient8_EA'  and i != 'Patient11_EA'\
                    and i != "Statistical Analysis - EA evol" and i != "Statistical Analysis - EA over3":
                ws_EA_over30 = wb_EA.get_sheet_by_name(i)
                mean_v_x_EA_over30.append(ws_EA_over30['Q' + str(count)].value)
                mean_v_y_EA_over30.append(ws_EA_over30['R' + str(count)].value)
                std_x_EA_over30.append(ws_EA_over30['Q' + str(count + 1)].value)
                std_y_EA_over30.append(ws_EA_over30['R' + str(count + 1)].value)
                amp_x_EA_over30.append(ws_EA_over30['Q' + str(count + 2)].value)
                amp_y_EA_over30.append(ws_EA_over30['R' + str(count + 2)].value)
                area_value_EA_over30.append(ws_EA_over30['Q' + str(count + 5)].value)

            if i == 'Patient4_EA' or i == 'Patient8_EA' or i == 'Patient11_EA':
                ws_EA_evoluted = wb_EA.get_sheet_by_name(i)
                mean_v_x_EA_evaluted.append(ws_EA_evoluted['Q' + str(count)].value)
                mean_v_y_EA_evaluted.append(ws_EA_evoluted['R' + str(count)].value)
                std_x_EA_evaluted.append(ws_EA_evoluted['Q' + str(count + 1)].value)
                std_y_EA_evaluted.append(ws_EA_evoluted['R' + str(count + 1)].value)
                amp_x_EA_evaluted.append(ws_EA_evoluted['Q' + str(count + 2)].value)
                amp_y_EA_evaluted.append(ws_EA_evoluted['R' + str(count + 2)].value)
                area_value_EA_evaluted.append(ws_EA_evoluted['Q' + str(count + 5)].value)


        cop_parameters_over30[task] = {"Velocity_X": mean_v_x_over30, "Velocity_Y": mean_v_y_over30,
                                 "STD_X": std_x_over30, "STD_Y": std_y_over30,
                                 "Amp_X": amp_x_over30, "Amp_Y": amp_y_over30,
                                 "Area": area_value_over30}

        cop_parameters_male[task] = {"Velocity_X": mean_v_x_male, "Velocity_Y": mean_v_y_male,
                                       "STD_X": std_x_male, "STD_Y": std_y_male,
                                       "Amp_X": amp_x_male, "Amp_Y": amp_y_male,
                                       "Area": area_value_male}

        cop_parameters_female[task] = {"Velocity_X": mean_v_x_female, "Velocity_Y": mean_v_y_female,
                                       "STD_X": std_x_female, "STD_Y": std_y_female,
                                       "Amp_X": amp_x_female, "Amp_Y": amp_y_female,
                                       "Area": area_value_female}

        cop_parameters_EA_over30[task] = {"Velocity_X": mean_v_x_EA_over30, "Velocity_Y": mean_v_y_EA_over30,
                                       "STD_X": std_x_EA_over30, "STD_Y": std_y_EA_over30,
                                       "Amp_X": amp_x_EA_over30, "Amp_Y": amp_y_EA_over30,
                                       "Area": area_value_EA_over30}

        cop_parameters_EA_evoluted[task] = {"Velocity_X": mean_v_x_EA_evaluted, "Velocity_Y": mean_v_y_EA_evaluted,
                                   "STD_X": std_x_EA_evaluted, "STD_Y": std_y_EA_evaluted,
                                   "Amp_X": amp_x_EA_evaluted, "Amp_Y": amp_y_EA_evaluted,
                                   "Area": area_value_EA_evaluted}

        count = count + 13

    return cop_parameters_over30, cop_parameters_male, cop_parameters_female, cop_parameters_EA_over30, cop_parameters_EA_evoluted


def get_values_COP_freq(COP_freq):

    wb_over30 = load_workbook(file_excel_over30)
    wb_20 = load_workbook(file_excel_20)
    wb_EA = load_workbook(file_excel_EA)

    count = 260

    cop_freqs_over30 = {}
    cop_freqs_male = {}
    cop_freqs_female = {}
    cop_freqs_EA_over30 = {}
    cop_freqs_EA_evoluted = {}

    for task in COP_freq:

        peak_over30_copx = []
        mean_over30_copx = []
        median_over30_copx = []
        eighty_over30_copx = []

        peak_over30_copy = []
        mean_over30_copy = []
        median_over30_copy = []
        eighty_over30_copy = []

        peak_male_copx = []
        mean_male_copx = []
        median_male_copx = []
        eighty_male_copx = []

        peak_male_copy = []
        mean_male_copy = []
        median_male_copy = []
        eighty_male_copy = []

        peak_female_copx = []
        mean_female_copx = []
        median_female_copx = []
        eighty_female_copx = []

        peak_female_copy = []
        mean_female_copy = []
        median_female_copy = []
        eighty_female_copy = []

        peak_EA_copx_over30 = []
        mean_EA_copx_over30= []
        median_EA_copx_over30 = []
        eighty_EA_copx_over30 = []

        peak_EA_copy_over30= []
        mean_EA_copy_over30 = []
        median_EA_copy_over30 = []
        eighty_EA_copy_over30 = []

        peak_EA_copx_evaluted = []
        mean_EA_copx_evaluted = []
        median_EA_copx_evaluted = []
        eighty_EA_copx_evaluted = []

        peak_EA_copy_evaluted = []
        mean_EA_copy_evaluted = []
        median_EA_copy_evaluted = []
        eighty_EA_copy_evaluted = []



        for i in wb_over30.sheetnames:
            if i != 'Sheet1' and i != 'Statistical Analysis_20_Female' and i != 'Statistical Analysis_20_Male' \
                    and i != 'Statistical Analysis over 30' and i != 'Folha1':
                ws_over30 = wb_over30.get_sheet_by_name(i)

                peak_over30_copx.append(ws_over30['K' + str(count)].value)
                peak_over30_copy.append(ws_over30['K' + str(count + 1)].value)
                mean_over30_copx.append(ws_over30['L' + str(count)].value)
                mean_over30_copy.append(ws_over30['L' + str(count + 1)].value)
                median_over30_copx.append(ws_over30['M' + str(count)].value)
                median_over30_copy.append(ws_over30['M' + str(count + 1)].value)
                eighty_over30_copx.append(ws_over30['N' + str(count)].value)
                eighty_over30_copy.append(ws_over30['N' + str(count + 1)].value)

        for i in wb_20.sheetnames:
            if i != 'Sheet1' and i != 'Statistical Analysis_20_Female' and i != 'Statistical Analysis_20_Male' \
                    and i != 'Statistical Analysis over 30' and i != 'Folha1' and i != 'Patient5_Healthy' \
                    and i != 'Patient8_Healthy' and i != 'Patient10_Healthy' and i != 'Patient13_Healthy' \
                    and i != 'Patient17_Healthy' and i != 'Patient18_Healthy' and i != 'Patient23_Healthy' \
                    and i != 'Patient25_Healthy' and i != 'Patient27_Healthy' and i != 'Patient30_Healthy' \
                    and i != 'Patient37_Healthy' and i != 'Patient38_Healthy':
                ws_20 = wb_20.get_sheet_by_name(i)

                peak_female_copx.append(ws_20['K' + str(count)].value)
                peak_female_copy.append(ws_20['K' + str(count + 1)].value)
                mean_female_copx.append(ws_20['L' + str(count)].value)
                mean_female_copy.append(ws_20['L' + str(count + 1)].value)
                median_female_copx.append(ws_20['M' + str(count)].value)
                median_female_copy.append(ws_20['M' + str(count + 1)].value)
                eighty_female_copx.append(ws_20['N' + str(count)].value)
                eighty_female_copy.append(ws_20['N' + str(count + 1)].value)

            if i == 'Patient8_Healthy' or i == 'Patient10_Healthy' or i == 'Patient13_Healthy' \
                    or i == 'Patient17_Healthy' or i == 'Patient18_Healthy' or i == 'Patient23_Healthy' \
                    or i == 'Patient25_Healthy' or i == 'Patient27_Healthy' or i == 'Patient30_Healthy' \
                    or i == 'Patient5_Healthy' or i == 'Patient37_Healthy' or i == 'Patient38_Healthy':
                ws_20 = wb_20.get_sheet_by_name(i)

                peak_male_copx.append(ws_20['K' + str(count)].value)
                peak_male_copy.append(ws_20['K' + str(count + 1)].value)
                mean_male_copx.append(ws_20['L' + str(count)].value)
                mean_male_copy.append(ws_20['L' + str(count + 1)].value)
                median_male_copx.append(ws_20['M' + str(count)].value)
                median_male_copy.append(ws_20['M' + str(count + 1)].value)
                eighty_male_copx.append(ws_20['N' + str(count)].value)
                eighty_male_copy.append(ws_20['N' + str(count + 1)].value)

        for i in wb_EA.sheetnames:
            if i != 'Sheet1' and i != 'Patient4_EA' and i != 'Patient8_EA' and i != 'Patient11_EA' \
                    and i != "Statistical Analysis - EA evol" and i != "Statistical Analysis - EA over3":
                ws_EA_over30 = wb_EA.get_sheet_by_name(i)

                peak_EA_copx_over30.append(ws_EA_over30['K' + str(count)].value)
                peak_EA_copy_over30.append(ws_EA_over30['K' + str(count + 1)].value)
                mean_EA_copx_over30.append(ws_EA_over30['L' + str(count)].value)
                mean_EA_copy_over30.append(ws_EA_over30['L' + str(count + 1)].value)
                median_EA_copx_over30.append(ws_EA_over30['M' + str(count)].value)
                median_EA_copy_over30.append(ws_EA_over30['M' + str(count + 1)].value)
                eighty_EA_copx_over30.append(ws_EA_over30['N' + str(count)].value)
                eighty_EA_copy_over30.append(ws_EA_over30['N' + str(count + 1)].value)

            if i == 'Patient4_EA' or i == 'Patient8_EA' or i == 'Patient11_EA':
                ws_EA_evoluted = wb_EA.get_sheet_by_name(i)
                peak_EA_copx_evaluted.append(ws_EA_evoluted['K' + str(count)].value)
                peak_EA_copy_evaluted.append(ws_EA_evoluted['K' + str(count + 1)].value)
                mean_EA_copx_evaluted.append(ws_EA_evoluted['L' + str(count)].value)
                mean_EA_copy_evaluted.append(ws_EA_evoluted['L' + str(count + 1)].value)
                median_EA_copx_evaluted.append(ws_EA_evoluted['M' + str(count)].value)
                median_EA_copy_evaluted.append(ws_EA_evoluted['M' + str(count + 1)].value)
                eighty_EA_copx_evaluted.append(ws_EA_evoluted['N' + str(count)].value)
                eighty_EA_copy_evaluted.append(ws_EA_evoluted['N' + str(count + 1)].value)


        cop_freqs_over30[task] = {"Peak_X": peak_over30_copx, "Peak_Y": peak_over30_copy,
                                 "Mean_X": mean_over30_copx, "Mean_Y": mean_female_copy,
                                 "Median_X": median_over30_copx, "Median_Y": median_over30_copy,
                                 "80_X": eighty_over30_copx, "80_Y": eighty_over30_copy}

        cop_freqs_male[task] = {"Peak_X": peak_male_copx, "Peak_Y": peak_male_copy,
                                 "Mean_X": mean_male_copx, "Mean_Y": mean_male_copy,
                                 "Median_X": median_male_copx, "Median_Y": median_male_copy,
                                 "80_X": eighty_male_copx, "80_Y": eighty_male_copy}

        cop_freqs_female[task] = {"Peak_X": peak_female_copx, "Peak_Y": peak_female_copy,
                                 "Mean_X": mean_female_copx, "Mean_Y": mean_female_copy,
                                 "Median_X": median_female_copx, "Median_Y": median_female_copy,
                                 "80_X": eighty_female_copx, "80_Y": eighty_female_copy}

        cop_freqs_EA_over30[task] = {"Peak_X": peak_EA_copx_over30, "Peak_Y": peak_EA_copy_over30,
                                 "Mean_X": mean_EA_copx_over30, "Mean_Y": mean_EA_copy_over30,
                                 "Median_X": median_EA_copx_over30, "Median_Y": median_EA_copy_over30,
                                 "80_X": eighty_EA_copx_over30, "80_Y": eighty_EA_copy_over30}

        cop_freqs_EA_evoluted[task] = {"Peak_X": peak_EA_copx_evaluted, "Peak_Y": peak_EA_copy_evaluted,
                              "Mean_X": mean_EA_copx_evaluted, "Mean_Y": mean_EA_copy_evaluted,
                              "Median_X": median_EA_copx_evaluted, "Median_Y": median_EA_copy_evaluted,
                              "80_X": eighty_EA_copx_evaluted, "80_Y": eighty_EA_copy_evaluted}

        count = count + 12

    return cop_freqs_over30, cop_freqs_male, cop_freqs_female, cop_freqs_EA_over30, cop_freqs_EA_evoluted



def get_value_freq_rest():

    wb_over30 = load_workbook(file_excel_over30)
    wb_20 = load_workbook(file_excel_20)
    wb_EA = load_workbook(file_excel_EA)

    count = 392
    over30 = {}
    male = {}
    female = {}
    EA_over30 = {}
    EA_evaluted = {}


    for index in range(0, 8):
        peak_30 = []
        mean_30 = []
        median_30 = []
        eighty_value_30 = []

        peak_male = []
        mean_male = []
        median_male = []
        eighty_value_male = []

        peak_female = []
        mean_female = []
        median_female = []
        eighty_value_female = []

        peak_EA_over30 = []
        mean_EA_over30 = []
        median_EA_over30 = []
        eighty_value_EA_over30 = []

        peak_EA_evaluted = []
        mean_EA_evaluted = []
        median_EA_evaluted = []
        eighty_value_EA_evaluted = []

        for i in wb_over30.sheetnames:
            if i != 'Sheet1' and i != 'Statistical Analysis_20_Female' and i!= 'Statistical Analysis_20_Male'\
                        and i != 'Statistical Analysis over 30' and i != 'Folha1':

                ws_over30 = wb_over30.get_sheet_by_name(i)
                if ws_over30['B' + str(count)].value != 'None':

                    peak_30.append(ws_over30['B' + str(count)].value)
                    mean_30.append(ws_over30['C' + str(count)].value)
                    median_30.append(ws_over30['D' + str(count)].value)
                    eighty_value_30.append(ws_over30['E' + str(count)].value)

        for i in wb_20.sheetnames:
            if i != 'Sheet1' and i != 'Statistical Analysis_20_Female' and i != 'Statistical Analysis_20_Male' \
                    and i != 'Statistical Analysis over 30' and i != 'Folha1' and i != 'Patient5_Healthy' \
                    and i != 'Patient8_Healthy' and i != 'Patient10_Healthy' and i != 'Patient13_Healthy' \
                    and i != 'Patient17_Healthy' and i != 'Patient18_Healthy' and i != 'Patient23_Healthy' \
                    and i != 'Patient25_Healthy' and i != 'Patient27_Healthy' and i != 'Patient30_Healthy' \
                    and i != 'Patient37_Healthy' and i != 'Patient38_Healthy':
                ws_20 = wb_20.get_sheet_by_name(i)
                if ws_20['B' + str(count)].value != None:

                    peak_female.append(ws_20['B' + str(count)].value)
                    mean_female.append(ws_20['C' + str(count)].value)
                    median_female.append(ws_20['D' + str(count)].value)
                    eighty_value_female.append(ws_20['E' + str(count)].value)

            if i == 'Patient8_Healthy' or i == 'Patient10_Healthy' or i == 'Patient13_Healthy' \
                    or i == 'Patient17_Healthy' or i == 'Patient18_Healthy' or i == 'Patient23_Healthy' \
                    or i == 'Patient25_Healthy' or i == 'Patient27_Healthy' or i == 'Patient30_Healthy' \
                    or i == 'Patient5_Healthy' or i == 'Patient37_Healthy' or i == 'Patient38_Healthy':

                ws_20 = wb_20.get_sheet_by_name(i)
                if ws_20['B' + str(count)].value != None:

                    peak_male.append(ws_20['B' + str(count)].value)
                    mean_male.append(ws_20['C' + str(count)].value)
                    median_male.append(ws_20['D' + str(count)].value)
                    eighty_value_male.append(ws_20['E' + str(count)].value)

        for i in wb_EA.sheetnames:
            if i != 'Sheet1' and i != 'Patient4_EA' and i != 'Patient8_EA'  and i != 'Patient11_EA'\
                    and i != "Statistical Analysis - EA evol" and i != "Statistical Analysis - EA over3":
                ws_EA_over30 = wb_EA.get_sheet_by_name(i)
                if ws_EA_over30['B' + str(count)].value != None:

                    peak_EA_over30.append(ws_EA_over30['B' + str(count)].value)
                    mean_EA_over30.append(ws_EA_over30['C' + str(count)].value)
                    median_EA_over30.append(ws_EA_over30['D' + str(count)].value)
                    eighty_value_EA_over30.append(ws_EA_over30['E' + str(count)].value)

            if i == 'Patient4_EA' or i == 'Patient8_EA' or i == 'Patient11_EA':
                    ws_EA_evaluted = wb_EA.get_sheet_by_name(i)
                    if ws_EA_evaluted['B' + str(count)].value != None:
                        peak_EA_evaluted.append(ws_EA_evaluted['B' + str(count)].value)
                        mean_EA_evaluted.append(ws_EA_evaluted['C' + str(count)].value)
                        median_EA_evaluted.append(ws_EA_evaluted['D' + str(count)].value)
                        eighty_value_EA_evaluted.append(ws_EA_evaluted['E' + str(count)].value)

        if index == 0:
            male["Rectus_L"] = {"Peak": peak_male, "Mean": mean_male, "Median": median_male,
                                          "80_freq": eighty_value_male}
            female["Rectus_L"] = {"Peak": peak_female, "Mean": mean_female, "Median": median_female,
                                          "80_freq": eighty_value_female}
            over30["Rectus_L"] = {"Peak": peak_30, "Mean": mean_30, "Median": median_30,
                                          "80_freq": eighty_value_30}
            EA_over30["Rectus_L"] = {"Peak": peak_EA_over30, "Mean": mean_EA_over30, "Median": median_EA_over30,
                                      "80_freq": eighty_value_EA_over30}
            EA_evaluted["Rectus_L"] = {"Peak": peak_EA_evaluted, "Mean": mean_EA_evaluted, "Median": median_EA_evaluted,
                              "80_freq": eighty_value_EA_evaluted}
        if index == 1:
            male["Rectus_R"] = {"Peak": peak_male, "Mean": mean_male, "Median": median_male,
                                          "80_freq": eighty_value_male}
            female["Rectus_R"] = {"Peak": peak_female, "Mean": mean_female, "Median": median_female,
                                          "80_freq": eighty_value_female}
            over30["Rectus_R"] = {"Peak": peak_30, "Mean": mean_30, "Median": median_30,
                                          "80_freq": eighty_value_30}
            EA_over30["Rectus_R"] = {"Peak": peak_EA_over30, "Mean": mean_EA_over30, "Median": median_EA_over30,
                                     "80_freq": eighty_value_EA_over30}
            EA_evaluted["Rectus_R"] = {"Peak": peak_EA_evaluted, "Mean": mean_EA_evaluted, "Median": median_EA_evaluted,
                                       "80_freq": eighty_value_EA_evaluted}
        if index == 2:
            male["Obliques_L"] = {"Peak": peak_male, "Mean": mean_male, "Median": median_male,
                                            "80_freq": eighty_value_male}
            female["Obliques_L"] = {"Peak": peak_female, "Mean": mean_female, "Median": median_female,
                                            "80_freq": eighty_value_female}
            over30["Obliques_L"] = {"Peak": peak_30, "Mean": mean_30, "Median": median_30,
                                            "80_freq": eighty_value_30}
            EA_over30["Obliques_L"] = {"Peak": peak_EA_over30, "Mean": mean_EA_over30, "Median": median_EA_over30,
                                     "80_freq": eighty_value_EA_over30}
            EA_evaluted["Obliques_L"] = {"Peak": peak_EA_evaluted, "Mean": mean_EA_evaluted, "Median": median_EA_evaluted,
                                       "80_freq": eighty_value_EA_evaluted}
        if index == 3:
            male["Obliques_R"] = {"Peak": peak_male, "Mean": mean_male, "Median": median_male,
                                            "80_freq": eighty_value_male}
            female["Obliques_R"] = {"Peak": peak_female, "Mean": mean_female, "Median": median_female,
                                          "80_freq": eighty_value_female}
            over30["Obliques_R"] = {"Peak": peak_30, "Mean": mean_30, "Median": median_30,
                                            "80_freq": eighty_value_30}
            EA_over30["Obliques_R"] = {"Peak": peak_EA_over30, "Mean": mean_EA_over30, "Median": median_EA_over30,
                                       "80_freq": eighty_value_EA_over30}
            EA_evaluted["Obliques_R"] = {"Peak": peak_EA_evaluted, "Mean": mean_EA_evaluted,
                                         "Median": median_EA_evaluted,
                                         "80_freq": eighty_value_EA_evaluted}
        if index == 4:
            male["Ilicostalis_L"] = {"Peak": peak_male, "Mean": mean_male, "Median": median_male,
                                               "80_freq": eighty_value_male}
            female["Ilicostalis_L"] = {"Peak": peak_female, "Mean": mean_female, "Median": median_female,
                                             "80_freq": eighty_value_female}
            over30["Ilicostalis_L"] = {"Peak": peak_30, "Mean": mean_30, "Median": median_30,
                                               "80_freq": eighty_value_30}
            EA_over30["Ilicostalis_L"] = {"Peak": peak_EA_over30, "Mean": mean_EA_over30, "Median": median_EA_over30,
                                       "80_freq": eighty_value_EA_over30}
            EA_evaluted["Ilicostalis_L"] = {"Peak": peak_EA_evaluted, "Mean": mean_EA_evaluted,
                                         "Median": median_EA_evaluted,
                                         "80_freq": eighty_value_EA_evaluted}
        if index == 5:
            male["Ilicostalis_R"] = {"Peak": peak_male, "Mean": mean_male, "Median": median_male,
                                               "80_freq": eighty_value_male}
            female["Ilicostalis_R"] = {"Peak": peak_female, "Mean": mean_female, "Median": median_female,
                                             "80_freq": eighty_value_female}
            over30["Ilicostalis_R"] = {"Peak": peak_30, "Mean": mean_30, "Median": median_30,
                                               "80_freq": eighty_value_30}
            EA_over30["Ilicostalis_R"] = {"Peak": peak_EA_over30, "Mean": mean_EA_over30, "Median": median_EA_over30,
                                          "80_freq": eighty_value_EA_over30}
            EA_evaluted["Ilicostalis_R"] = {"Peak": peak_EA_evaluted, "Mean": mean_EA_evaluted,
                                            "Median": median_EA_evaluted,
                                            "80_freq": eighty_value_EA_evaluted}
        if index == 6:
            male["Multi_L"] = {"Peak": peak_male, "Mean": mean_male, "Median": median_male,
                                         "80_freq": eighty_value_male}
            female["Multi_L"] = {"Peak": peak_female, "Mean": mean_female, "Median": median_female,
                                       "80_freq": eighty_value_female}
            over30["Multi_L"] = {"Peak": peak_30, "Mean": mean_30, "Median": median_30,
                                         "80_freq": eighty_value_30}
            EA_over30["Multi_L"] = {"Peak": peak_EA_over30, "Mean": mean_EA_over30, "Median": median_EA_over30,
                                          "80_freq": eighty_value_EA_over30}
            EA_evaluted["Multi_L"] = {"Peak": peak_EA_evaluted, "Mean": mean_EA_evaluted,
                                            "Median": median_EA_evaluted,
                                            "80_freq": eighty_value_EA_evaluted}
        if index == 7:
            male["Multi_R"] = {"Peak": peak_male, "Mean": mean_male, "Median": median_male,
                                         "80_freq": eighty_value_male}
            female["Multi_R"] = {"Peak": peak_female, "Mean": mean_female, "Median": median_female,
                                       "80_freq": eighty_value_female}
            over30["Multi_R"] = {"Peak": peak_30, "Mean": mean_30, "Median": median_30,
                                         "80_freq": eighty_value_30}
            EA_over30["Multi_R"] = {"Peak": peak_EA_over30, "Mean": mean_EA_over30, "Median": median_EA_over30,
                                    "80_freq": eighty_value_EA_over30}
            EA_evaluted["Multi_R"] = {"Peak": peak_EA_evaluted, "Mean": mean_EA_evaluted,
                                      "Median": median_EA_evaluted,
                                      "80_freq": eighty_value_EA_evaluted}

        count = count + 1


    return over30, male, female, EA_over30, EA_evaluted


def get_value_tonus_rest():

    wb_over30 = load_workbook(file_excel_over30)
    wb_20 = load_workbook(file_excel_20)
    wb_EA = load_workbook(file_excel_EA)


    count = 146

    value_rectus_l_over30 = []
    value_rectus_r_over30 = []
    value_obliques_l_over30 = []
    value_obliques_r_over30 = []
    value_ilicostalis_l_over30 = []
    value_ilicostalis_r_over30 = []
    value_multi_l_over30 = []
    value_multi_r_over30 = []

    value_rectus_l_female = []
    value_rectus_r_female = []
    value_obliques_l_female = []
    value_obliques_r_female = []
    value_ilicostalis_l_female = []
    value_ilicostalis_r_female = []
    value_multi_l_female = []
    value_multi_r_female = []

    value_rectus_l_male = []
    value_rectus_r_male = []
    value_obliques_l_male = []
    value_obliques_r_male = []
    value_ilicostalis_l_male = []
    value_ilicostalis_r_male = []
    value_multi_l_male = []
    value_multi_r_male = []

    value_rectus_l_EA_over30 = []
    value_rectus_r_EA_over30 = []
    value_obliques_l_EA_over30 = []
    value_obliques_r_EA_over30 = []
    value_ilicostalis_l_EA_over30 = []
    value_ilicostalis_r_EA_over30 = []
    value_multi_l_EA_over30 = []
    value_multi_r_EA_over30 = []

    value_rectus_l_EA_evaluted = []
    value_rectus_r_EA_evaluted = []
    value_obliques_l_EA_evaluted = []
    value_obliques_r_EA_evaluted = []
    value_ilicostalis_l_EA_evaluted = []
    value_ilicostalis_r_EA_evaluted = []
    value_multi_l_EA_evaluted = []
    value_multi_r_EA_evaluted = []

    for i in wb_over30.sheetnames:
        if i != 'Sheet1' and i != 'Statistical Analysis_20_Female' and i != 'Statistical Analysis_20_Male' \
                and i != 'Statistical Analysis over 30' and i != 'Folha1':
            ws_over30 = wb_over30.get_sheet_by_name(i)
            value_rectus_l_over30.append(ws_over30['B' + str(count)].value)
            value_rectus_r_over30.append(ws_over30['C' + str(count)].value)
            value_obliques_l_over30.append(ws_over30['B' + str(count + 1)].value)
            value_obliques_r_over30.append(ws_over30['C' + str(count + 1)].value)
            value_ilicostalis_l_over30.append(ws_over30['B' + str(count + 2)].value)
            value_ilicostalis_r_over30.append(ws_over30['C' + str(count + 2)].value)
            value_multi_l_over30.append(ws_over30['B' + str(count + 3)].value)
            value_multi_r_over30.append(ws_over30['C' + str(count + 3)].value)

    for i in wb_20.sheetnames:
        if i != 'Sheet1' and i != 'Statistical Analysis_20_Female' and i != 'Statistical Analysis_20_Male' \
                and i != 'Statistical Analysis over 30' and i != 'Folha1' and i != 'Patient5_Healthy' \
                and i != 'Patient8_Healthy' and i != 'Patient10_Healthy' and i != 'Patient13_Healthy' \
                and i != 'Patient17_Healthy' and i != 'Patient18_Healthy' and i != 'Patient23_Healthy' \
                and i != 'Patient25_Healthy' and i != 'Patient27_Healthy' and i != 'Patient30_Healthy' \
                and i != 'Patient37_Healthy' and i != 'Patient38_Healthy':
            ws_female = wb_20.get_sheet_by_name(i)
            value_rectus_l_female.append(ws_female['B' + str(count)].value)
            value_rectus_r_female.append(ws_female['C' + str(count)].value)
            value_obliques_l_female.append(ws_female['B' + str(count + 1)].value)
            value_obliques_r_female.append(ws_female['C' + str(count + 1)].value)
            value_ilicostalis_l_female.append(ws_female['B' + str(count + 2)].value)
            value_ilicostalis_r_female.append(ws_female['C' + str(count + 2)].value)
            value_multi_l_female.append(ws_female['B' + str(count + 3)].value)
            value_multi_r_female.append(ws_female['C' + str(count + 3)].value)

        if i == 'Patient8_Healthy' or i == 'Patient10_Healthy' or i == 'Patient13_Healthy' \
                or i == 'Patient17_Healthy' or i == 'Patient18_Healthy' or i == 'Patient23_Healthy' \
                or i == 'Patient25_Healthy' or i == 'Patient27_Healthy' or i == 'Patient30_Healthy' \
                or i == 'Patient5_Healthy' or i == 'Patient37_Healthy' or i == 'Patient38_Healthy':
            ws_male = wb_20.get_sheet_by_name(i)
            value_rectus_l_male.append(ws_male['B' + str(count)].value)
            value_rectus_r_male.append(ws_male['C' + str(count)].value)
            value_obliques_l_male.append(ws_male['B' + str(count + 1)].value)
            value_obliques_r_male.append(ws_male['C' + str(count + 1)].value)
            value_ilicostalis_l_male.append(ws_male['B' + str(count + 2)].value)
            value_ilicostalis_r_male.append(ws_male['C' + str(count + 2)].value)
            value_multi_l_male.append(ws_male['B' + str(count + 3)].value)
            value_multi_r_male.append(ws_male['C' + str(count + 3)].value)

    for i in wb_EA.sheetnames:
        if i != 'Sheet1' and i != 'Patient4_EA' and i != 'Patient8_EA'  and i != 'Patient11_EA'\
                and i != "Statistical Analysis - EA evol" and i != "Statistical Analysis - EA over3":
            ws_EA_over30 = wb_EA.get_sheet_by_name(i)
            value_rectus_l_EA_over30.append(ws_EA_over30['B' + str(count)].value)
            value_rectus_r_EA_over30.append(ws_EA_over30['C' + str(count)].value)
            value_obliques_l_EA_over30.append(ws_EA_over30['B' + str(count + 1)].value)
            value_obliques_r_EA_over30.append(ws_EA_over30['C' + str(count + 1)].value)
            value_ilicostalis_l_EA_over30.append(ws_EA_over30['B' + str(count + 2)].value)
            value_ilicostalis_r_EA_over30.append(ws_EA_over30['C' + str(count + 2)].value)
            value_multi_l_EA_over30.append(ws_EA_over30['B' + str(count + 3)].value)
            value_multi_r_EA_over30.append(ws_EA_over30['C' + str(count + 3)].value)

        if i == 'Patient4_EA' or i == 'Patient8_EA' or i == 'Patient11_EA':
            ws_EA_evaluted = wb_EA.get_sheet_by_name(i)
            value_rectus_l_EA_evaluted.append(ws_EA_evaluted['B' + str(count)].value)
            value_rectus_r_EA_evaluted.append(ws_EA_evaluted['C' + str(count)].value)
            value_obliques_l_EA_evaluted.append(ws_EA_evaluted['B' + str(count + 1)].value)
            value_obliques_r_EA_evaluted.append(ws_EA_evaluted['C' + str(count + 1)].value)
            value_ilicostalis_l_EA_evaluted.append(ws_EA_evaluted['B' + str(count + 2)].value)
            value_ilicostalis_r_EA_evaluted.append(ws_EA_evaluted['C' + str(count + 2)].value)
            value_multi_l_EA_evaluted.append(ws_EA_evaluted['B' + str(count + 3)].value)
            value_multi_r_EA_evaluted.append(ws_EA_evaluted['C' + str(count + 3)].value)


    value_final_over30 = {"Rectus_L": value_rectus_l_over30, "Rectus_R": value_rectus_r_over30,
                                "Obliques_L": value_obliques_l_over30,
                                "Obliques_R": value_obliques_r_over30, "Ilicostalis_L": value_ilicostalis_l_over30,
                                "Ilicostalis_R": value_ilicostalis_r_over30,
                                "Multifidus_L": value_multi_l_over30, "Multifidus_R": value_multi_r_over30}

    value_final_male = {"Rectus_L": value_rectus_l_male, "Rectus_R": value_rectus_r_male,
                              "Obliques_L": value_obliques_l_male,
                              "Obliques_R": value_obliques_r_male, "Ilicostalis_L": value_ilicostalis_l_male,
                              "Ilicostalis_R": value_ilicostalis_r_male,
                              "Multifidus_L": value_multi_l_male, "Multifidus_R": value_multi_r_male}

    value_final_female = {"Rectus_L": value_rectus_l_female, "Rectus_R": value_rectus_r_female,
                                "Obliques_L": value_obliques_l_female,
                                "Obliques_R": value_obliques_r_female, "Ilicostalis_L": value_ilicostalis_l_female,
                                "Ilicostalis_R": value_ilicostalis_r_female,
                                "Multifidus_L": value_multi_l_female, "Multifidus_R": value_multi_r_female}

    value_final_EA_over30 = {"Rectus_L": value_rectus_l_EA_over30, "Rectus_R": value_rectus_r_EA_over30,
                            "Obliques_L": value_obliques_l_EA_over30,
                            "Obliques_R": value_obliques_r_EA_over30, "Ilicostalis_L": value_ilicostalis_l_EA_over30,
                            "Ilicostalis_R": value_ilicostalis_r_EA_over30,
                            "Multifidus_L": value_multi_l_EA_over30, "Multifidus_R": value_multi_r_EA_over30}

    value_final_EA_evaluted = {"Rectus_L": value_rectus_l_EA_evaluted, "Rectus_R": value_rectus_r_EA_evaluted,
                      "Obliques_L": value_obliques_l_EA_evaluted,
                      "Obliques_R": value_obliques_r_EA_evaluted, "Ilicostalis_L": value_ilicostalis_l_EA_evaluted,
                      "Ilicostalis_R": value_ilicostalis_r_EA_evaluted,
                      "Multifidus_L": value_multi_l_EA_evaluted, "Multifidus_R": value_multi_r_EA_evaluted}


    return value_final_over30, value_final_male, value_final_female, value_final_EA_over30, value_final_EA_evaluted



def EMG_values_boxplot(over30_tonus, male_tonus, female_tonus, EA_over_tonus, EA_more_tonus):
    l = 0
    pp = PdfPages('Tonus_Muscular_BoxPlot_Analysis_Healthy_EA.pdf')

    for i in over30_tonus:

        fig = plt.figure(l)
        l = l + 1
        fig.suptitle(str(i) + " - Max Value of normalized EMG", fontsize=21)

        plot1 = plt.subplot2grid((4, 2), (0, 0))
        data_rectus_l = [[male_tonus[i]["Rectus_L"]], [female_tonus[i]["Rectus_L"]],
                         [over30_tonus[i]["Rectus_L"]], [EA_over_tonus[i]["Rectus_L"]],
                         [EA_more_tonus[i]["Rectus_L"]]]
        plt.boxplot(data_rectus_l, positions = [1,2,3,4,5], widths = 0.6, showmeans= True)
        plt.xticks([1,2,3,4,5], ("Male", "Female", "Over 30", "EA over30", "EA more evolve"), fontsize=8)
        plt.ylim([0, 140])
        plot1.set_title("Rectus Left", fontsize=12)

        plot2 = plt.subplot2grid((4, 2), (0, 1))
        data_rectus_R = [[male_tonus[i]["Rectus_R"]], [female_tonus[i]["Rectus_R"]], [over30_tonus[i]["Rectus_R"]],
                         [EA_over_tonus[i]["Rectus_R"]], [EA_more_tonus[i]["Rectus_R"]]]
        plt.boxplot(data_rectus_R, positions = [1,2,3,4,5], widths=0.6, showmeans= True)
        plt.xticks([1, 2, 3, 4, 5], ("Male", "Female", "Over 30", "EA over30", "EA more evolve"),
                   fontsize=8)
        plt.ylim([0, 140])
        plot2.set_title("Rectus Right", fontsize=12)

        plot3 = plt.subplot2grid((4, 2), (1, 0))
        data_obliques_l = [[male_tonus[i]["Obliques_L"]], [female_tonus[i]["Obliques_L"]],
                           [over30_tonus[i]["Obliques_L"]],[EA_over_tonus[i]["Rectus_R"]], [EA_more_tonus[i]["Obliques_L"]]]
        plt.boxplot(data_obliques_l, positions = [1,2,3,4,5], widths=0.6, showmeans=True)
        plt.xticks([1, 2, 3, 4, 5], ("Male", "Female", "Over 30", "EA over30", "EA more evolve"),
                   fontsize=8)
        plt.ylim([0, 140])
        plot3.set_title("Obliques Left", fontsize=12)

        plot4 = plt.subplot2grid((4, 2), (1, 1))
        data_obliques_r = [[male_tonus[i]["Obliques_R"]], [female_tonus[i]["Obliques_R"]],
                           [over30_tonus[i]["Obliques_R"]],
                           [EA_over_tonus[i]["Obliques_R"]], [EA_more_tonus[i]["Obliques_R"]]]
        plt.boxplot(data_obliques_r, positions = [1,2,3,4,5], widths=0.6, showmeans=True)
        plt.xticks([1, 2, 3, 4, 5], ("Male", "Female", "Over 30", "EA over30", "EA more evolve"),
                   fontsize=8)
        plt.ylim([0, 140])
        plot4.set_title("Obliques Right", fontsize=12)

        plot5 = plt.subplot2grid((4, 2), (2, 0))
        data_ilicostalis_l = [[male_tonus[i]["Ilicostalis_L"]], [female_tonus[i]["Ilicostalis_L"]],
                              [over30_tonus[i]["Ilicostalis_L"]],
                              [EA_over_tonus[i]["Ilicostalis_L"]], [EA_more_tonus[i]["Ilicostalis_L"]]]
        plt.boxplot(data_ilicostalis_l, positions = [1,2,3,4,5], widths=0.6, showmeans=True)
        plt.xticks([1, 2, 3, 4, 5], ("Male", "Female", "Over 30", "EA over30", "EA under30", "EA more evolve"),
                   fontsize=8)
        plt.ylim([0, 140])
        plot5.set_title("Iliocostalis Left", fontsize=12)

        plot6 = plt.subplot2grid((4, 2), (2, 1))
        data_ilicostalis_r = [[male_tonus[i]["Ilicostalis_R"]], [female_tonus[i]["Ilicostalis_R"]],
                              [over30_tonus[i]["Ilicostalis_R"]],
                              [EA_over_tonus[i]["Ilicostalis_R"]], [EA_more_tonus[i]["Ilicostalis_R"]]]
        plt.boxplot(data_ilicostalis_r, positions = [1,2,3,4,5], widths=0.6, showmeans=True)
        plt.xticks([1, 2, 3, 4, 5], ("Male", "Female", "Over 30", "EA over30", "EA more evolve"),
                   fontsize=8)
        plt.ylim([0, 140])
        plot6.set_title("Iliocostalis Right", fontsize=12)

        plot7 = plt.subplot2grid((4, 2), (3, 0))
        data_multi_l = [[male_tonus[i]["Multifidus_L"]], [female_tonus[i]["Multifidus_L"]],
                        [over30_tonus[i]["Multifidus_L"]],
                        [EA_over_tonus[i]["Multifidus_L"]], [EA_more_tonus[i]["Multifidus_L"]]]
        plt.boxplot(data_multi_l, positions = [1,2,3,4,5], widths=0.6, showmeans=True)
        plt.xticks([1, 2, 3, 4, 5], ("Male", "Female", "Over 30", "EA over30", "EA more evolve"),
                   fontsize=8)
        plt.ylim([0, 140])
        plot7.set_title("Multifidus Left", fontsize=12)

        plot8 = plt.subplot2grid((4, 2), (3, 1))
        data_multi_r = [[male_tonus[i]["Multifidus_R"]], [female_tonus[i]["Multifidus_R"]],
                        [over30_tonus[i]["Multifidus_R"]],
                        [EA_over_tonus[i]["Multifidus_R"]], [EA_more_tonus[i]["Multifidus_R"]]]
        plt.boxplot(data_multi_r, positions = [1,2,3,4,5], widths=0.6, showmeans=True)
        plt.xticks([1, 2, 3, 4, 5], ("Male", "Female", "Over 30", "EA over30", "EA more evolve"),
                   fontsize=8)
        plt.ylim([0, 140])
        plot8.set_title("Multifidus Right", fontsize=12)

        plt.subplots_adjust(top=0.90, bottom=0.10, left=0.12, right=0.90, wspace=0.44, hspace=0.45)
        plt.show()
        pp.savefig(fig)
    pp.close()



def EMG_freq_front_boxplot(over30_freq, male_freq, female_freq, EA_over_freq, EA_more_freq):
    l = 0
    pp = PdfPages('EMG_Freqs_BoxPlot_Analysis.pdf')

    for i in over30_freq:
        fig = plt.figure(l)
        l = l + 1
        fig.suptitle(str(i) + " - EMG Frequencies Front Muscles", fontsize=21)

        row_peak_mean = 0
        col_peak_mean = 0
        row_80_median = 2
        col_80_median = 0

        colors = ['blue', 'green']
        color_index = 0
        plt.figtext(0.29, 0.91, "Peak Frequency", horizontalalignment='center', multialignment='center', size=18,
                    clip_on=True)
        plt.figtext(0.71, 0.91, "Mean Frequency", horizontalalignment='center', multialignment='center', size=18,
                    clip_on=True)
        plt.figtext(0.29, 0.46, "80% Frequency", horizontalalignment='center', multialignment='center', size=18,
                    clip_on=True)
        plt.figtext(0.71, 0.46, "Median Frequency", horizontalalignment='center', multialignment='center', size=18,
                    clip_on=True)

        for freq in over30_freq[i]["Rectus_L"]:
            if freq == "Peak" or freq == "Mean":
                row_l = row_peak_mean
                col_l = col_peak_mean
                row_r = row_peak_mean
                col_r = col_peak_mean + 1

                for muscle in over30_freq[i]:
                    if muscle == "Rectus_L" or muscle == "Obliques_L":
                        plot = plt.subplot2grid((4, 4), (row_l, col_l))
                        data = [[male_freq[i][muscle][freq]], [female_freq[i][muscle][freq]],
                                [over30_freq[i][muscle][freq]],
                                [EA_over_freq[i][muscle][freq]],
                                [EA_more_freq[i][muscle][freq]]]
                        box = plt.boxplot(data, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
                        plt.xticks([1, 2, 3, 4, 5],
                                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                                   fontsize=7)
                        plot.set_title(str(muscle), fontsize=12, style = 'italic')
                        row_l = row_l + 1

                    if muscle == "Rectus_R" or muscle == "Obliques_R":
                        plot = plt.subplot2grid((4, 4), (row_r, col_r))
                        data = [[male_freq[i][muscle][freq]], [female_freq[i][muscle][freq]],
                                [over30_freq[i][muscle][freq]],
                                [EA_over_freq[i][muscle][freq]],
                                [EA_more_freq[i][muscle][freq]]]
                        box = plt.boxplot(data, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
                        plt.xticks([1, 2, 3, 4, 5],
                                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                                   fontsize=7)
                        plot.set_title(str(muscle), fontsize=12, style='italic')
                        row_r = row_r + 1

                col_peak_mean = col_peak_mean + 2
                color_index = color_index + 1

            if freq == "Median" or freq == "80_freq":
                row_l = row_80_median
                col_l = col_80_median
                row_r = row_80_median
                col_r = col_80_median + 1

                for muscle in over30_freq[i]:
                    if muscle == "Rectus_L" or muscle == "Obliques_L":
                        plot = plt.subplot2grid((4, 4), (row_l, col_l))
                        data = [[male_freq[i][muscle][freq]], [female_freq[i][muscle][freq]],
                                [over30_freq[i][muscle][freq]],
                                [EA_over_freq[i][muscle][freq]],
                                [EA_more_freq[i][muscle][freq]]]
                        box = plt.boxplot(data, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
                        plt.xticks([1, 2, 3, 4, 5],
                                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                                   fontsize=7)
                        plot.set_title(str(muscle), fontsize=12, style='italic')
                        row_l = row_l + 1

                    if muscle == "Rectus_R" or muscle == "Obliques_R":
                        plot = plt.subplot2grid((4, 4), (row_r, col_r))
                        data = [[male_freq[i][muscle][freq]], [female_freq[i][muscle][freq]],
                                [over30_freq[i][muscle][freq]],
                                [EA_over_freq[i][muscle][freq]],
                                [EA_more_freq[i][muscle][freq]]]
                        box = plt.boxplot(data, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
                        plt.xticks([1, 2, 3, 4, 5],
                                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                                   fontsize=7)
                        plot.set_title(str(muscle), fontsize=12, style='italic')
                        row_r = row_r + 1

                col_80_median = col_80_median + 2
                color_index = color_index + 1

        plt.subplots_adjust(top=0.89, bottom=0.05, left=0.12, right=0.90, wspace=0.56, hspace=0.67)
        plt.show()
        pp.savefig(fig)
    pp.close()

def COP_parameters_boxplot(over_30, male, female, EA_over, EA_more):
    l = 0
    pp = PdfPages('COP_Parameters_BoxPlot_Analysis_Healthy_EA.pdf')

    for i in over_30:
        fig = plt.figure(l)
        l = l + 1
        fig.suptitle(str(i) + " - COP Parameters", fontsize=21)

        plot1 = plt.subplot2grid((3, 3), (0, 0))
        data_velocity_x = [[male[i]["Velocity_X"]], [female[i]["Velocity_X"]], [over_30[i]["Velocity_X"]],
                          [EA_over[i]["Velocity_X"]],
                          [EA_more[i]["Velocity_X"]]]
        plt.boxplot(data_velocity_x, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
        plt.xticks([1, 2, 3, 4, 5],
                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                   fontsize=7)
        plot1.set_title("Velocity COP X", fontsize=12)

        plot2 = plt.subplot2grid((3, 3), (0, 1))
        data_velocity_y = [[male[i]["Velocity_Y"]], [female[i]["Velocity_Y"]], [over_30[i]["Velocity_Y"]],
                            [EA_over[i]["Velocity_Y"]],
                            [EA_more[i]["Velocity_Y"]]]
        plt.boxplot(data_velocity_y, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
        plt.xticks([1, 2, 3, 4, 5],
                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                   fontsize=7)
        plot2.set_title("Velocity COP Y", fontsize=12)

        plot3 = plt.subplot2grid((3, 3), (1, 0))
        data_std_x = [[male[i]["STD_X"]], [female[i]["STD_X"]], [over_30[i]["STD_X"]],
                       [EA_over[i]["STD_X"]],
                       [EA_more[i]["STD_X"]]]
        plt.boxplot(data_std_x, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
        plt.xticks([1, 2, 3, 4, 5],
                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                   fontsize=7)
        plot3.set_title("STD COP X", fontsize=12)

        plot4 = plt.subplot2grid((3, 3), (1, 1))
        data_std_y = [[male[i]["STD_Y"]], [female[i]["STD_Y"]], [over_30[i]["STD_Y"]],
                       [EA_over[i]["STD_Y"]],
                       [EA_more[i]["STD_Y"]]]
        plt.boxplot(data_std_y, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
        plt.xticks([1, 2, 3, 4, 5],
                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                   fontsize=7)
        plot4.set_title("STD COP Y", fontsize=12)

        plot5 = plt.subplot2grid((3, 3), (2, 0))
        data_amp_x = [[male[i]["Amp_X"]], [female[i]["Amp_X"]], [over_30[i]["Amp_X"]],
                       [EA_over[i]["Amp_X"]],
                       [EA_more[i]["Amp_X"]]]
        plt.boxplot(data_amp_x, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
        plt.xticks([1, 2, 3, 4, 5],
                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                   fontsize=7)
        plot5.set_title("Amplitude COP X", fontsize=12)

        plot6 = plt.subplot2grid((3, 3), (2, 1))
        data_amp_y = [[male[i]["Amp_Y"]], [female[i]["Amp_Y"]], [over_30[i]["Amp_Y"]],
                       [EA_over[i]["Amp_Y"]],
                       [EA_more[i]["Amp_Y"]]]
        plt.boxplot(data_amp_y, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
        plt.xticks([1, 2, 3, 4, 5],
                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                   fontsize=7)
        plot6.set_title("Amplitude COP Y", fontsize=12)

        plot7 = plt.subplot2grid((3, 3), (1, 2))
        data_area = [[male[i]["Area"]], [female[i]["Area"]], [over_30[i]["Area"]],
                      [EA_over[i]["Area"]],
                      [EA_more[i]["Area"]]]
        plt.boxplot(data_area, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
        plt.xticks([1, 2, 3, 4, 5],
                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                   fontsize=7)
        plot7.set_title("Area", fontsize=12)

        plt.subplots_adjust(top=0.90, bottom=0.10, left=0.12, right=0.90, wspace=0.31, hspace=0.27)
        plt.show()
        pp.savefig(fig)
    pp.close()


def COP_freq_boxplot(over30, male, female, EA_over, EA_more):
    l = 0
    pp = PdfPages('COP_Freqs_BoxPlot_Analysis_Healthy_EA.pdf')

    for i in over30:
        fig = plt.figure(l)
        l = l + 1
        fig.suptitle(str(i) + " - COP Frequencies", fontsize=21)

        plot1 = plt.subplot2grid((2, 4), (0, 0))
        data_peak_x = [[male[i]["Peak_X"]], [female[i]["Peak_X"]], [over30[i]["Peak_X"]],
                       [EA_over[i]["Peak_X"]],
                       [EA_more[i]["Peak_X"]]]
        plt.boxplot(data_peak_x, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
        plt.xticks([1, 2, 3, 4, 5],
                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                   fontsize=7)
        plot1.set_title("Peak Frequency - COP X", fontsize=12)

        plot2 = plt.subplot2grid((2, 4), (0, 1))
        data_peak_y = [[male[i]["Peak_Y"]], [female[i]["Peak_Y"]], [over30[i]["Peak_Y"]],
                       [EA_over[i]["Peak_Y"]],
                       [EA_more[i]["Peak_Y"]]]
        plt.boxplot(data_peak_y, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
        plt.xticks([1, 2, 3, 4, 5],
                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                   fontsize=7)
        plot2.set_title("Peak Frequency - COP Y", fontsize=12)

        plot3 = plt.subplot2grid((2, 4), (0, 2))
        data_mean_x = [[male[i]["Mean_X"]], [female[i]["Mean_X"]], [over30[i]["Mean_X"]],
                       [EA_over[i]["Mean_X"]],
                       [EA_more[i]["Mean_X"]]]
        plt.boxplot(data_mean_x, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
        plt.xticks([1, 2, 3, 4, 5],
                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                   fontsize=7)
        plot3.set_title("Mean Frequency - COP X", fontsize=12)

        plot4 = plt.subplot2grid((2, 4), (0, 3))
        data_mean_y = [[male[i]["Mean_Y"]], [female[i]["Mean_Y"]], [over30[i]["Mean_Y"]],
                       [EA_over[i]["Mean_Y"]],
                       [EA_more[i]["Mean_Y"]]]
        plt.boxplot(data_mean_y, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
        plt.xticks([1, 2, 3, 4, 5],
                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                   fontsize=7)
        plot4.set_title("Mean Frequency - COP Y", fontsize=12)

        plot5 = plt.subplot2grid((2, 4), (1, 0))
        data_median_x = [[male[i]["Median_X"]], [female[i]["Median_X"]], [over30[i]["Median_X"]],
                         [EA_over[i]["Median_X"]],
                         [EA_more[i]["Median_X"]]]
        plt.boxplot(data_median_x,positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
        plt.xticks([1, 2, 3, 4, 5],
                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                   fontsize=7)
        plot5.set_title("Median Frequency - COP X", fontsize=12)

        plot6 = plt.subplot2grid((2, 4), (1, 1))
        data_median_y = [[male[i]["Median_Y"]], [female[i]["Median_Y"]], [over30[i]["Median_Y"]],
                         [EA_over[i]["Median_Y"]],
                         [EA_more[i]["Median_Y"]]]
        plt.boxplot(data_median_x, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
        plt.xticks([1, 2, 3, 4, 5],
                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                   fontsize=7)
        plot6.set_title("Median Frequency - COP Y", fontsize=12)

        plot7 = plt.subplot2grid((2, 4), (1, 2))
        data_80_X = [[male[i]["80_X"]], [female[i]["80_X"]], [over30[i]["80_X"]],
                     [EA_over[i]["80_X"]],
                     [EA_more[i]["80_X"]]]
        plt.boxplot(data_80_X, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
        plt.xticks([1, 2, 3, 4, 5],
                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                   fontsize=7)
        plot7.set_title("80% Frequency - COP X", fontsize=12)

        plot8 = plt.subplot2grid((2, 4), (1, 3))
        data_80_Y = [[male[i]["80_Y"]], [female[i]["80_Y"]], [over30[i]["80_Y"]],
                     [EA_over[i]["80_Y"]],
                     [EA_more[i]["80_Y"]]]
        plt.boxplot(data_80_Y, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
        plt.xticks([1, 2, 3, 4, 5],
                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                   fontsize=7)
        plot8.set_title("80% Frequency - COP Y", fontsize=12)

        plt.subplots_adjust(top=0.90, bottom=0.10, left=0.12, right=0.90, wspace=0.31, hspace=0.27)
        plt.show()
        pp.savefig(fig)
    pp.close()

def task_vs_relax(mean_over30, mean_male, mean_female, mean_EA_over, mean_EA_more,mean_rest_over30, mean_rest_male, mean_rest_female, mean_rest_EA_over, mean_rest_EA_more):

    l = 0
    pp = PdfPages('RelaxVSTask_MaxValues_Analysis_Healthy_EA.pdf')

    for i in mean_over30:

        fig = plt.figure(l)
        l = l + 1
        fig.suptitle("Relax VS Task (Max Value) - " + str(i), fontsize=21)

        plot1 = plt.subplot2grid((2, 4), (0, 0))

        plt.plot(mean_rest_over30["Rectus_L"], mean_over30[i]["Rectus_L"], 'ro', color='blue', label = "Over 30")
        plt.plot(mean_rest_male["Rectus_L"], mean_male[i]["Rectus_L"], 'ro', color='yellow', label = "Male 20-30")
        plt.plot(mean_rest_female["Rectus_L"], mean_female[i]["Rectus_L"], 'ro', color='green', label = "Female 20-30")
        plt.plot(mean_rest_EA_over["Rectus_L"], mean_EA_over[i]["Rectus_L"], 'ro', color='red', label = "EA Over 30")
        plt.plot(mean_rest_EA_more["Rectus_L"], mean_EA_more[i]["Rectus_L"], 'ro', color='magenta', label="EA evolved")

        plt.xlabel('Rest max Values', fontsize=10)
        plt.xlim([0,50])
        plt.ylim([0, 140])
        plt.ylabel('Rectus Left max Values', fontsize=10)
        plt.legend(bbox_to_anchor=(1.0, 1.0), loc=2, borderaxespad=0., fontsize=10)
        plot1.set_title("Rectus Left", fontsize=12)

        plot2 = plt.subplot2grid((2, 4), (0, 1))

        plt.plot(mean_rest_over30["Rectus_R"], mean_over30[i]["Rectus_R"], 'ro', color='blue', label="Over 30")
        plt.plot(mean_rest_male["Rectus_R"], mean_male[i]["Rectus_R"], 'ro', color='yellow', label="Male 20-30")
        plt.plot(mean_rest_female["Rectus_R"], mean_female[i]["Rectus_R"], 'ro', color='green', label="Female 20-30")
        plt.plot(mean_rest_EA_over["Rectus_R"], mean_EA_over[i]["Rectus_R"], 'ro', color='red', label="EA Over 30")
        plt.plot(mean_rest_EA_more["Rectus_R"], mean_EA_more[i]["Rectus_R"], 'ro', color='magenta', label="EA evolved")

        plt.xlabel('Rest max Values', fontsize=10)
        plt.ylabel('Rectus Right max Values', fontsize=10)
        plt.xlim([0, 50])
        plt.ylim([0, 140])
        plt.legend(bbox_to_anchor=(1.0, 1.0), loc=2, borderaxespad=0., fontsize=10)
        plot2.set_title("Rectus Right", fontsize=12)

        plot3 = plt.subplot2grid((2, 4), (0, 2))

        plt.plot(mean_rest_over30["Obliques_L"], mean_over30[i]["Obliques_L"], 'ro', color='blue', label="Over 30")
        plt.plot(mean_rest_male["Obliques_L"], mean_male[i]["Obliques_L"], 'ro', color='yellow', label="Male 20-30")
        plt.plot(mean_rest_female["Obliques_L"], mean_female[i]["Obliques_L"], 'ro', color='green', label="Female 20-30")
        plt.plot(mean_rest_EA_over["Obliques_L"], mean_EA_over[i]["Obliques_L"], 'ro', color='red', label="EA Over 30")
        plt.plot(mean_rest_EA_more["Obliques_L"], mean_EA_more[i]["Obliques_L"], 'ro', color='magenta', label="EA evolved")

        plt.xlabel('Rest max Values', fontsize=10)
        plt.ylabel('Obliques Left max Values', fontsize=10)
        plt.xlim([0, 50])
        plt.ylim([0, 140])
        plt.legend(bbox_to_anchor=(1.0, 1.0), loc=2, borderaxespad=0., fontsize=10)
        plot3.set_title("Obliques Left", fontsize=12)

        plot4 = plt.subplot2grid((2, 4), (0, 3))
        plt.plot(mean_rest_over30["Obliques_R"], mean_over30[i]["Obliques_R"], 'ro', color='blue', label="Over 30")
        plt.plot(mean_rest_male["Obliques_R"], mean_male[i]["Obliques_R"], 'ro', color='yellow', label="Male 20-30")
        plt.plot(mean_rest_female["Obliques_R"], mean_female[i]["Obliques_R"], 'ro', color='green',
                 label="Female 20-30")
        plt.plot(mean_rest_EA_over["Obliques_R"], mean_EA_over[i]["Obliques_R"], 'ro', color='red', label="EA Over 30")
        plt.plot(mean_rest_EA_more["Obliques_R"], mean_EA_more[i]["Obliques_R"], 'ro', color='magenta',
                 label="EA evolved")

        plt.xlabel('Rest max Values', fontsize=10)
        plt.ylabel('Obliques Right max Values', fontsize=10)
        plt.xlim([0, 50])
        plt.ylim([0, 140])
        plt.legend(bbox_to_anchor=(1.0, 1.0), loc=2, borderaxespad=0., fontsize=10)
        plot4.set_title("Obliques Right", fontsize=12)

        plot5 = plt.subplot2grid((2, 4), (1, 0))

        plt.plot(mean_rest_over30["Ilicostalis_L"], mean_over30[i]["Ilicostalis_L"], 'ro', color='blue', label="Over 30")
        plt.plot(mean_rest_male["Ilicostalis_L"], mean_male[i]["Ilicostalis_L"], 'ro', color='yellow', label="Male 20-30")
        plt.plot(mean_rest_female["Ilicostalis_L"], mean_female[i]["Ilicostalis_L"], 'ro', color='green',
                 label="Female 20-30")
        plt.plot(mean_rest_EA_over["Ilicostalis_L"], mean_EA_over[i]["Ilicostalis_L"], 'ro', color='red', label="EA Over 30")
        plt.plot(mean_rest_EA_more["Ilicostalis_L"], mean_EA_more[i]["Ilicostalis_L"], 'ro', color='magenta',
                 label="EA evolved")

        plt.xlabel('Rest max Values', fontsize=10)
        plt.ylabel('Iliocostalis Left max Values', fontsize=10)
        plt.xlim([0, 50])
        plt.ylim([0, 140])
        plt.legend(bbox_to_anchor=(1.0, 1.0), loc=2, borderaxespad=0., fontsize=10)
        plot5.set_title("Iliocostalis Left", fontsize=12)

        plot6 = plt.subplot2grid((2, 4), (1, 1))

        plt.plot(mean_rest_over30["Ilicostalis_R"], mean_over30[i]["Ilicostalis_R"], 'ro', color='blue',
                 label="Over 30")
        plt.plot(mean_rest_male["Ilicostalis_R"], mean_male[i]["Ilicostalis_R"], 'ro', color='yellow',
                 label="Male 20-30")
        plt.plot(mean_rest_female["Ilicostalis_R"], mean_female[i]["Ilicostalis_R"], 'ro', color='green',
                 label="Female 20-30")
        plt.plot(mean_rest_EA_over["Ilicostalis_R"], mean_EA_over[i]["Ilicostalis_R"], 'ro', color='red',
                 label="EA Over 30")
        plt.plot(mean_rest_EA_more["Ilicostalis_R"], mean_EA_more[i]["Ilicostalis_R"], 'ro', color='magenta',
                 label="EA evolved")

        plt.xlabel('Rest max Values', fontsize=10)
        plt.ylabel('Iliocostalis Right max Values', fontsize=10)
        plt.xlim([0, 50])
        plt.ylim([0, 140])
        plt.legend(bbox_to_anchor=(1.0, 1.0), loc=2, borderaxespad=0., fontsize=10)
        plot6.set_title("Iliocostalis Right", fontsize=12)

        plot7 = plt.subplot2grid((2, 4), (1, 2))

        plt.plot(mean_rest_over30["Multifidus_L"], mean_over30[i]["Multifidus_L"], 'ro', color='blue',
                 label="Over 30")
        plt.plot(mean_rest_male["Multifidus_L"], mean_male[i]["Multifidus_L"], 'ro', color='yellow',
                 label="Male 20-30")
        plt.plot(mean_rest_female["Multifidus_L"], mean_female[i]["Multifidus_L"], 'ro', color='green',
                 label="Female 20-30")
        plt.plot(mean_rest_EA_over["Multifidus_L"], mean_EA_over[i]["Multifidus_L"], 'ro', color='red',
                 label="EA Over 30")
        plt.plot(mean_rest_EA_more["Multifidus_L"], mean_EA_more[i]["Multifidus_L"], 'ro', color='magenta',
                 label="EA evolved")

        plt.xlabel('Rest max Values', fontsize=10)
        plt.ylabel('Multifidus Left max Values', fontsize=10)
        plt.xlim([0, 50])
        plt.ylim([0, 140])
        plt.legend(bbox_to_anchor=(1.0, 1.0), loc=2, borderaxespad=0., fontsize=10)
        plot7.set_title("Multifidus Left", fontsize=12)

        plot8 = plt.subplot2grid((2, 4), (1, 3))

        plt.plot(mean_rest_over30["Multifidus_R"], mean_over30[i]["Multifidus_R"], 'ro', color='blue',
                 label="Over 30")
        plt.plot(mean_rest_male["Multifidus_R"], mean_male[i]["Multifidus_R"], 'ro', color='yellow',
                 label="Male 20-30")
        plt.plot(mean_rest_female["Multifidus_R"], mean_female[i]["Multifidus_R"], 'ro', color='green',
                 label="Female 20-30")
        plt.plot(mean_rest_EA_over["Multifidus_R"], mean_EA_over[i]["Multifidus_R"], 'ro', color='red',
                 label="EA Over 30")
        plt.plot(mean_rest_EA_more["Multifidus_R"], mean_EA_more[i]["Multifidus_R"], 'ro', color='magenta',
                 label="EA evolved")

        plt.xlabel('Rest max Values', fontsize=10)
        plt.ylabel('Multifidus Right max Values', fontsize=10)
        plt.xlim([0, 50])
        plt.ylim([0, 140])
        plt.legend(bbox_to_anchor=(1.0, 1.0), loc=2, borderaxespad=0., fontsize=10)
        plot8.set_title("Multifidus Right", fontsize=12)

        plt.subplots_adjust(top=0.90, bottom=0.10, left=0.07, right=0.86, wspace=0.95, hspace=0.68)
        plt.show()
        pp.savefig(fig)
    pp.close()


def means_database(over30, male, female, EA,rest_over30, restmale, restfemale, restEA):
    mean_over30 = {}
    mean_male = {}
    mean_female = {}
    mean_EA = {}

    mean_rest_over30 = {}
    mean_rest_male = {}
    mean_rest_female = {}
    mean_rest_EA = {}

    for i in over30:
        mean_EMG_over30 = {}
        mean_EMG_male = {}
        mean_EMG_female = {}
        mean_EMG_EA = {}

        for n in over30[i]:
            mean_EMG_over30[n] = np.mean(over30[i][n])
            mean_EMG_male[n] = np.mean(male[i][n])
            mean_EMG_female[n] = np.mean(female[i][n])
            mean_EMG_EA[n] = np.mean(EA[i][n])

        mean_over30[i] = mean_EMG_over30
        mean_male[i] = mean_EMG_male
        mean_female[i] = mean_EMG_female
        mean_EA[i] = mean_EMG_EA

    for muscle in rest_over30:
        rest = []
        for index in rest_over30[muscle]:
            if index != None:
                rest.append(index)
        mean_rest_over30[muscle] = np.mean(rest)

    for muscle in restmale:
        rest = []
        for index in restmale[muscle]:
            if index != None:
                rest.append(index)
        mean_rest_male[muscle] = np.mean(rest)

    for muscle in restfemale:
        rest = []
        for index in restfemale[muscle]:
            if index != None:
                rest.append(index)
        mean_rest_female[muscle] = np.mean(rest)

    for muscle in restEA:
        rest = []
        for index in restEA[muscle]:
            if index != None:
                rest.append(index)
        mean_rest_EA[muscle] = np.mean(rest)

    return mean_over30, mean_male, mean_female, mean_EA, mean_rest_over30, mean_rest_male, mean_rest_female, mean_rest_EA

def EMG_freq_back_boxplot(over30_freq, male_freq, female_freq, EA_over_freq, EA_more_freq):
    l = 0
    pp = PdfPages('EMG_Freqs_Back_BoxPlot_Analysis_Healthy_EA.pdf')

    for i in over30_freq:
        fig = plt.figure(l)
        l = l + 1
        fig.suptitle(str(i) + " - EMG Frequencies Back Muscles", fontsize=21)

        row_peak_mean = 0
        col_peak_mean = 0
        row_80_median = 2
        col_80_median = 0

        colors = ['blue', 'green']
        color_index = 0
        plt.figtext(0.29, 0.91, "Peak Frequency", horizontalalignment = 'center', multialignment = 'center', size = 18,
                    clip_on = True)
        plt.figtext(0.71, 0.91, "Mean Frequency", horizontalalignment='center', multialignment='center', size=18,
                    clip_on=True)
        plt.figtext(0.29, 0.46, "80% Frequency", horizontalalignment='center', multialignment='center', size=18,
                    clip_on=True)
        plt.figtext(0.71, 0.46, "Median Frequency", horizontalalignment='center', multialignment='center', size=18,
                    clip_on=True)

        for freq in over30_freq[i]["Rectus_L"]:
            if freq == "Peak" or freq == "Mean":
                row_l = row_peak_mean
                col_l = col_peak_mean
                row_r = row_peak_mean
                col_r = col_peak_mean + 1

                for muscle in over30_freq[i]:
                    if muscle == "Ilicostalis_L" or muscle == "Multi_L":
                        plot = plt.subplot2grid((4, 4), (row_l, col_l))
                        data = [[male_freq[i][muscle][freq]], [female_freq[i][muscle][freq]],
                                [over30_freq[i][muscle][freq]],
                                [EA_over_freq[i][muscle][freq]],
                                [EA_more_freq[i][muscle][freq]]]
                        box = plt.boxplot(data, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
                        plt.xticks([1, 2, 3, 4, 5],
                                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                                   fontsize=7)
                        plot.set_title(str(muscle), fontsize=12, style = 'italic')
                        row_l = row_l + 1

                    if muscle == "Ilicostalis_R" or muscle == "Multi_R":
                        plot = plt.subplot2grid((4, 4), (row_r, col_r))
                        data = [[male_freq[i][muscle][freq]], [female_freq[i][muscle][freq]],
                                [over30_freq[i][muscle][freq]],
                                [EA_over_freq[i][muscle][freq]],
                                [EA_more_freq[i][muscle][freq]]]
                        box = plt.boxplot(data, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
                        plt.xticks([1, 2, 3, 4, 5],
                                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                                   fontsize=7)
                        plot.set_title(str(muscle), fontsize=12, style='italic')
                        row_r = row_r + 1

                col_peak_mean = col_peak_mean + 2
                color_index = color_index + 1

            if freq == "Median" or freq == "80_freq":
                row_l = row_80_median
                col_l = col_80_median
                row_r = row_80_median
                col_r = col_80_median + 1

                for muscle in over30_freq[i]:
                    if muscle == "Ilicostalis_L" or muscle == "Multi_L":
                        plot = plt.subplot2grid((4, 4), (row_l, col_l))
                        data = [[male_freq[i][muscle][freq]], [female_freq[i][muscle][freq]],
                                [over30_freq[i][muscle][freq]],
                                [EA_over_freq[i][muscle][freq]],
                                [EA_more_freq[i][muscle][freq]]]
                        box = plt.boxplot(data, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
                        plt.xticks([1, 2, 3, 4, 5],
                                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                                   fontsize=7)
                        plot.set_title(str(muscle), fontsize=12, style='italic')
                        row_l = row_l + 1

                    if muscle == "Ilicostalis_R" or muscle == "Multi_R":
                        plot = plt.subplot2grid((4, 4), (row_r, col_r))
                        data = [[male_freq[i][muscle][freq]], [female_freq[i][muscle][freq]],
                                [over30_freq[i][muscle][freq]],
                                [EA_over_freq[i][muscle][freq]],
                                [EA_more_freq[i][muscle][freq]]]
                        box = plt.boxplot(data, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
                        plt.xticks([1, 2, 3, 4, 5],
                                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                                   fontsize=7)
                        plt.xticks([1, 2, 3, 4, 5],
                                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                                   fontsize=7)
                        plot.set_title(str(muscle), fontsize=12, style='italic')
                        row_r = row_r + 1

                col_80_median = col_80_median + 2
                color_index = color_index + 1

        plt.subplots_adjust(top=0.89, bottom=0.05, left=0.12, right=0.90, wspace=0.56, hspace=0.67)
        plt.show()
        pp.savefig(fig)
    pp.close()


def get_value_mean_tonus(EMG_array):

    wb_over30 = load_workbook(file_excel_over30)
    wb_20 = load_workbook(file_excel_20)
    wb_EA = load_workbook(file_excel_EA)


    count = 616
    value_final_over30 = {}
    value_final_male = {}
    value_final_female = {}
    value_final_EA_over30 = {}
    value_final_EA_evoluted = {}

    for task in EMG_array:

        value_rectus_l_over30 = []
        value_rectus_r_over30 = []
        value_obliques_l_over30 = []
        value_obliques_r_over30 = []
        value_ilicostalis_l_over30 = []
        value_ilicostalis_r_over30 = []
        value_multi_l_over30 = []
        value_multi_r_over30 = []

        value_rectus_l_female = []
        value_rectus_r_female = []
        value_obliques_l_female = []
        value_obliques_r_female = []
        value_ilicostalis_l_female = []
        value_ilicostalis_r_female = []
        value_multi_l_female = []
        value_multi_r_female = []

        value_rectus_l_male = []
        value_rectus_r_male = []
        value_obliques_l_male = []
        value_obliques_r_male = []
        value_ilicostalis_l_male = []
        value_ilicostalis_r_male = []
        value_multi_l_male = []
        value_multi_r_male = []

        value_rectus_l_EA_over30 = []
        value_rectus_r_EA_over30 = []
        value_obliques_l_EA_over30 = []
        value_obliques_r_EA_over30 = []
        value_ilicostalis_l_EA_over30 = []
        value_ilicostalis_r_EA_over30 = []
        value_multi_l_EA_over30 = []
        value_multi_r_EA_over30 = []

        value_rectus_l_EA_evoluted = []
        value_rectus_r_EA_evoluted = []
        value_obliques_l_EA_evoluted = []
        value_obliques_r_EA_evoluted = []
        value_ilicostalis_l_EA_evoluted = []
        value_ilicostalis_r_EA_evoluted = []
        value_multi_l_EA_evoluted = []
        value_multi_r_EA_evoluted = []

        for i in wb_over30.sheetnames:
            if i != 'Sheet1' and i != 'Statistical Analysis_20_Female' and i!= 'Statistical Analysis_20_Male'\
                    and i != 'Statistical Analysis over 30' and i != 'Folha1':

                ws_over30 = wb_over30.get_sheet_by_name(i)
                value_rectus_l_over30.append(ws_over30['B' + str(count)].value)
                value_rectus_r_over30.append(ws_over30['C' + str(count)].value)
                value_obliques_l_over30.append(ws_over30['B' + str(count+1)].value)
                value_obliques_r_over30.append(ws_over30['C' + str(count+1)].value)
                value_ilicostalis_l_over30.append(ws_over30['B' + str(count+2)].value)
                value_ilicostalis_r_over30.append(ws_over30['C' + str(count+2)].value)
                value_multi_l_over30.append(ws_over30['B' + str(count+3)].value)
                value_multi_r_over30.append(ws_over30['C' + str(count+3)].value)

        for i in wb_20.sheetnames:
            if i != 'Sheet1' and i != 'Statistical Analysis_20_Female' and i != 'Statistical Analysis_20_Male' \
                    and i != 'Statistical Analysis over 30' and i != 'Folha1' and i != 'Patient5_Healthy' \
                    and i != 'Patient8_Healthy' and i != 'Patient10_Healthy' and i != 'Patient13_Healthy' \
                    and i != 'Patient17_Healthy' and i != 'Patient18_Healthy' and i != 'Patient23_Healthy' \
                    and i != 'Patient25_Healthy' and i != 'Patient27_Healthy' and i != 'Patient30_Healthy' \
                    and i != 'Patient37_Healthy' and i != 'Patient38_Healthy':

                ws_female = wb_20.get_sheet_by_name(i)
                value_rectus_l_female.append(ws_female['B' + str(count)].value)
                value_rectus_r_female.append(ws_female['C' + str(count)].value)
                value_obliques_l_female.append(ws_female['B' + str(count+1)].value)
                value_obliques_r_female.append(ws_female['C' + str(count+1)].value)
                value_ilicostalis_l_female.append(ws_female['B' + str(count+2)].value)
                value_ilicostalis_r_female.append(ws_female['C' + str(count+2)].value)
                value_multi_l_female.append(ws_female['B' + str(count+3)].value)
                value_multi_r_female.append(ws_female['C' + str(count+3)].value)

            if i == 'Patient8_Healthy' or i == 'Patient10_Healthy' or i == 'Patient13_Healthy' \
                    or i == 'Patient17_Healthy' or i == 'Patient18_Healthy' or i == 'Patient23_Healthy' \
                    or i == 'Patient25_Healthy' or i == 'Patient27_Healthy' or i == 'Patient30_Healthy' \
                    or i == 'Patient5_Healthy' or i == 'Patient37_Healthy' or i == 'Patient38_Healthy':

                ws_male = wb_20.get_sheet_by_name(i)
                value_rectus_l_male.append(ws_male['B' + str(count)].value)
                value_rectus_r_male.append(ws_male['C' + str(count)].value)
                value_obliques_l_male.append(ws_male['B' + str(count + 1)].value)
                value_obliques_r_male.append(ws_male['C' + str(count + 1)].value)
                value_ilicostalis_l_male.append(ws_male['B' + str(count + 2)].value)
                value_ilicostalis_r_male.append(ws_male['C' + str(count + 2)].value)
                value_multi_l_male.append(ws_male['B' + str(count + 3)].value)
                value_multi_r_male.append(ws_male['C' + str(count + 3)].value)


        for i in wb_EA.sheetnames:
            if i != 'Sheet1' and i != 'Patient4_EA' and i != 'Patient8_EA' and i != 'Patient11_EA'\
                    and i != "Statistical Analysis - EA evol" and i != "Statistical Analysis - EA over3":

                ws_EA_over30 = wb_EA.get_sheet_by_name(i)
                value_rectus_l_EA_over30.append(ws_EA_over30['B' + str(count)].value)
                value_rectus_r_EA_over30.append(ws_EA_over30['C' + str(count)].value)
                value_obliques_l_EA_over30.append(ws_EA_over30['B' + str(count+1)].value)
                value_obliques_r_EA_over30.append(ws_EA_over30['C' + str(count+1)].value)
                value_ilicostalis_l_EA_over30.append(ws_EA_over30['B' + str(count+2)].value)
                value_ilicostalis_r_EA_over30.append(ws_EA_over30['C' + str(count+2)].value)
                value_multi_l_EA_over30.append(ws_EA_over30['B' + str(count+3)].value)
                value_multi_r_EA_over30.append(ws_EA_over30['C' + str(count+3)].value)

            if i == 'Patient4_EA' or i == 'Patient8_EA' or i == 'Patient11_EA':

                ws_EA_evoluted = wb_EA.get_sheet_by_name(i)
                value_rectus_l_EA_evoluted.append(ws_EA_evoluted['B' + str(count)].value)
                value_rectus_r_EA_evoluted.append(ws_EA_evoluted['C' + str(count)].value)
                value_obliques_l_EA_evoluted.append(ws_EA_evoluted['B' + str(count+1)].value)
                value_obliques_r_EA_evoluted.append(ws_EA_evoluted['C' + str(count+1)].value)
                value_ilicostalis_l_EA_evoluted.append(ws_EA_evoluted['B' + str(count+2)].value)
                value_ilicostalis_r_EA_evoluted.append(ws_EA_evoluted['C' + str(count+2)].value)
                value_multi_l_EA_evoluted.append(ws_EA_evoluted['B' + str(count+3)].value)
                value_multi_r_EA_evoluted.append(ws_EA_evoluted['C' + str(count+3)].value)

        count = count + 13

        value_final_over30[task] = {"Rectus_L": value_rectus_l_over30,"Rectus_R": value_rectus_r_over30, "Obliques_L": value_obliques_l_over30,
                             "Obliques_R": value_obliques_r_over30, "Ilicostalis_L": value_ilicostalis_l_over30, "Ilicostalis_R": value_ilicostalis_r_over30,
                             "Multifidus_L": value_multi_l_over30, "Multifidus_R": value_multi_r_over30}

        value_final_male[task] = {"Rectus_L": value_rectus_l_male, "Rectus_R": value_rectus_r_male,
                                    "Obliques_L": value_obliques_l_male,
                                    "Obliques_R": value_obliques_r_male, "Ilicostalis_L": value_ilicostalis_l_male,
                                    "Ilicostalis_R": value_ilicostalis_r_male,
                                    "Multifidus_L": value_multi_l_male, "Multifidus_R": value_multi_r_male}

        value_final_female[task] = {"Rectus_L": value_rectus_l_female, "Rectus_R": value_rectus_r_female,
                                    "Obliques_L": value_obliques_l_female,
                                    "Obliques_R": value_obliques_r_female, "Ilicostalis_L": value_ilicostalis_l_female,
                                    "Ilicostalis_R": value_ilicostalis_r_female,
                                    "Multifidus_L": value_multi_l_female, "Multifidus_R": value_multi_r_female}

        value_final_EA_over30[task] = {"Rectus_L": value_rectus_l_EA_over30, "Rectus_R": value_rectus_r_EA_over30,
                                    "Obliques_L": value_obliques_l_EA_over30,
                                    "Obliques_R": value_obliques_r_EA_over30, "Ilicostalis_L": value_ilicostalis_l_EA_over30,
                                    "Ilicostalis_R": value_ilicostalis_r_EA_over30,
                                    "Multifidus_L": value_multi_l_EA_over30, "Multifidus_R": value_multi_r_EA_over30}

        value_final_EA_evoluted[task] = {"Rectus_L": value_rectus_l_EA_evoluted, "Rectus_R": value_rectus_r_EA_evoluted,
                                "Obliques_L": value_obliques_l_EA_evoluted,
                                "Obliques_R": value_obliques_r_EA_evoluted, "Ilicostalis_L": value_ilicostalis_l_EA_evoluted,
                                "Ilicostalis_R": value_ilicostalis_r_EA_evoluted,
                                "Multifidus_L": value_multi_l_EA_evoluted, "Multifidus_R": value_multi_r_EA_evoluted}

    return value_final_over30, value_final_male, value_final_female, value_final_EA_over30,value_final_EA_evoluted

def get_value_mean_tonus_rest():

    wb_over30 = load_workbook(file_excel_over30)
    wb_20 = load_workbook(file_excel_20)
    wb_EA = load_workbook(file_excel_EA)


    count = 736

    value_rectus_l_over30 = []
    value_rectus_r_over30 = []
    value_obliques_l_over30 = []
    value_obliques_r_over30 = []
    value_ilicostalis_l_over30 = []
    value_ilicostalis_r_over30 = []
    value_multi_l_over30 = []
    value_multi_r_over30 = []

    value_rectus_l_female = []
    value_rectus_r_female = []
    value_obliques_l_female = []
    value_obliques_r_female = []
    value_ilicostalis_l_female = []
    value_ilicostalis_r_female = []
    value_multi_l_female = []
    value_multi_r_female = []

    value_rectus_l_male = []
    value_rectus_r_male = []
    value_obliques_l_male = []
    value_obliques_r_male = []
    value_ilicostalis_l_male = []
    value_ilicostalis_r_male = []
    value_multi_l_male = []
    value_multi_r_male = []

    value_rectus_l_EA_over30 = []
    value_rectus_r_EA_over30 = []
    value_obliques_l_EA_over30 = []
    value_obliques_r_EA_over30 = []
    value_ilicostalis_l_EA_over30 = []
    value_ilicostalis_r_EA_over30 = []
    value_multi_l_EA_over30 = []
    value_multi_r_EA_over30 = []

    value_rectus_l_EA_evaluted = []
    value_rectus_r_EA_evaluted = []
    value_obliques_l_EA_evaluted = []
    value_obliques_r_EA_evaluted = []
    value_ilicostalis_l_EA_evaluted = []
    value_ilicostalis_r_EA_evaluted = []
    value_multi_l_EA_evaluted = []
    value_multi_r_EA_evaluted = []

    for i in wb_over30.sheetnames:
        if i != 'Sheet1' and i != 'Statistical Analysis_20_Female' and i != 'Statistical Analysis_20_Male' \
                and i != 'Statistical Analysis over 30' and i != 'Folha1':
            ws_over30 = wb_over30.get_sheet_by_name(i)
            value_rectus_l_over30.append(ws_over30['B' + str(count)].value)
            value_rectus_r_over30.append(ws_over30['C' + str(count)].value)
            value_obliques_l_over30.append(ws_over30['B' + str(count + 1)].value)
            value_obliques_r_over30.append(ws_over30['C' + str(count + 1)].value)
            value_ilicostalis_l_over30.append(ws_over30['B' + str(count + 2)].value)
            value_ilicostalis_r_over30.append(ws_over30['C' + str(count + 2)].value)
            value_multi_l_over30.append(ws_over30['B' + str(count + 3)].value)
            value_multi_r_over30.append(ws_over30['C' + str(count + 3)].value)

    for i in wb_20.sheetnames:
        if i != 'Sheet1' and i != 'Statistical Analysis_20_Female' and i != 'Statistical Analysis_20_Male' \
                and i != 'Statistical Analysis over 30' and i != 'Folha1' and i != 'Patient5_Healthy' \
                and i != 'Patient8_Healthy' and i != 'Patient10_Healthy' and i != 'Patient13_Healthy' \
                and i != 'Patient17_Healthy' and i != 'Patient18_Healthy' and i != 'Patient23_Healthy' \
                and i != 'Patient25_Healthy' and i != 'Patient27_Healthy' and i != 'Patient30_Healthy' \
                and i != 'Patient37_Healthy' and i != 'Patient38_Healthy':
            ws_female = wb_20.get_sheet_by_name(i)
            value_rectus_l_female.append(ws_female['B' + str(count)].value)
            value_rectus_r_female.append(ws_female['C' + str(count)].value)
            value_obliques_l_female.append(ws_female['B' + str(count + 1)].value)
            value_obliques_r_female.append(ws_female['C' + str(count + 1)].value)
            value_ilicostalis_l_female.append(ws_female['B' + str(count + 2)].value)
            value_ilicostalis_r_female.append(ws_female['C' + str(count + 2)].value)
            value_multi_l_female.append(ws_female['B' + str(count + 3)].value)
            value_multi_r_female.append(ws_female['C' + str(count + 3)].value)

        if i == 'Patient8_Healthy' or i == 'Patient10_Healthy' or i == 'Patient13_Healthy' \
                or i == 'Patient17_Healthy' or i == 'Patient18_Healthy' or i == 'Patient23_Healthy' \
                or i == 'Patient25_Healthy' or i == 'Patient27_Healthy' or i == 'Patient30_Healthy' \
                or i == 'Patient5_Healthy' or i == 'Patient37_Healthy' or i == 'Patient38_Healthy':
            ws_male = wb_20.get_sheet_by_name(i)
            value_rectus_l_male.append(ws_male['B' + str(count)].value)
            value_rectus_r_male.append(ws_male['C' + str(count)].value)
            value_obliques_l_male.append(ws_male['B' + str(count + 1)].value)
            value_obliques_r_male.append(ws_male['C' + str(count + 1)].value)
            value_ilicostalis_l_male.append(ws_male['B' + str(count + 2)].value)
            value_ilicostalis_r_male.append(ws_male['C' + str(count + 2)].value)
            value_multi_l_male.append(ws_male['B' + str(count + 3)].value)
            value_multi_r_male.append(ws_male['C' + str(count + 3)].value)

    for i in wb_EA.sheetnames:
        if i != 'Sheet1' and i != 'Patient4_EA' and i != 'Patient8_EA' and i != 'Patient11_EA'\
                and i != "Statistical Analysis - EA evol" and i != "Statistical Analysis - EA over3":
            ws_EA_over30 = wb_EA.get_sheet_by_name(i)
            value_rectus_l_EA_over30.append(ws_EA_over30['B' + str(count)].value)
            value_rectus_r_EA_over30.append(ws_EA_over30['C' + str(count)].value)
            value_obliques_l_EA_over30.append(ws_EA_over30['B' + str(count + 1)].value)
            value_obliques_r_EA_over30.append(ws_EA_over30['C' + str(count + 1)].value)
            value_ilicostalis_l_EA_over30.append(ws_EA_over30['B' + str(count + 2)].value)
            value_ilicostalis_r_EA_over30.append(ws_EA_over30['C' + str(count + 2)].value)
            value_multi_l_EA_over30.append(ws_EA_over30['B' + str(count + 3)].value)
            value_multi_r_EA_over30.append(ws_EA_over30['C' + str(count + 3)].value)

        if i == 'Patient4_EA' or i == 'Patient8_EA' or i == 'Patient11_EA':
            ws_EA_evaluted = wb_EA.get_sheet_by_name(i)
            value_rectus_l_EA_evaluted.append(ws_EA_evaluted['B' + str(count)].value)
            value_rectus_r_EA_evaluted.append(ws_EA_evaluted['C' + str(count)].value)
            value_obliques_l_EA_evaluted.append(ws_EA_evaluted['B' + str(count + 1)].value)
            value_obliques_r_EA_evaluted.append(ws_EA_evaluted['C' + str(count + 1)].value)
            value_ilicostalis_l_EA_evaluted.append(ws_EA_evaluted['B' + str(count + 2)].value)
            value_ilicostalis_r_EA_evaluted.append(ws_EA_evaluted['C' + str(count + 2)].value)
            value_multi_l_EA_evaluted.append(ws_EA_evaluted['B' + str(count + 3)].value)
            value_multi_r_EA_evaluted.append(ws_EA_evaluted['C' + str(count + 3)].value)


    value_final_over30 = {"Rectus_L": value_rectus_l_over30, "Rectus_R": value_rectus_r_over30,
                                "Obliques_L": value_obliques_l_over30,
                                "Obliques_R": value_obliques_r_over30, "Ilicostalis_L": value_ilicostalis_l_over30,
                                "Ilicostalis_R": value_ilicostalis_r_over30,
                                "Multifidus_L": value_multi_l_over30, "Multifidus_R": value_multi_r_over30}

    value_final_male = {"Rectus_L": value_rectus_l_male, "Rectus_R": value_rectus_r_male,
                              "Obliques_L": value_obliques_l_male,
                              "Obliques_R": value_obliques_r_male, "Ilicostalis_L": value_ilicostalis_l_male,
                              "Ilicostalis_R": value_ilicostalis_r_male,
                              "Multifidus_L": value_multi_l_male, "Multifidus_R": value_multi_r_male}

    value_final_female = {"Rectus_L": value_rectus_l_female, "Rectus_R": value_rectus_r_female,
                                "Obliques_L": value_obliques_l_female,
                                "Obliques_R": value_obliques_r_female, "Ilicostalis_L": value_ilicostalis_l_female,
                                "Ilicostalis_R": value_ilicostalis_r_female,
                                "Multifidus_L": value_multi_l_female, "Multifidus_R": value_multi_r_female}

    value_final_EA_over30 = {"Rectus_L": value_rectus_l_EA_over30, "Rectus_R": value_rectus_r_EA_over30,
                            "Obliques_L": value_obliques_l_EA_over30,
                            "Obliques_R": value_obliques_r_EA_over30, "Ilicostalis_L": value_ilicostalis_l_EA_over30,
                            "Ilicostalis_R": value_ilicostalis_r_EA_over30,
                            "Multifidus_L": value_multi_l_EA_over30, "Multifidus_R": value_multi_r_EA_over30}

    value_final_EA_evaluted = {"Rectus_L": value_rectus_l_EA_evaluted, "Rectus_R": value_rectus_r_EA_evaluted,
                      "Obliques_L": value_obliques_l_EA_evaluted,
                      "Obliques_R": value_obliques_r_EA_evaluted, "Ilicostalis_L": value_ilicostalis_l_EA_evaluted,
                      "Ilicostalis_R": value_ilicostalis_r_EA_evaluted,
                      "Multifidus_L": value_multi_l_EA_evaluted, "Multifidus_R": value_multi_r_EA_evaluted}


    return value_final_over30, value_final_male, value_final_female, value_final_EA_over30, value_final_EA_evaluted

def EMG_mean_values_boxplot(over30_tonus, male_tonus, female_tonus, EA_over_tonus, EA_more_tonus):
    l = 0
    pp = PdfPages('Tonus_Muscular_BoxPlot_Analysis_Healthy_EA.pdf')

    for i in over30_tonus:

        fig = plt.figure(l)
        l = l + 1
        fig.suptitle(str(i) + " - Mean Value of normalized EMG", fontsize=21)

        plot1 = plt.subplot2grid((4, 2), (0, 0))
        data_rectus_l = [[male_tonus[i]["Rectus_L"]], [female_tonus[i]["Rectus_L"]],
                         [over30_tonus[i]["Rectus_L"]], [EA_over_tonus[i]["Rectus_L"]],
                         [EA_more_tonus[i]["Rectus_L"]]]
        plt.boxplot(data_rectus_l, positions = [1,2,3,4,5], widths = 0.6, showmeans= True)
        plt.xticks([1,2,3,4,5], ("Male", "Female", "Over 30", "EA over30", "EA more evolve"), fontsize=8)
        plt.ylim([0, 140])
        plot1.set_title("Rectus Left", fontsize=12)

        plot2 = plt.subplot2grid((4, 2), (0, 1))
        data_rectus_R = [[male_tonus[i]["Rectus_R"]], [female_tonus[i]["Rectus_R"]], [over30_tonus[i]["Rectus_R"]],
                         [EA_over_tonus[i]["Rectus_R"]], [EA_more_tonus[i]["Rectus_R"]]]
        plt.boxplot(data_rectus_R, positions = [1,2,3,4,5], widths=0.6, showmeans= True)
        plt.xticks([1, 2, 3, 4, 5], ("Male", "Female", "Over 30", "EA over30", "EA more evolve"),
                   fontsize=8)
        plt.ylim([0, 140])
        plot2.set_title("Rectus Right", fontsize=12)

        plot3 = plt.subplot2grid((4, 2), (1, 0))
        data_obliques_l = [[male_tonus[i]["Obliques_L"]], [female_tonus[i]["Obliques_L"]],
                           [over30_tonus[i]["Obliques_L"]],[EA_over_tonus[i]["Rectus_R"]], [EA_more_tonus[i]["Obliques_L"]]]
        plt.boxplot(data_obliques_l, positions = [1,2,3,4,5], widths=0.6, showmeans=True)
        plt.xticks([1, 2, 3, 4, 5], ("Male", "Female", "Over 30", "EA over30", "EA more evolve"),
                   fontsize=8)
        plt.ylim([0, 140])
        plot3.set_title("Obliques Left", fontsize=12)

        plot4 = plt.subplot2grid((4, 2), (1, 1))
        data_obliques_r = [[male_tonus[i]["Obliques_R"]], [female_tonus[i]["Obliques_R"]],
                           [over30_tonus[i]["Obliques_R"]],
                           [EA_over_tonus[i]["Obliques_R"]], [EA_more_tonus[i]["Obliques_R"]]]
        plt.boxplot(data_obliques_r, positions = [1,2,3,4,5], widths=0.6, showmeans=True)
        plt.xticks([1, 2, 3, 4, 5], ("Male", "Female", "Over 30", "EA over30", "EA more evolve"),
                   fontsize=8)
        plt.ylim([0, 140])
        plot4.set_title("Obliques Right", fontsize=12)

        plot5 = plt.subplot2grid((4, 2), (2, 0))
        data_ilicostalis_l = [[male_tonus[i]["Ilicostalis_L"]], [female_tonus[i]["Ilicostalis_L"]],
                              [over30_tonus[i]["Ilicostalis_L"]],
                              [EA_over_tonus[i]["Ilicostalis_L"]], [EA_more_tonus[i]["Ilicostalis_L"]]]
        plt.boxplot(data_ilicostalis_l, positions = [1,2,3,4,5], widths=0.6, showmeans=True)
        plt.xticks([1, 2, 3, 4, 5], ("Male", "Female", "Over 30", "EA over30", "EA under30", "EA more evolve"),
                   fontsize=8)
        plt.ylim([0, 140])
        plot5.set_title("Iliocostalis Left", fontsize=12)

        plot6 = plt.subplot2grid((4, 2), (2, 1))
        data_ilicostalis_r = [[male_tonus[i]["Ilicostalis_R"]], [female_tonus[i]["Ilicostalis_R"]],
                              [over30_tonus[i]["Ilicostalis_R"]],
                              [EA_over_tonus[i]["Ilicostalis_R"]], [EA_more_tonus[i]["Ilicostalis_R"]]]
        plt.boxplot(data_ilicostalis_r, positions = [1,2,3,4,5], widths=0.6, showmeans=True)
        plt.xticks([1, 2, 3, 4, 5], ("Male", "Female", "Over 30", "EA over30", "EA more evolve"),
                   fontsize=8)
        plt.ylim([0, 140])
        plot6.set_title("Iliocostalis Right", fontsize=12)

        plot7 = plt.subplot2grid((4, 2), (3, 0))
        data_multi_l = [[male_tonus[i]["Multifidus_L"]], [female_tonus[i]["Multifidus_L"]],
                        [over30_tonus[i]["Multifidus_L"]],
                        [EA_over_tonus[i]["Multifidus_L"]], [EA_more_tonus[i]["Multifidus_L"]]]
        plt.boxplot(data_multi_l, positions = [1,2,3,4,5], widths=0.6, showmeans=True)
        plt.xticks([1, 2, 3, 4, 5], ("Male", "Female", "Over 30", "EA over30", "EA more evolve"),
                   fontsize=8)
        plt.ylim([0, 140])
        plot7.set_title("Multifidus Left", fontsize=12)

        plot8 = plt.subplot2grid((4, 2), (3, 1))
        data_multi_r = [[male_tonus[i]["Multifidus_R"]], [female_tonus[i]["Multifidus_R"]],
                        [over30_tonus[i]["Multifidus_R"]],
                        [EA_over_tonus[i]["Multifidus_R"]], [EA_more_tonus[i]["Multifidus_R"]]]
        plt.boxplot(data_multi_r, positions = [1,2,3,4,5], widths=0.6, showmeans=True)
        plt.xticks([1, 2, 3, 4, 5], ("Male", "Female", "Over 30", "EA over30", "EA more evolve"),
                   fontsize=8)
        plt.ylim([0, 140])
        plot8.set_title("Multifidus Right", fontsize=12)

        plt.subplots_adjust(top=0.90, bottom=0.10, left=0.12, right=0.90, wspace=0.44, hspace=0.45)
        plt.show()
        pp.savefig(fig)
    pp.close()

def mean_task_vs_relax(mean_over30, mean_male, mean_female, mean_EA_over, mean_EA_more,mean_rest_over30, mean_rest_male, mean_rest_female, mean_rest_EA_over, mean_rest_EA_more):

    l = 0
    pp = PdfPages('RelaxVSTask_MeanValues_Analysis_Healthy_EA.pdf')

    for i in mean_over30:

        fig = plt.figure(l)
        l = l + 1
        fig.suptitle("Relax VS Task (Mean Value) - " + str(i), fontsize=21)

        plot1 = plt.subplot2grid((2, 4), (0, 0))

        plt.plot(mean_rest_over30["Rectus_L"], mean_over30[i]["Rectus_L"], 'ro', color='blue', label = "Over 30")
        plt.plot(mean_rest_male["Rectus_L"], mean_male[i]["Rectus_L"], 'ro', color='yellow', label = "Male 20-30")
        plt.plot(mean_rest_female["Rectus_L"], mean_female[i]["Rectus_L"], 'ro', color='green', label = "Female 20-30")
        plt.plot(mean_rest_EA_over["Rectus_L"], mean_EA_over[i]["Rectus_L"], 'ro', color='red', label = "EA Over 30")
        plt.plot(mean_rest_EA_more["Rectus_L"], mean_EA_more[i]["Rectus_L"], 'ro', color='magenta', label="EA evolved")

        plt.xlabel('Rest mean Values', fontsize=10)
        plt.xlim([0,50])
        plt.ylim([0, 50])
        plt.ylabel('Rectus Left mean Values', fontsize=10)
        plt.legend(bbox_to_anchor=(1.0, 1.0), loc=2, borderaxespad=0., fontsize=10)
        plot1.set_title("Rectus Left", fontsize=12)

        plot2 = plt.subplot2grid((2, 4), (0, 1))

        plt.plot(mean_rest_over30["Rectus_R"], mean_over30[i]["Rectus_R"], 'ro', color='blue', label="Over 30")
        plt.plot(mean_rest_male["Rectus_R"], mean_male[i]["Rectus_R"], 'ro', color='yellow', label="Male 20-30")
        plt.plot(mean_rest_female["Rectus_R"], mean_female[i]["Rectus_R"], 'ro', color='green', label="Female 20-30")
        plt.plot(mean_rest_EA_over["Rectus_R"], mean_EA_over[i]["Rectus_R"], 'ro', color='red', label="EA Over 30")
        plt.plot(mean_rest_EA_more["Rectus_R"], mean_EA_more[i]["Rectus_R"], 'ro', color='magenta', label="EA evolved")

        plt.xlabel('Rest mean Values', fontsize=10)
        plt.ylabel('Rectus Right mean Values', fontsize=10)
        plt.xlim([0, 50])
        plt.ylim([0, 50])
        plt.legend(bbox_to_anchor=(1.0, 1.0), loc=2, borderaxespad=0., fontsize=10)
        plot2.set_title("Rectus Right", fontsize=12)

        plot3 = plt.subplot2grid((2, 4), (0, 2))

        plt.plot(mean_rest_over30["Obliques_L"], mean_over30[i]["Obliques_L"], 'ro', color='blue', label="Over 30")
        plt.plot(mean_rest_male["Obliques_L"], mean_male[i]["Obliques_L"], 'ro', color='yellow', label="Male 20-30")
        plt.plot(mean_rest_female["Obliques_L"], mean_female[i]["Obliques_L"], 'ro', color='green', label="Female 20-30")
        plt.plot(mean_rest_EA_over["Obliques_L"], mean_EA_over[i]["Obliques_L"], 'ro', color='red', label="EA Over 30")
        plt.plot(mean_rest_EA_more["Obliques_L"], mean_EA_more[i]["Obliques_L"], 'ro', color='magenta', label="EA evolved")

        plt.xlabel('Rest mean Values', fontsize=10)
        plt.ylabel('Obliques Left mean Values', fontsize=10)
        plt.xlim([0, 50])
        plt.ylim([0, 50])
        plt.legend(bbox_to_anchor=(1.0, 1.0), loc=2, borderaxespad=0., fontsize=10)
        plot3.set_title("Obliques Left", fontsize=12)

        plot4 = plt.subplot2grid((2, 4), (0, 3))
        plt.plot(mean_rest_over30["Obliques_R"], mean_over30[i]["Obliques_R"], 'ro', color='blue', label="Over 30")
        plt.plot(mean_rest_male["Obliques_R"], mean_male[i]["Obliques_R"], 'ro', color='yellow', label="Male 20-30")
        plt.plot(mean_rest_female["Obliques_R"], mean_female[i]["Obliques_R"], 'ro', color='green',
                 label="Female 20-30")
        plt.plot(mean_rest_EA_over["Obliques_R"], mean_EA_over[i]["Obliques_R"], 'ro', color='red', label="EA Over 30")
        plt.plot(mean_rest_EA_more["Obliques_R"], mean_EA_more[i]["Obliques_R"], 'ro', color='magenta',
                 label="EA evolved")

        plt.xlabel('Rest mean Values', fontsize=10)
        plt.ylabel('Obliques Right mean Values', fontsize=10)
        plt.xlim([0, 50])
        plt.ylim([0, 50])
        plt.legend(bbox_to_anchor=(1.0, 1.0), loc=2, borderaxespad=0., fontsize=10)
        plot4.set_title("Obliques Right", fontsize=12)

        plot5 = plt.subplot2grid((2, 4), (1, 0))

        plt.plot(mean_rest_over30["Ilicostalis_L"], mean_over30[i]["Ilicostalis_L"], 'ro', color='blue', label="Over 30")
        plt.plot(mean_rest_male["Ilicostalis_L"], mean_male[i]["Ilicostalis_L"], 'ro', color='yellow', label="Male 20-30")
        plt.plot(mean_rest_female["Ilicostalis_L"], mean_female[i]["Ilicostalis_L"], 'ro', color='green',
                 label="Female 20-30")
        plt.plot(mean_rest_EA_over["Ilicostalis_L"], mean_EA_over[i]["Ilicostalis_L"], 'ro', color='red', label="EA Over 30")
        plt.plot(mean_rest_EA_more["Ilicostalis_L"], mean_EA_more[i]["Ilicostalis_L"], 'ro', color='magenta',
                 label="EA evolved")

        plt.xlabel('Rest mean Values', fontsize=10)
        plt.ylabel('Iliocostalis Left mean Values', fontsize=10)
        plt.xlim([0, 50])
        plt.ylim([0, 50])
        plt.legend(bbox_to_anchor=(1.0, 1.0), loc=2, borderaxespad=0., fontsize=10)
        plot5.set_title("Iliocostalis Left", fontsize=12)

        plot6 = plt.subplot2grid((2, 4), (1, 1))

        plt.plot(mean_rest_over30["Ilicostalis_R"], mean_over30[i]["Ilicostalis_R"], 'ro', color='blue',
                 label="Over 30")
        plt.plot(mean_rest_male["Ilicostalis_R"], mean_male[i]["Ilicostalis_R"], 'ro', color='yellow',
                 label="Male 20-30")
        plt.plot(mean_rest_female["Ilicostalis_R"], mean_female[i]["Ilicostalis_R"], 'ro', color='green',
                 label="Female 20-30")
        plt.plot(mean_rest_EA_over["Ilicostalis_R"], mean_EA_over[i]["Ilicostalis_R"], 'ro', color='red',
                 label="EA Over 30")
        plt.plot(mean_rest_EA_more["Ilicostalis_R"], mean_EA_more[i]["Ilicostalis_R"], 'ro', color='magenta',
                 label="EA evolved")

        plt.xlabel('Rest mean Values', fontsize=10)
        plt.ylabel('Iliocostalis Right mean Values', fontsize=10)
        plt.xlim([0, 50])
        plt.ylim([0, 50])
        plt.legend(bbox_to_anchor=(1.0, 1.0), loc=2, borderaxespad=0., fontsize=10)
        plot6.set_title("Iliocostalis Right", fontsize=12)

        plot7 = plt.subplot2grid((2, 4), (1, 2))

        plt.plot(mean_rest_over30["Multifidus_L"], mean_over30[i]["Multifidus_L"], 'ro', color='blue',
                 label="Over 30")
        plt.plot(mean_rest_male["Multifidus_L"], mean_male[i]["Multifidus_L"], 'ro', color='yellow',
                 label="Male 20-30")
        plt.plot(mean_rest_female["Multifidus_L"], mean_female[i]["Multifidus_L"], 'ro', color='green',
                 label="Female 20-30")
        plt.plot(mean_rest_EA_over["Multifidus_L"], mean_EA_over[i]["Multifidus_L"], 'ro', color='red',
                 label="EA Over 30")
        plt.plot(mean_rest_EA_more["Multifidus_L"], mean_EA_more[i]["Multifidus_L"], 'ro', color='magenta',
                 label="EA evolved")

        plt.xlabel('Rest mean Values', fontsize=10)
        plt.ylabel('Multifidus Left mean Values', fontsize=10)
        plt.xlim([0, 50])
        plt.ylim([0, 50])
        plt.legend(bbox_to_anchor=(1.0, 1.0), loc=2, borderaxespad=0., fontsize=10)
        plot7.set_title("Multifidus Left", fontsize=12)

        plot8 = plt.subplot2grid((2, 4), (1, 3))

        plt.plot(mean_rest_over30["Multifidus_R"], mean_over30[i]["Multifidus_R"], 'ro', color='blue',
                 label="Over 30")
        plt.plot(mean_rest_male["Multifidus_R"], mean_male[i]["Multifidus_R"], 'ro', color='yellow',
                 label="Male 20-30")
        plt.plot(mean_rest_female["Multifidus_R"], mean_female[i]["Multifidus_R"], 'ro', color='green',
                 label="Female 20-30")
        plt.plot(mean_rest_EA_over["Multifidus_R"], mean_EA_over[i]["Multifidus_R"], 'ro', color='red',
                 label="EA Over 30")
        plt.plot(mean_rest_EA_more["Multifidus_R"], mean_EA_more[i]["Multifidus_R"], 'ro', color='magenta',
                 label="EA evolved")

        plt.xlabel('Rest mean Values', fontsize=10)
        plt.ylabel('Multifidus Right mean Values', fontsize=10)
        plt.xlim([0, 50])
        plt.ylim([0, 50])
        plt.legend(bbox_to_anchor=(1.0, 1.0), loc=2, borderaxespad=0., fontsize=10)
        plot8.set_title("Multifidus Right", fontsize=12)

        plt.subplots_adjust(top=0.90, bottom=0.10, left=0.07, right=0.86, wspace=0.95, hspace=0.68)
        plt.show()
        pp.savefig(fig)
    pp.close()

    def EMG_freq_back_boxplot(over30_freq, male_freq, female_freq, EA_over_freq, EA_more_freq):
        l = 0
        pp = PdfPages('EMG_Freqs_Back_BoxPlot_Analysis_Healthy_EA.pdf')

        for i in over30_freq:
            fig = plt.figure(l)
            l = l + 1
            fig.suptitle(str(i) + " - EMG Frequencies Back Muscles", fontsize=21)

            row_peak_mean = 0
            col_peak_mean = 0
            row_80_median = 2
            col_80_median = 0

            colors = ['blue', 'green']
            color_index = 0
            plt.figtext(0.29, 0.91, "Peak Frequency", horizontalalignment='center', multialignment='center', size=18,
                        clip_on=True)
            plt.figtext(0.71, 0.91, "Mean Frequency", horizontalalignment='center', multialignment='center', size=18,
                        clip_on=True)
            plt.figtext(0.29, 0.46, "80% Frequency", horizontalalignment='center', multialignment='center', size=18,
                        clip_on=True)
            plt.figtext(0.71, 0.46, "Median Frequency", horizontalalignment='center', multialignment='center', size=18,
                        clip_on=True)

            for freq in over30_freq[i]["Rectus_L"]:
                if freq == "Peak" or freq == "Mean":
                    row_l = row_peak_mean
                    col_l = col_peak_mean
                    row_r = row_peak_mean
                    col_r = col_peak_mean + 1

                    for muscle in over30_freq[i]:
                        if muscle == "Ilicostalis_L" or muscle == "Multi_L":
                            plot = plt.subplot2grid((4, 4), (row_l, col_l))
                            data = [[male_freq[i][muscle][freq]], [female_freq[i][muscle][freq]],
                                    [over30_freq[i][muscle][freq]],
                                    [EA_over_freq[i][muscle][freq]],
                                    [EA_more_freq[i][muscle][freq]]]
                            box = plt.boxplot(data, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
                            plt.xticks([1, 2, 3, 4, 5],
                                       ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                                       fontsize=7)
                            plot.set_title(str(muscle), fontsize=12, style='italic')
                            row_l = row_l + 1

                        if muscle == "Ilicostalis_R" or muscle == "Multi_R":
                            plot = plt.subplot2grid((4, 4), (row_r, col_r))
                            data = [[male_freq[i][muscle][freq]], [female_freq[i][muscle][freq]],
                                    [over30_freq[i][muscle][freq]],
                                    [EA_over_freq[i][muscle][freq]],
                                    [EA_more_freq[i][muscle][freq]]]
                            box = plt.boxplot(data, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
                            plt.xticks([1, 2, 3, 4, 5],
                                       ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                                       fontsize=7)
                            plot.set_title(str(muscle), fontsize=12, style='italic')
                            row_r = row_r + 1

                    col_peak_mean = col_peak_mean + 2
                    color_index = color_index + 1

                if freq == "Median" or freq == "80_freq":
                    row_l = row_80_median
                    col_l = col_80_median
                    row_r = row_80_median
                    col_r = col_80_median + 1

                    for muscle in over30_freq[i]:
                        if muscle == "Ilicostalis_L" or muscle == "Multi_L":
                            plot = plt.subplot2grid((4, 4), (row_l, col_l))
                            data = [[male_freq[i][muscle][freq]], [female_freq[i][muscle][freq]],
                                    [over30_freq[i][muscle][freq]],
                                    [EA_over_freq[i][muscle][freq]],
                                    [EA_more_freq[i][muscle][freq]]]
                            box = plt.boxplot(data, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
                            plt.xticks([1, 2, 3, 4, 5],
                                       ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                                       fontsize=7)
                            plot.set_title(str(muscle), fontsize=12, style='italic')
                            row_l = row_l + 1

                        if muscle == "Ilicostalis_R" or muscle == "Multi_R":
                            plot = plt.subplot2grid((4, 4), (row_r, col_r))
                            data = [[male_freq[i][muscle][freq]], [female_freq[i][muscle][freq]],
                                    [over30_freq[i][muscle][freq]],
                                    [EA_over_freq[i][muscle][freq]],
                                    [EA_more_freq[i][muscle][freq]]]
                            box = plt.boxplot(data, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
                            plt.xticks([1, 2, 3, 4, 5],
                                       ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                                       fontsize=7)
                            plt.xticks([1, 2, 3, 4, 5],
                                       ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                                       fontsize=7)
                            plot.set_title(str(muscle), fontsize=12, style='italic')
                            row_r = row_r + 1

                    col_80_median = col_80_median + 2
                    color_index = color_index + 1

            plt.subplots_adjust(top=0.89, bottom=0.05, left=0.12, right=0.90, wspace=0.56, hspace=0.67)
            plt.show()
            pp.savefig(fig)
        pp.close()

def EMG_freq_rest_back_boxplot(over30_freq, male_freq, female_freq, EA_over_freq, EA_more_freq):
    l = 0
    pp = PdfPages('EMG_Freqs_Rest_Back_BoxPlot_Analysis_Healthy_EA.pdf')

    fig = plt.figure(l)
    fig.suptitle("EMG Frequencies Back Muscles during Rest", fontsize=21)

    row_peak_mean = 0
    col_peak_mean = 0
    row_80_median = 2
    col_80_median = 0

    colors = ['blue', 'green']
    color_index = 0
    plt.figtext(0.29, 0.91, "Peak Frequency", horizontalalignment = 'center', multialignment = 'center', size = 18,
                    clip_on = True)
    plt.figtext(0.71, 0.91, "Mean Frequency", horizontalalignment='center', multialignment='center', size=18,
                    clip_on=True)
    plt.figtext(0.29, 0.46, "80% Frequency", horizontalalignment='center', multialignment='center', size=18,
                    clip_on=True)
    plt.figtext(0.71, 0.46, "Median Frequency", horizontalalignment='center', multialignment='center', size=18,
                    clip_on=True)

    for freq in over30_freq["Rectus_L"]:
        if freq == "Peak" or freq == "Mean":
            row_l = row_peak_mean
            col_l = col_peak_mean
            row_r = row_peak_mean
            col_r = col_peak_mean + 1

            for muscle in over30_freq:
                if muscle == "Ilicostalis_L" or muscle == "Multi_L":
                    plot = plt.subplot2grid((4, 4), (row_l, col_l))
                    data = [[male_freq[muscle][freq]], [female_freq[muscle][freq]],
                                [over30_freq[muscle][freq]],
                                [EA_over_freq[muscle][freq]],
                                [EA_more_freq[muscle][freq]]]
                    box = plt.boxplot(data, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
                    plt.xticks([1, 2, 3, 4, 5],
                                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                                   fontsize=7)
                    plot.set_title(str(muscle), fontsize=12, style = 'italic')
                    row_l = row_l + 1

                if muscle == "Ilicostalis_R" or muscle == "Multi_R":
                    plot = plt.subplot2grid((4, 4), (row_r, col_r))
                    data = [[male_freq[muscle][freq]], [female_freq[muscle][freq]],
                                [over30_freq[muscle][freq]],
                                [EA_over_freq[muscle][freq]],
                                [EA_more_freq[muscle][freq]]]
                    box = plt.boxplot(data, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
                    plt.xticks([1, 2, 3, 4, 5],
                                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                                   fontsize=7)
                    plot.set_title(str(muscle), fontsize=12, style='italic')
                    row_r = row_r + 1

            col_peak_mean = col_peak_mean + 2
            color_index = color_index + 1

        if freq == "Median" or freq == "80_freq":
            row_l = row_80_median
            col_l = col_80_median
            row_r = row_80_median
            col_r = col_80_median + 1

            for muscle in over30_freq:
                if muscle == "Ilicostalis_L" or muscle == "Multi_L":
                    plot = plt.subplot2grid((4, 4), (row_l, col_l))
                    data = [[male_freq[muscle][freq]], [female_freq[muscle][freq]],
                                [over30_freq[muscle][freq]],
                                [EA_over_freq[muscle][freq]],
                                [EA_more_freq[muscle][freq]]]
                    box = plt.boxplot(data, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
                    plt.xticks([1, 2, 3, 4, 5],
                                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                                   fontsize=7)
                    plot.set_title(str(muscle), fontsize=12, style='italic')
                    row_l = row_l + 1

                if muscle == "Ilicostalis_R" or muscle == "Multi_R":
                    plot = plt.subplot2grid((4, 4), (row_r, col_r))
                    data = [[male_freq[muscle][freq]], [female_freq[muscle][freq]],
                                [over30_freq[muscle][freq]],
                                [EA_over_freq[muscle][freq]],
                                [EA_more_freq[muscle][freq]]]
                    box = plt.boxplot(data, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
                    plt.xticks([1, 2, 3, 4, 5],
                                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                                   fontsize=7)
                    plt.xticks([1, 2, 3, 4, 5],
                                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                                   fontsize=7)
                    plot.set_title(str(muscle), fontsize=12, style='italic')
                    row_r = row_r + 1

            col_80_median = col_80_median + 2
            color_index = color_index + 1

    plt.subplots_adjust(top=0.89, bottom=0.05, left=0.12, right=0.90, wspace=0.56, hspace=0.67)
    plt.show()
    pp.savefig(fig)
    pp.close()

def EMG_freq_rest_front_boxplot(over30_freq, male_freq, female_freq, EA_over_freq, EA_more_freq):
    l = 0
    pp = PdfPages('EMG_Freqs_Rest_Front_BoxPlot_Analysis.pdf')


    fig = plt.figure(l)
    fig.suptitle("EMG Frequencies Front Muscles during Rest", fontsize=21)

    row_peak_mean = 0
    col_peak_mean = 0
    row_80_median = 2
    col_80_median = 0

    colors = ['blue', 'green']
    color_index = 0
    plt.figtext(0.29, 0.91, "Peak Frequency", horizontalalignment='center', multialignment='center', size=18,
                    clip_on=True)
    plt.figtext(0.71, 0.91, "Mean Frequency", horizontalalignment='center', multialignment='center', size=18,
                    clip_on=True)
    plt.figtext(0.29, 0.46, "80% Frequency", horizontalalignment='center', multialignment='center', size=18,
                    clip_on=True)
    plt.figtext(0.71, 0.46, "Median Frequency", horizontalalignment='center', multialignment='center', size=18,
                    clip_on=True)

    for freq in over30_freq["Rectus_L"]:
        if freq == "Peak" or freq == "Mean":
            row_l = row_peak_mean
            col_l = col_peak_mean
            row_r = row_peak_mean
            col_r = col_peak_mean + 1

            for muscle in over30_freq:
                if muscle == "Rectus_L" or muscle == "Obliques_L":
                    plot = plt.subplot2grid((4, 4), (row_l, col_l))
                    data = [[male_freq[muscle][freq]], [female_freq[muscle][freq]],
                                [over30_freq[muscle][freq]],
                                [EA_over_freq[muscle][freq]],
                                [EA_more_freq[muscle][freq]]]
                    box = plt.boxplot(data, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
                    plt.xticks([1, 2, 3, 4, 5],
                                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                                   fontsize=7)
                    plot.set_title(str(muscle), fontsize=12, style = 'italic')
                    row_l = row_l + 1

                if muscle == "Rectus_R" or muscle == "Obliques_R":
                    plot = plt.subplot2grid((4, 4), (row_r, col_r))
                    data = [[male_freq[muscle][freq]], [female_freq[muscle][freq]],
                                [over30_freq[muscle][freq]],
                                [EA_over_freq[muscle][freq]],
                                [EA_more_freq[muscle][freq]]]
                    box = plt.boxplot(data, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
                    plt.xticks([1, 2, 3, 4, 5],
                                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                                   fontsize=7)
                    plot.set_title(str(muscle), fontsize=12, style='italic')
                    row_r = row_r + 1

            col_peak_mean = col_peak_mean + 2
            color_index = color_index + 1

        if freq == "Median" or freq == "80_freq":
            row_l = row_80_median
            col_l = col_80_median
            row_r = row_80_median
            col_r = col_80_median + 1

            for muscle in over30_freq:
                if muscle == "Rectus_L" or muscle == "Obliques_L":
                    plot = plt.subplot2grid((4, 4), (row_l, col_l))
                    data = [[male_freq[muscle][freq]], [female_freq[muscle][freq]],
                                [over30_freq[muscle][freq]],
                                [EA_over_freq[muscle][freq]],
                                [EA_more_freq[muscle][freq]]]
                    box = plt.boxplot(data, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
                    plt.xticks([1, 2, 3, 4, 5],
                                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                                   fontsize=7)
                    plot.set_title(str(muscle), fontsize=12, style='italic')
                    row_l = row_l + 1

                if muscle == "Rectus_R" or muscle == "Obliques_R":
                    plot = plt.subplot2grid((4, 4), (row_r, col_r))
                    data = [[male_freq[muscle][freq]], [female_freq[muscle][freq]],
                                [over30_freq[muscle][freq]],
                                [EA_over_freq[muscle][freq]],
                                [EA_more_freq[muscle][freq]]]
                    box = plt.boxplot(data, positions=[1, 2, 3, 4, 5], widths=0.6, showmeans=True)
                    plt.xticks([1, 2, 3, 4, 5],
                                   ("Male", "Female", ">30", "EA >30", "EA\n envolved"),
                                   fontsize=7)
                    plot.set_title(str(muscle), fontsize=12, style='italic')
                    row_r = row_r + 1

            col_80_median = col_80_median + 2
            color_index = color_index + 1

    plt.subplots_adjust(top=0.89, bottom=0.05, left=0.12, right=0.90, wspace=0.56, hspace=0.67)
    plt.show()
    pp.savefig(fig)
    pp.close()


def EMG_rest_values_boxplot(over30_tonus, male_tonus, female_tonus, EA_over_tonus, EA_more_tonus):
    l = 0
    pp = PdfPages('PeakValue_Muscular_Rest_BoxPlot_Analysis_Healthy_EA.pdf')

    fig = plt.figure(l)
    fig.suptitle("Max Value of normalized EMG during Rest", fontsize=21)

    plot1 = plt.subplot2grid((4, 2), (0, 0))
    data_rectus_l = [[male_tonus["Rectus_L"]], [female_tonus["Rectus_L"]],
                         [over30_tonus["Rectus_L"]], [EA_over_tonus["Rectus_L"]],
                         [EA_more_tonus["Rectus_L"]]]
    plt.boxplot(data_rectus_l, positions = [1,2,3,4,5], widths = 0.6, showmeans= True)
    plt.xticks([1,2,3,4,5], ("Male", "Female", "Over 30", "EA over30", "EA more evolve"), fontsize=8)
    plt.ylim([0,50])
    plot1.set_title("Rectus Left", fontsize=12)

    plot2 = plt.subplot2grid((4, 2), (0, 1))
    data_rectus_R = [[male_tonus["Rectus_R"]], [female_tonus["Rectus_R"]], [over30_tonus["Rectus_R"]],
                         [EA_over_tonus["Rectus_R"]], [EA_more_tonus["Rectus_R"]]]
    plt.boxplot(data_rectus_R, positions = [1,2,3,4,5], widths=0.6, showmeans= True)
    plt.xticks([1, 2, 3, 4, 5], ("Male", "Female", "Over 30", "EA over30", "EA more evolve"),
                   fontsize=8)
    plt.ylim([0, 50])
    plot2.set_title("Rectus Right", fontsize=12)

    plot3 = plt.subplot2grid((4, 2), (1, 0))
    data_obliques_l = [[male_tonus["Obliques_L"]], [female_tonus["Obliques_L"]],
                           [over30_tonus["Obliques_L"]],[EA_over_tonus["Rectus_R"]], [EA_more_tonus["Obliques_L"]]]
    plt.boxplot(data_obliques_l, positions = [1,2,3,4,5], widths=0.6, showmeans=True)
    plt.xticks([1, 2, 3, 4, 5], ("Male", "Female", "Over 30", "EA over30", "EA more evolve"),
                   fontsize=8)
    plt.ylim([0, 50])
    plot3.set_title("Obliques Left", fontsize=12)

    plot4 = plt.subplot2grid((4, 2), (1, 1))
    data_obliques_r = [[male_tonus["Obliques_R"]], [female_tonus["Obliques_R"]],
                           [over30_tonus["Obliques_R"]],
                           [EA_over_tonus["Obliques_R"]], [EA_more_tonus["Obliques_R"]]]
    plt.boxplot(data_obliques_r, positions = [1,2,3,4,5], widths=0.6, showmeans=True)
    plt.xticks([1, 2, 3, 4, 5], ("Male", "Female", "Over 30", "EA over30", "EA more evolve"),
                   fontsize=8)
    plt.ylim([0, 50])
    plot4.set_title("Obliques Right", fontsize=12)

    plot5 = plt.subplot2grid((4, 2), (2, 0))
    data_ilicostalis_l = [[male_tonus["Ilicostalis_L"]], [female_tonus["Ilicostalis_L"]],
                              [over30_tonus["Ilicostalis_L"]],
                              [EA_over_tonus["Ilicostalis_L"]], [EA_more_tonus["Ilicostalis_L"]]]
    plt.boxplot(data_ilicostalis_l, positions = [1,2,3,4,5], widths=0.6, showmeans=True)
    plt.xticks([1, 2, 3, 4, 5], ("Male", "Female", "Over 30", "EA over30", "EA under30", "EA more evolve"),
                   fontsize=8)
    plt.ylim([0, 50])
    plot5.set_title("Iliocostalis Left", fontsize=12)

    plot6 = plt.subplot2grid((4, 2), (2, 1))
    data_ilicostalis_r = [[male_tonus["Ilicostalis_R"]], [female_tonus["Ilicostalis_R"]],
                              [over30_tonus["Ilicostalis_R"]],
                              [EA_over_tonus["Ilicostalis_R"]], [EA_more_tonus["Ilicostalis_R"]]]
    plt.boxplot(data_ilicostalis_r, positions = [1,2,3,4,5], widths=0.6, showmeans=True)
    plt.xticks([1, 2, 3, 4, 5], ("Male", "Female", "Over 30", "EA over30", "EA more evolve"),
                   fontsize=8)
    plt.ylim([0, 50])
    plot6.set_title("Iliocostalis Right", fontsize=12)

    plot7 = plt.subplot2grid((4, 2), (3, 0))
    data_multi_l = [[male_tonus["Multifidus_L"]], [female_tonus["Multifidus_L"]],
                        [over30_tonus["Multifidus_L"]],
                        [EA_over_tonus["Multifidus_L"]], [EA_more_tonus["Multifidus_L"]]]
    plt.boxplot(data_multi_l, positions = [1,2,3,4,5], widths=0.6, showmeans=True)
    plt.xticks([1, 2, 3, 4, 5], ("Male", "Female", "Over 30", "EA over30", "EA more evolve"),
                   fontsize=8)
    plt.ylim([0, 50])
    plot7.set_title("Multifidus Left", fontsize=12)

    plot8 = plt.subplot2grid((4, 2), (3, 1))
    data_multi_r = [[male_tonus["Multifidus_R"]], [female_tonus["Multifidus_R"]],
                        [over30_tonus["Multifidus_R"]],
                        [EA_over_tonus["Multifidus_R"]], [EA_more_tonus["Multifidus_R"]]]
    plt.boxplot(data_multi_r, positions = [1,2,3,4,5], widths=0.6, showmeans=True)
    plt.xticks([1, 2, 3, 4, 5], ("Male", "Female", "Over 30", "EA over30", "EA more evolve"),
                   fontsize=8)
    plt.ylim([0, 50])
    plot8.set_title("Multifidus Right", fontsize=12)

    plt.subplots_adjust(top=0.90, bottom=0.10, left=0.12, right=0.90, wspace=0.44, hspace=0.45)
    plt.show()
    pp.savefig(fig)
    pp.close()

def EMG_rest_meanvalues_boxplot(over30_tonus, male_tonus, female_tonus, EA_over_tonus, EA_more_tonus):
    l = 0
    pp = PdfPages('MeanValue_Muscular_Rest_BoxPlot_Analysis_Healthy_EA.pdf')

    fig = plt.figure(l)
    fig.suptitle("Mean Value of normalized EMG during Rest", fontsize=21)

    plot1 = plt.subplot2grid((4, 2), (0, 0))
    data_rectus_l = [[male_tonus["Rectus_L"]], [female_tonus["Rectus_L"]],
                         [over30_tonus["Rectus_L"]], [EA_over_tonus["Rectus_L"]],
                         [EA_more_tonus["Rectus_L"]]]
    plt.boxplot(data_rectus_l, positions = [1,2,3,4,5], widths = 0.6, showmeans= True)
    plt.xticks([1,2,3,4,5], ("Male", "Female", "Over 30", "EA over30", "EA more evolve"), fontsize=8)
    plt.ylim([0,30])
    plot1.set_title("Rectus Left", fontsize=12)

    plot2 = plt.subplot2grid((4, 2), (0, 1))
    data_rectus_R = [[male_tonus["Rectus_R"]], [female_tonus["Rectus_R"]], [over30_tonus["Rectus_R"]],
                         [EA_over_tonus["Rectus_R"]], [EA_more_tonus["Rectus_R"]]]
    plt.boxplot(data_rectus_R, positions = [1,2,3,4,5], widths=0.6, showmeans= True)
    plt.xticks([1, 2, 3, 4, 5], ("Male", "Female", "Over 30", "EA over30", "EA more evolve"),
                   fontsize=8)
    plt.ylim([0, 30])
    plot2.set_title("Rectus Right", fontsize=12)

    plot3 = plt.subplot2grid((4, 2), (1, 0))
    data_obliques_l = [[male_tonus["Obliques_L"]], [female_tonus["Obliques_L"]],
                           [over30_tonus["Obliques_L"]],[EA_over_tonus["Rectus_R"]], [EA_more_tonus["Obliques_L"]]]
    plt.boxplot(data_obliques_l, positions = [1,2,3,4,5], widths=0.6, showmeans=True)
    plt.xticks([1, 2, 3, 4, 5], ("Male", "Female", "Over 30", "EA over30", "EA more evolve"),
                   fontsize=8)
    plt.ylim([0, 30])
    plot3.set_title("Obliques Left", fontsize=12)

    plot4 = plt.subplot2grid((4, 2), (1, 1))
    data_obliques_r = [[male_tonus["Obliques_R"]], [female_tonus["Obliques_R"]],
                           [over30_tonus["Obliques_R"]],
                           [EA_over_tonus["Obliques_R"]], [EA_more_tonus["Obliques_R"]]]
    plt.boxplot(data_obliques_r, positions = [1,2,3,4,5], widths=0.6, showmeans=True)
    plt.xticks([1, 2, 3, 4, 5], ("Male", "Female", "Over 30", "EA over30", "EA more evolve"),
                   fontsize=8)
    plt.ylim([0, 30])
    plot4.set_title("Obliques Right", fontsize=12)

    plot5 = plt.subplot2grid((4, 2), (2, 0))
    data_ilicostalis_l = [[male_tonus["Ilicostalis_L"]], [female_tonus["Ilicostalis_L"]],
                              [over30_tonus["Ilicostalis_L"]],
                              [EA_over_tonus["Ilicostalis_L"]], [EA_more_tonus["Ilicostalis_L"]]]
    plt.boxplot(data_ilicostalis_l, positions = [1,2,3,4,5], widths=0.6, showmeans=True)
    plt.xticks([1, 2, 3, 4, 5], ("Male", "Female", "Over 30", "EA over30", "EA under30", "EA more evolve"),
                   fontsize=8)
    plt.ylim([0, 30])
    plot5.set_title("Iliocostalis Left", fontsize=12)

    plot6 = plt.subplot2grid((4, 2), (2, 1))
    data_ilicostalis_r = [[male_tonus["Ilicostalis_R"]], [female_tonus["Ilicostalis_R"]],
                              [over30_tonus["Ilicostalis_R"]],
                              [EA_over_tonus["Ilicostalis_R"]], [EA_more_tonus["Ilicostalis_R"]]]
    plt.boxplot(data_ilicostalis_r, positions = [1,2,3,4,5], widths=0.6, showmeans=True)
    plt.xticks([1, 2, 3, 4, 5], ("Male", "Female", "Over 30", "EA over30", "EA more evolve"),
                   fontsize=8)
    plt.ylim([0, 30])
    plot6.set_title("Iliocostalis Right", fontsize=12)

    plot7 = plt.subplot2grid((4, 2), (3, 0))
    data_multi_l = [[male_tonus["Multifidus_L"]], [female_tonus["Multifidus_L"]],
                        [over30_tonus["Multifidus_L"]],
                        [EA_over_tonus["Multifidus_L"]], [EA_more_tonus["Multifidus_L"]]]
    plt.boxplot(data_multi_l, positions = [1,2,3,4,5], widths=0.6, showmeans=True)
    plt.xticks([1, 2, 3, 4, 5], ("Male", "Female", "Over 30", "EA over30", "EA more evolve"),
                   fontsize=8)
    plt.ylim([0, 30])
    plot7.set_title("Multifidus Left", fontsize=12)

    plot8 = plt.subplot2grid((4, 2), (3, 1))
    data_multi_r = [[male_tonus["Multifidus_R"]], [female_tonus["Multifidus_R"]],
                        [over30_tonus["Multifidus_R"]],
                        [EA_over_tonus["Multifidus_R"]], [EA_more_tonus["Multifidus_R"]]]
    plt.boxplot(data_multi_r, positions = [1,2,3,4,5], widths=0.6, showmeans=True)
    plt.xticks([1, 2, 3, 4, 5], ("Male", "Female", "Over 30", "EA over30", "EA more evolve"),
                   fontsize=8)
    plt.ylim([0, 30])
    plot8.set_title("Multifidus Right", fontsize=12)

    plt.subplots_adjust(top=0.90, bottom=0.10, left=0.12, right=0.90, wspace=0.44, hspace=0.45)
    plt.show()
    pp.savefig(fig)
    pp.close()

def get_IMC():

    wb_over30 = load_workbook(file_excel_over30)
    wb_20 = load_workbook(file_excel_20)
    wb_EA = load_workbook(file_excel_EA)


    count = 1
    count2 = 2

    hight_over30 = []
    weight_over30 = []

    hight_female = []
    weight_female = []

    hight_male = []
    weight_male = []

    hight_EA_over30 = []
    weight_EA_over30 = []

    hight_EA_evaluted = []
    weight_EA_evaluted = []

    for i in wb_over30.sheetnames:
        if i != 'Sheet1' and i != 'Statistical Analysis_20_Female' and i != 'Statistical Analysis_20_Male' \
                and i != 'Statistical Analysis over 30' and i != 'Folha1':
            ws_over30 = wb_over30.get_sheet_by_name(i)
            hight_over30.append(ws_over30['D' + str(count)].value)
            weight_over30.append(ws_over30['D' + str(count2)].value)


    for i in wb_20.sheetnames:
        if i != 'Sheet1' and i != 'Statistical Analysis_20_Female' and i != 'Statistical Analysis_20_Male' \
                and i != 'Statistical Analysis over 30' and i != 'Folha1' and i != 'Patient5_Healthy' \
                and i != 'Patient8_Healthy' and i != 'Patient10_Healthy' and i != 'Patient13_Healthy' \
                and i != 'Patient17_Healthy' and i != 'Patient18_Healthy' and i != 'Patient23_Healthy' \
                and i != 'Patient25_Healthy' and i != 'Patient27_Healthy' and i != 'Patient30_Healthy' \
                and i != 'Patient37_Healthy' and i != 'Patient38_Healthy':
            ws_female = wb_20.get_sheet_by_name(i)
            hight_female.append(ws_female['D' + str(count)].value)
            weight_female.append(ws_female['D' + str(count2)].value)


        if i == 'Patient8_Healthy' or i == 'Patient10_Healthy' or i == 'Patient13_Healthy' \
                or i == 'Patient17_Healthy' or i == 'Patient18_Healthy' or i == 'Patient23_Healthy' \
                or i == 'Patient25_Healthy' or i == 'Patient27_Healthy' or i == 'Patient30_Healthy' \
                or i == 'Patient5_Healthy' or i == 'Patient37_Healthy' or i == 'Patient38_Healthy':
            ws_male = wb_20.get_sheet_by_name(i)
            hight_male.append(ws_male['D' + str(count)].value)
            weight_male.append(ws_male['D' + str(count2)].value)


    for i in wb_EA.sheetnames:
        if i != 'Sheet1' and i != 'Patient4_EA' and i != 'Patient8_EA'  and i != 'Patient11_EA'\
                and i != "Statistical Analysis - EA evol" and i != "Statistical Analysis - EA over3":
            ws_EA_over30 = wb_EA.get_sheet_by_name(i)
            hight_EA_over30.append(ws_EA_over30['D' + str(count)].value)
            weight_EA_over30.append(ws_EA_over30['D' + str(count2)].value)

        if i == 'Patient4_EA' or i == 'Patient8_EA' or i == 'Patient11_EA':
            ws_EA_evaluted = wb_EA.get_sheet_by_name(i)
            hight_EA_evaluted.append(ws_EA_evaluted['D' + str(count)].value)
            weight_EA_evaluted.append(ws_EA_evaluted['D' + str(count2)].value)


    value_final_over30 = {"Height": hight_over30, "Weight": weight_over30}

    value_final_male = {"Height": hight_male, "Weight": weight_male}

    value_final_female = {"Height": hight_female, "Weight": weight_female}

    value_final_EA_over30 = {"Height": hight_EA_over30, "Weight": weight_EA_over30}

    value_final_EA_evaluted = {"Height": hight_EA_evaluted, "Weight": weight_EA_evaluted}

    return value_final_over30, value_final_male, value_final_female, value_final_EA_over30, value_final_EA_evaluted


