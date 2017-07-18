import h5py
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
import numpy as np
from tools import *
#from openpyxl.worksheet.table import Table, TableStyleInfo

file_excel = 'C:/Users/Rita/PycharmProjects/tonusMuscular/Excel_database/database_posturography_20-30_final.xlsx'
patient = 'Patient41_Healthy'

def create_database():
    workbook = xlsxwriter.Workbook('C:/Users/Rita/PycharmProjects/tonusMuscular/Excel_database/database_posturography_20-30_final.xlsx')

def personal_data(file_name):
    file = h5py.File(file_name, 'r')

    wb2 = load_workbook(file_excel)
    sheet1 = wb2.create_sheet(patient)

    fill = PatternFill(start_color='000080', end_color='000080', fill_type='solid')

    A1 = sheet1.cell('A1')
    A1.value = "Gender"
    A1.font = A1.font.copy(bold=True)


    A2 = sheet1.cell('A2')
    A2.value = "Age"
    A2.font = A2.font.copy(bold=True)


    A3 = sheet1.cell('A3')
    A3.value = "Condition"
    A3.font = A3.font.copy(bold=True)


    A4 = sheet1.cell('A4')
    A4.value = "Dominant Hand"
    A4.font = A4.font.copy(bold=True)


    C5 = sheet1.cell('C1')
    C5.value = "Height(cm)"
    C5.font = C5.font.copy(bold=True)

    C6 = sheet1.cell('C2')
    C6.value = "Weight(kg)"
    C6.font = C6.font.copy(bold=True)

    C7 = sheet1.cell('C3')
    C7.value = "Sports"
    C7.font = C7.font.copy(bold=True)



    B1 = sheet1.cell('B1')
    B1.value = file.attrs.values()[0]
    B1.alignment = Alignment(horizontal='center', vertical='center')

    B2 = sheet1.cell('B2')
    B2.value = file.attrs.values()[1]
    B2.alignment = Alignment(horizontal='center', vertical='center')

    B3 = sheet1.cell('B3')
    B3.value = file.attrs.values()[2]
    B3.alignment = Alignment(horizontal='center', vertical='center')

    B4 = sheet1.cell('B4')
    B4.value = file.attrs.values()[3]
    B4.alignment = Alignment(horizontal='center', vertical='center')

    D1 = sheet1.cell('D1')
    D1.value = file.attrs.values()[4]
    D1.alignment = Alignment(horizontal='center', vertical='center')

    D2 = sheet1.cell('D2')
    D2.value = file.attrs.values()[5]
    D2.alignment = Alignment(horizontal='center', vertical='center')

    D3 = sheet1.cell('D3')
    D3.value = file.attrs.values()[6]
    D3.alignment = Alignment(horizontal='center', vertical='center')


    wb2.save(file_excel)


def parameters(EMG_values):
    wb2 = load_workbook(file_excel)
    sheet1 = wb2.get_sheet_by_name(patient)
    count = 6
    count2 = 8
    count_col = 2
    for i in EMG_values:
        if i != 'Reach_Ground' and i != 'Arms_extension':
            title = sheet1.cell('B' + str(count))
            title.value = "Max values of each muscle - " + str(i)
            title.font = title.font.copy(bold=True)

            first_col = sheet1.cell('A' + str(count + 1))
            first_col.value = ''
            first_col.font = first_col.font.copy(bold=True)


            second_col = sheet1.cell('B' + str(count + 1))
            second_col.value = 'Left'
            second_col.font = second_col.font.copy(bold=True)


            third_col = sheet1.cell('C' + str(count + 1))
            third_col.value = 'Right'
            third_col.font = third_col.font.copy(bold=True)


            first_row = sheet1.cell('A' + str(count + 2))
            first_row.value = 'Rectus_A'
            first_row.font = first_row.font.copy(bold=True)


            second_row = sheet1.cell('A' + str(count + 3))
            second_row.value = 'Obliques'
            second_row.font = second_row.font.copy(bold=True)


            third_row = sheet1.cell('A' + str(count + 4))
            third_row.value = 'Ilicostalis'
            third_row.font = third_row.font.copy(bold=True)


            fourth_row = sheet1.cell('A' + str(count + 5))
            fourth_row.value = 'Multifidus'
            fourth_row.font = fourth_row.font.copy(bold=True)



            data = [[EMG_values[i][0], EMG_values[i][1]],
                    [EMG_values[i][2], EMG_values[i][3]],
                    [EMG_values[i][4], EMG_values[i][5]],
                    [EMG_values[i][6], EMG_values[i][7]]
                    ]
            for row in data:
                sheet1.cell(row=count2, column=count_col, value=row[0])
                sheet1.cell(row=count2, column=count_col + 1, value=row[1])
                count2 = count2 + 1

            count = count + 13
            count2 = count2 + 9
    wb2.save(file_excel)

def coherency(coherency_values):
    wb2 = load_workbook(file_excel)
    sheet1 = wb2.get_sheet_by_name(patient)

    fill = PatternFill(start_color='BCC2BC', end_color='BCC2BC', fill_type='solid')

    count = 6
    count1 = 7
    count2 = 8
    count_col = 8
    for i in coherency_values:
        if i != 'Reach_Ground' and i != 'Arms_extension':
            title = sheet1.cell('H' + str(count))
            title.value = "Coherency values between each muscle and each COP direction - " + str(i)
            title.font = title.font.copy(bold=True)

            first_col = sheet1.cell('H' + str(count1))
            first_col.value = ''
            first_col.font = first_col.font.copy(bold=True)
            first_col.fill = fill

            second_col = sheet1.cell('I' + str(count1))
            second_col.value = 'COP X'
            second_col.font = second_col.font.copy(bold=True)
            second_col.fill = fill

            third_col = sheet1.cell('J' + str(count1))
            third_col.value = 'COP Y'
            third_col.font = third_col.font.copy(bold=True)
            third_col.fill = fill

            first_row = sheet1.cell('H' + str(count1+1))
            first_row.value = 'Rectus_A_L'
            first_row.font = first_row.font.copy(bold=True)
            first_row.fill = fill

            second_row = sheet1.cell('H' + str(count1 + 2))
            second_row.value = 'Obliques_L'
            second_row.font = second_row.font.copy(bold=True)
            second_row.fill = fill

            third_row = sheet1.cell('H' + str(count1 + 3))
            third_row.value = 'Ilicostalis_L'
            third_row.font = third_row.font.copy(bold=True)
            third_row.fill = fill

            fourth_row = sheet1.cell('H' + str(count1 + 4))
            fourth_row.value = 'Multifidus_L'
            fourth_row.font = fourth_row.font.copy(bold=True)
            fourth_row.fill = fill

            fifth_row = sheet1.cell('H' + str(count1 + 5))
            fifth_row.value = 'Rectus_A_R'
            fifth_row.font = fifth_row.font.copy(bold=True)
            fifth_row.fill = fill

            sixth_row = sheet1.cell('H' + str(count1 + 6))
            sixth_row.value = 'Obliques_R'
            sixth_row.font = sixth_row.font.copy(bold=True)
            sixth_row.fill = fill

            seventh_row = sheet1.cell('H' + str(count1 + 7))
            seventh_row.value = 'Ilicostalis_R'
            seventh_row.font = seventh_row.font.copy(bold=True)
            seventh_row.fill = fill

            eigth_row = sheet1.cell('H' + str(count1 + 8))
            eigth_row.value = 'Multifidus_R'
            eigth_row.font = eigth_row.font.copy(bold=True)
            eigth_row.fill = fill


            data = [[np.max(coherency_values[i]["coherency_x"][0:40, 0]),  np.max(coherency_values[i]["coherency_y"][0:40, 0])],
                    [np.max(coherency_values[i]["coherency_x"][0:40, 2]),  np.max(coherency_values[i]["coherency_y"][0:40, 2])],
                    [np.max(coherency_values[i]["coherency_x"][0:40, 4]),  np.max(coherency_values[i]["coherency_y"][0:40, 4])],
                    [np.max(coherency_values[i]["coherency_x"][0:40, 6]),  np.max(coherency_values[i]["coherency_y"][0:40, 6])],
                    [np.max(coherency_values[i]["coherency_x"][0:40, 1]),  np.max(coherency_values[i]["coherency_y"][0:40, 1])],
                    [np.max(coherency_values[i]["coherency_x"][0:40, 3]),  np.max(coherency_values[i]["coherency_y"][0:40, 3])],
                    [np.max(coherency_values[i]["coherency_x"][0:40, 5]),  np.max(coherency_values[i]["coherency_y"][0:40, 5])],
                    [np.max(coherency_values[i]["coherency_x"][0:40, 7]),  np.max(coherency_values[i]["coherency_y"][0:40, 7])]
                    ]

            for row in data:
                sheet1.cell(row= count2, column= count_col + 1, value= row[0])
                sheet1.cell(row= count2, column= count_col + 2, value= row[1])
                count2 = count2 + 1

            count = count + 13
            count1 = count1 + 13
            count2 = (count2-1) + 6
    wb2.save(file_excel)

def COP_parameters (mean_velocity, platform_COP, std, amplitude):
    wb2 = load_workbook(file_excel)
    sheet1 = wb2.get_sheet_by_name(patient)

    fill = PatternFill(start_color='BCC2BC', end_color='BCC2BC', fill_type='solid')

    count = 6
    count1 = 7

    for i in mean_velocity:
        if i != 'Reach_Ground' and i != 'Arms_extension':
            title = sheet1.cell('P' + str(count))
            title.value = "COP important values - " + str(i)
            title.font = title.font.copy(bold=True)

            first_col = sheet1.cell('P' + str(count1))
            first_col.value = ''
            first_col.font = first_col.font.copy(bold=True)
            first_col.fill = fill

            second_col = sheet1.cell('Q' + str(count1))
            second_col.value = 'COP X'
            second_col.font = second_col.font.copy(bold=True)
            second_col.fill = fill

            third_col = sheet1.cell('R' + str(count1))
            third_col.value = 'COP Y'
            third_col.font = third_col.font.copy(bold=True)
            third_col.fill = fill

            first_row = sheet1.cell('P' + str(count1 + 1))
            first_row.value = 'Mean Velocity'
            first_row.font = first_row.font.copy(bold=True)
            first_row.fill = fill

            second_row = sheet1.cell('P' + str(count1 + 2))
            second_row.value = 'STD'
            second_row.font = second_row.font.copy(bold=True)
            second_row.fill = fill

            third_row = sheet1.cell('P' + str(count1 + 3))
            third_row.value = 'Amplitude'
            third_row.font = third_row.font.copy(bold=True)
            third_row.fill = fill
            wb2 = load_workbook(file_excel)
            sheet1 = wb2.get_sheet_by_name(patient)

            fill = PatternFill(start_color='BCC2BC', end_color='BCC2BC', fill_type='solid')

            count = 6
            count1 = 7

            for i in mean_velocity:
                if i != 'Reach_Ground' and i != 'Arms_extension':
                    title = sheet1.cell('P' + str(count))
                    title.value = "COP important values - " + str(i)
                    title.font = title.font.copy(bold=True)

                    first_col = sheet1.cell('P' + str(count1))
                    first_col.value = ''
                    first_col.font = first_col.font.copy(bold=True)
                    first_col.fill = fill

                    second_col = sheet1.cell('Q' + str(count1))
                    second_col.value = 'COP X'
                    second_col.font = second_col.font.copy(bold=True)
                    second_col.fill = fill

                    third_col = sheet1.cell('R' + str(count1))
                    third_col.value = 'COP Y'
                    third_col.font = third_col.font.copy(bold=True)
                    third_col.fill = fill

                    first_row = sheet1.cell('P' + str(count1 + 1))
                    first_row.value = 'Mean Velocity'
                    first_row.font = first_row.font.copy(bold=True)
                    first_row.fill = fill

                    second_row = sheet1.cell('P' + str(count1 + 2))
                    second_row.value = 'STD'
                    second_row.font = second_row.font.copy(bold=True)
                    second_row.fill = fill

                    third_row = sheet1.cell('P' + str(count1 + 3))
                    third_row.value = 'Amplitude'
                    third_row.font = third_row.font.copy(bold=True)
                    third_row.fill = fill

                    velocity_x = sheet1.cell('Q' + str(count1+1))
                    velocity_x.value = mean_velocity[i]["COP_X"]

                    velocity_y = sheet1.cell('R' + str(count1+1))
                    velocity_y.value = mean_velocity[i]["COP_Y"]

                    std_x = sheet1.cell('Q' + str(count1 + 2))
                    std_x.value = std[i]["COP_X"]

                    std_y = sheet1.cell('R' + str(count1 + 2))
                    std_y.value = std[i]["COP_Y"]

                    amp_x = sheet1.cell('Q' + str(count1 + 3))
                    amp_x.value = amplitude[i]["COP_X"]

                    amp_y = sheet1.cell('R' + str(count1 + 3))
                    amp_y.value = amplitude[i]["COP_Y"]

                    area_traj = convex_hull(platform_COP[i]["COP_X"], platform_COP[i]["COP_Y"])
                    area_value = area_calc(area_traj)

                    area = sheet1.cell('P' + str(count1 + 6))
                    area.value = "Area COP"
                    area.font = area.font.copy(bold=True)
                    area.fill = fill

                    area_val = sheet1.cell(('Q' + str(count1 + 6)))
                    area_val.value = area_value


                    count = count + 13
                    count1 = count1 + 13

    wb2.save(file_excel)


def correlation_RL(correlation_RL):
    wb2 = load_workbook(file_excel)
    sheet1 = wb2.get_sheet_by_name(patient)

    fill = PatternFill(start_color='FF6600', end_color='FF6600', fill_type='solid')
    header = sheet1.cell('A151')
    header.value = "Correlation values between EMG, COP, velocity and accelaration"
    header.font = header.font.copy(bold=True)

    RL_description1 = sheet1.cell('A154')
    RL_description1.value = "**EMG data - Right muscle minus left muscle. Correlation between EMG data,"

    RL_description1 = sheet1.cell('A155')
    RL_description1.value = " COP in X direction, velocity in X direction and accelaration in X direction."

    count = 158
    count1 = 160
    count_col = 2

    for i in correlation_RL:
        if i != 'Reach_Ground' and i != 'Arms_extension':
            title = sheet1.cell('A' + str(count))
            title.value = "Correlation values - " + str(i)
            title.font = title.font.copy(bold=True)

            first_col = sheet1.cell('A' + str(count + 1))
            first_col.value = ''
            first_col.font = first_col.font.copy(bold=True)
            first_col.fill = fill

            second_col = sheet1.cell('B' + str(count + 1))
            second_col.value = 'COP X'
            second_col.font = second_col.font.copy(bold=True)
            second_col.fill = fill

            third_col = sheet1.cell('C' + str(count + 1))
            third_col.value = 'Velocity X'
            third_col.font = third_col.font.copy(bold=True)
            third_col.fill = fill

            fourth_col = sheet1.cell('D' + str(count + 1))
            fourth_col.value = 'Acelaration X'
            fourth_col.font = fourth_col.font.copy(bold=True)
            fourth_col.fill = fill

            first_row = sheet1.cell('A' + str(count + 2))
            first_row.value = 'Rectus_A'
            first_row.font = first_row.font.copy(bold=True)
            first_row.fill = fill

            second_row = sheet1.cell('A' + str(count + 3))
            second_row.value = 'Obliques'
            second_row.font = second_row.font.copy(bold=True)
            second_row.fill = fill

            third_row = sheet1.cell('A' + str(count + 4))
            third_row.value = 'Ilicostalis'
            third_row.font = third_row.font.copy(bold=True)
            third_row.fill = fill

            fourth_row = sheet1.cell('A' + str(count + 5))
            fourth_row.value = 'Multifidus'
            fourth_row.font = fourth_row.font.copy(bold=True)
            fourth_row.fill = fill

            data = [[correlation_RL[i]["RA_corr"]["RA_COP"], correlation_RL[i]["RA_corr"]["RA_vel"], correlation_RL[i]["RA_corr"]["RA_acel"]],
                    [correlation_RL[i]["O_corr"]["O_COP"], correlation_RL[i]["O_corr"]["O_vel"], correlation_RL[i]["O_corr"]["O_acel"]],
                    [correlation_RL[i]["I_corr"]["I_COP"], correlation_RL[i]["I_corr"]["I_vel"], correlation_RL[i]["I_corr"]["I_acel"]],
                    [correlation_RL[i]["M_corr"]["M_COP"], correlation_RL[i]["M_corr"]["M_vel"], correlation_RL[i]["M_corr"]["M_acel"]]
                    ]

            for row in data:
                sheet1.cell(row = count1, column = count_col, value = row[0])
                sheet1.cell(row = count1, column = count_col + 1, value = row[1])
                sheet1.cell(row = count1, column = count_col + 2, value = row[2])
                count1 = count1 + 1

            count = count + 8
            count1 = (count1 - 1) + 5

    wb2.save(file_excel)

def correlation_FB(correlation_FB):
    wb2 = load_workbook(file_excel)
    sheet1 = wb2.get_sheet_by_name(patient)

    fill = PatternFill(start_color='FF6600', end_color='FF6600', fill_type='solid')

    FB_description1 = sheet1.cell('H154')
    FB_description1.value = "**EMG data - Front muscle minus back muscle. Correlation between EMG data,"

    FB_description1 = sheet1.cell('H155')
    FB_description1.value = " COP in Y direction, velocity in Y direction and accelaration in Y direction."

    count = 158
    count1 = 160
    count_col = 9

    for i in correlation_FB:
        if i != 'Reach_Ground' and i != 'Arms_extension':
            title = sheet1.cell('H' + str(count))
            title.value = "Correlation values - " + str(i)
            title.font = title.font.copy(bold=True)

            first_col = sheet1.cell('H' + str(count + 1))
            first_col.value = ''
            first_col.font = first_col.font.copy(bold=True)
            first_col.fill = fill

            second_col = sheet1.cell('I' + str(count + 1))
            second_col.value = 'COP Y'
            second_col.font = second_col.font.copy(bold=True)
            second_col.fill = fill

            third_col = sheet1.cell('J' + str(count + 1))
            third_col.value = 'Velocity Y'
            third_col.font = third_col.font.copy(bold=True)
            third_col.fill = fill

            fourth_col = sheet1.cell('K' + str(count + 1))
            fourth_col.value = 'Acelaration Y'
            fourth_col.font = fourth_col.font.copy(bold=True)
            fourth_col.fill = fill

            first_row = sheet1.cell('H' + str(count + 2))
            first_row.value = 'R-M_L'
            first_row.font = first_row.font.copy(bold=True)
            first_row.fill = fill

            second_row = sheet1.cell('H' + str(count + 3))
            second_row.value = 'R-M_R'
            second_row.font = second_row.font.copy(bold=True)
            second_row.fill = fill

            third_row = sheet1.cell('H' + str(count + 4))
            third_row.value = 'O-I_L'
            third_row.font = third_row.font.copy(bold=True)
            third_row.fill = fill

            fourth_row = sheet1.cell('H' + str(count + 5))
            fourth_row.value = 'O-I_R'
            fourth_row.font = fourth_row.font.copy(bold=True)
            fourth_row.fill = fill

            data = [[correlation_FB[i]["MR_L_corr"]["MR_L_COP"], correlation_FB[i]["MR_L_corr"]["MR_L_vel"], correlation_FB[i]["MR_L_corr"]["MR_L_acel"]],
                    [correlation_FB[i]["MR_R_corr"]["MR_R_COP"], correlation_FB[i]["MR_R_corr"]["MR_R_vel"], correlation_FB[i]["MR_R_corr"]["MR_R_acel"]],
                    [correlation_FB[i]["IO_L_corr"]["IO_L_COP"], correlation_FB[i]["IO_L_corr"]["IO_L_vel"], correlation_FB[i]["IO_L_corr"]["IO_L_acel"]],
                    [correlation_FB[i]["IO_R_corr"]["IO_R_COP"], correlation_FB[i]["IO_R_corr"]["IO_R_vel"], correlation_FB[i]["IO_R_corr"]["IO_R_acel"]]
                    ]

            for row in data:
                sheet1.cell(row = count1, column = count_col, value = row[0])
                sheet1.cell(row = count1, column = count_col + 1, value = row[1])
                sheet1.cell(row = count1, column = count_col + 2, value = row[2])

                count1 = count1 + 1

            count = count + 8
            count1 = count1 + 4

    wb2.save(file_excel)

def correlation_FB_cross(correlation_FB_cross):
    wb2 = load_workbook(file_excel)
    sheet1 = wb2.get_sheet_by_name(patient)

    fill = PatternFill(start_color='FF6600', end_color='FF6600', fill_type='solid')

    FB_description1 = sheet1.cell('O154')
    FB_description1.value = "**EMG data - Front muscle minus back muscle in cross direction. Correlation between EMG data,"

    FB_description1 = sheet1.cell('O155')
    FB_description1.value = " COP in Y direction, velocity in Y direction and accelaration in Y direction."

    count = 158
    count1 = 160
    count_col = 16

    for i in correlation_FB_cross:
        if i != 'Reach_Ground' and i != 'Arms_extension':
            title = sheet1.cell('O' + str(count))
            title.value = "Correlation values - " + str(i)
            title.font = title.font.copy(bold=True)

            first_col = sheet1.cell('O' + str(count + 1))
            first_col.value = ''
            first_col.font = first_col.font.copy(bold=True)
            first_col.fill = fill

            second_col = sheet1.cell('P' + str(count + 1))
            second_col.value = 'COP X'
            second_col.font = second_col.font.copy(bold=True)
            second_col.fill = fill

            third_col = sheet1.cell('Q' + str(count + 1))
            third_col.value = 'COP Y'
            third_col.font = third_col.font.copy(bold=True)
            third_col.fill = fill

            first_row = sheet1.cell('O' + str(count + 2))
            first_row.value = 'RL-MR'
            first_row.font = first_row.font.copy(bold=True)
            first_row.fill = fill

            second_row = sheet1.cell('O' + str(count + 3))
            second_row.value = 'RR-LM'
            second_row.font = second_row.font.copy(bold=True)
            second_row.fill = fill

            third_row = sheet1.cell('O' + str(count + 4))
            third_row.value = 'OL-IR'
            third_row.font = third_row.font.copy(bold=True)
            third_row.fill = fill

            fourth_row = sheet1.cell('O' + str(count + 5))
            fourth_row.value = 'OR-IL'
            fourth_row.font = fourth_row.font.copy(bold=True)
            fourth_row.fill = fill

            data = [[correlation_FB_cross[i]["MR_LR_corr"]["MR_LR_COP_X"], correlation_FB_cross[i]["MR_LR_corr"]["MR_LR_COP_Y"]],
                    [correlation_FB_cross[i]["MR_RL_corr"]["MR_RL_COP_X"], correlation_FB_cross[i]["MR_RL_corr"]["MR_RL_COP_Y"]],
                    [correlation_FB_cross[i]["IO_LR_corr"]["IO_LR_COP_X"], correlation_FB_cross[i]["IO_LR_corr"]["IO_LR_COP_Y"]],
                    [correlation_FB_cross[i]["IO_RL_corr"]["IO_RL_COP_X"], correlation_FB_cross[i]["IO_RL_corr"]["IO_RL_COP_Y"]]
                    ]

            for row in data:
                sheet1.cell(row = count1, column = count_col, value = row[0])
                sheet1.cell(row = count1, column = count_col + 1, value = row[1])

                count1 = count1 + 1

            count = count + 8
            count1 = count1 + 4

    wb2.save(file_excel)

def correlation(simple_corr):
    wb2 = load_workbook(file_excel)
    sheet1 = wb2.get_sheet_by_name(patient)

    count = 158
    count2 = 160
    count_col = 24

    fill = PatternFill(start_color='FF6600', end_color='FF6600', fill_type='solid')
    for i in simple_corr:
        if i != 'Reach_Ground' and i != 'Arms_extension':
            title = sheet1.cell('W' + str(count))
            title.value = "Correlation coeficient between each muscle and each COP direction - " + str(i)
            title.font = title.font.copy(bold=True)

            first_col = sheet1.cell('W' + str(count + 1))
            first_col.value = ''
            first_col.font = first_col.font.copy(bold=True)
            first_col.fill = fill

            second_col = sheet1.cell('X' + str(count + 1))
            second_col.value = 'COP X'
            second_col.font = second_col.font.copy(bold=True)
            second_col.fill = fill

            third_col = sheet1.cell('Y' + str(count + 1))
            third_col.value = 'COP Y'
            third_col.font = third_col.font.copy(bold=True)
            third_col.fill = fill

            first_row = sheet1.cell('W' + str(count + 2))
            first_row.value = 'Rectus_A_L'
            first_row.font = first_row.font.copy(bold=True)
            first_row.fill = fill

            second_row = sheet1.cell('W' + str(count + 3))
            second_row.value = 'Obliques_L'
            second_row.font = second_row.font.copy(bold=True)
            second_row.fill = fill

            third_row = sheet1.cell('W' + str(count + 4))
            third_row.value = 'Ilicostalis_L'
            third_row.font = third_row.font.copy(bold=True)
            third_row.fill = fill

            fourth_row = sheet1.cell('W' + str(count + 5))
            fourth_row.value = 'Multifidus_L'
            fourth_row.font = fourth_row.font.copy(bold=True)
            fourth_row.fill = fill

            fifth_row = sheet1.cell('W' + str(count + 6))
            fifth_row.value = 'Rectus_A_R'
            fifth_row.font = fifth_row.font.copy(bold=True)
            fifth_row.fill = fill

            sixth_row = sheet1.cell('W' + str(count + 7))
            sixth_row.value = 'Obliques_R'
            sixth_row.font = sixth_row.font.copy(bold=True)
            sixth_row.fill = fill

            seventh_row = sheet1.cell('W' + str(count + 8))
            seventh_row.value = 'Ilicostalis_R'
            seventh_row.font = seventh_row.font.copy(bold=True)
            seventh_row.fill = fill

            eigth_row = sheet1.cell('W' + str(count + 9))
            eigth_row.value = 'Multifidus_R'
            eigth_row.font = eigth_row.font.copy(bold=True)
            eigth_row.fill = fill

            data = [[simple_corr[i]["COP_X"][0, 0], simple_corr[i]["COP_Y"][0, 0]],
                    [simple_corr[i]["COP_X"][0, 2], simple_corr[i]["COP_Y"][0, 2]],
                    [simple_corr[i]["COP_X"][0, 4], simple_corr[i]["COP_Y"][0, 4]],
                    [simple_corr[i]["COP_X"][0, 6], simple_corr[i]["COP_Y"][0, 6]],
                    [simple_corr[i]["COP_X"][0, 1], simple_corr[i]["COP_Y"][0, 1]],
                    [simple_corr[i]["COP_X"][0, 3], simple_corr[i]["COP_Y"][0, 3]],
                    [simple_corr[i]["COP_X"][0, 5], simple_corr[i]["COP_Y"][0, 5]],
                    [simple_corr[i]["COP_X"][0, 7], simple_corr[i]["COP_Y"][0, 7]]
                    ]

            for row in data:
                sheet1.cell(row=count2, column=count_col, value=row[0])
                sheet1.cell(row=count2, column=count_col + 1, value=row[1])
                count2 = count2 + 1


            count = count + 12
            count2 = count2 + 4

    wb2.save(file_excel)

def fourrier_parameters_EMG(peak_f, mean_f, f80, median_f):
    wb2 = load_workbook(file_excel)
    sheet1 = wb2.get_sheet_by_name(patient)

    freq_description = sheet1.cell('A252')
    freq_description.value = "**Frequency analysis of EMG and COP"

    fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    count = 258
    count2 = 260
    count_col = 2

    for i in peak_f:
        if i != 'Reach_Ground' and i != 'Arms_extension':
            title = sheet1.cell('A' + str(count))
            title.value = "Frequency EMG - " + str(i)
            title.font = title.font.copy(bold=True)

            first_col = sheet1.cell('A' + str(count + 1))
            first_col.value = ''
            first_col.font = first_col.font.copy(bold=True)
            first_col.fill = fill

            second_col = sheet1.cell('B' + str(count + 1))
            second_col.value = 'Peak Freq (Hz)'
            second_col.font = second_col.font.copy(bold=True)
            second_col.fill = fill

            third_col = sheet1.cell('C' + str(count + 1))
            third_col.value = 'Mean Freq (Hz)'
            third_col.font = third_col.font.copy(bold=True)
            third_col.fill = fill

            fourth_col = sheet1.cell('D' + str(count + 1))
            fourth_col.value = 'Median Freq (Hz)'
            fourth_col.font = fourth_col.font.copy(bold=True)
            fourth_col.fill = fill

            fifth_col = sheet1.cell('E' + str(count + 1))
            fifth_col.value = '80% Freq (Hz)'
            fifth_col.font = fifth_col.font.copy(bold=True)
            fifth_col.fill = fill

            first_row = sheet1.cell('A' + str(count + 2))
            first_row.value = 'Rectus_A_L'
            first_row.font = first_row.font.copy(bold=True)
            first_row.fill = fill

            second_row = sheet1.cell('A' + str(count + 3))
            second_row.value = 'Rectus_A_R'
            second_row.font = second_row.font.copy(bold=True)
            second_row.fill = fill

            third_row = sheet1.cell('A' + str(count + 4))
            third_row.value = 'Obliques_L'
            third_row.font = third_row.font.copy(bold=True)
            third_row.fill = fill

            fourth_row = sheet1.cell('A' + str(count + 5))
            fourth_row.value = 'Obliques_R'
            fourth_row.font = fourth_row.font.copy(bold=True)
            fourth_row.fill = fill

            fifth_row = sheet1.cell('A' + str(count + 6))
            fifth_row.value = 'Ilicostalis_L'
            fifth_row.font = fifth_row.font.copy(bold=True)
            fifth_row.fill = fill

            sixth_row = sheet1.cell('A' + str(count + 7))
            sixth_row.value = 'Ilicostalis_R'
            sixth_row.font = sixth_row.font.copy(bold=True)
            sixth_row.fill = fill

            seventh_row = sheet1.cell('A' + str(count + 8))
            seventh_row.value = 'Multifidus_L'
            seventh_row.font = seventh_row.font.copy(bold=True)
            seventh_row.fill = fill

            eigth_row = sheet1.cell('A' + str(count + 9))
            eigth_row.value = 'Multifidus_R'
            eigth_row.font = eigth_row.font.copy(bold=True)
            eigth_row.fill = fill

            data = [[peak_f[i][0, 0], mean_f[i][0, 0], median_f[i][0, 0], f80[i][0, 0]],
                    [peak_f[i][0, 1], mean_f[i][0, 1], median_f[i][0, 1], f80[i][0, 1]],
                    [peak_f[i][0, 2], mean_f[i][0, 2], median_f[i][0, 2], f80[i][0, 2]],
                    [peak_f[i][0, 3], mean_f[i][0, 3], median_f[i][0, 3], f80[i][0, 3]],
                    [peak_f[i][0, 4], mean_f[i][0, 4], median_f[i][0, 4], f80[i][0, 4]],
                    [peak_f[i][0, 5], mean_f[i][0, 5], median_f[i][0, 5], f80[i][0, 5]],
                    [peak_f[i][0, 6], mean_f[i][0, 6], median_f[i][0, 6], f80[i][0, 6]],
                    [peak_f[i][0, 7], mean_f[i][0, 7], median_f[i][0, 7], f80[i][0, 7]]]
            for row in data:
                sheet1.cell(row=count2, column=count_col, value=row[0])
                sheet1.cell(row=count2, column=count_col + 1, value=row[1])
                sheet1.cell(row=count2, column=count_col + 2, value=row[2])
                sheet1.cell(row=count2, column=count_col + 3, value=row[3])
                count2 = count2 + 1

            count = count + 12
            count2 = count2 + 4

    wb2.save(file_excel)


def fourrier_parameters_COP(peak_f_COP, mean_f_COP, f80_COP, median_f_COP):
    wb2 = load_workbook(file_excel)
    sheet1 = wb2.get_sheet_by_name(patient)

    freq_description = sheet1.cell('A252')
    freq_description.value = "**Frequency analysis of EMG and COP"

    fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    count = 258
    count2 = 260
    count_col = 11

    for i in peak_f_COP:
        if i != 'Reach_Ground' and i != 'Arms_extension':
            title = sheet1.cell('J' + str(count))
            title.value = "Frequency COP - " + str(i)
            title.font = title.font.copy(bold=True)

            first_col = sheet1.cell('J' + str(count + 1))
            first_col.value = ''
            first_col.font = first_col.font.copy(bold=True)
            first_col.fill = fill

            second_col = sheet1.cell('K' + str(count + 1))
            second_col.value = 'Peak Freq (Hz)'
            second_col.font = second_col.font.copy(bold=True)
            second_col.fill = fill

            third_col = sheet1.cell('L' + str(count + 1))
            third_col.value = 'Mean Freq (Hz)'
            third_col.font = third_col.font.copy(bold=True)
            third_col.fill = fill

            fourth_col = sheet1.cell('M' + str(count + 1))
            fourth_col.value = 'Median Freq (Hz)'
            fourth_col.font = fourth_col.font.copy(bold=True)
            fourth_col.fill = fill

            fifth_col = sheet1.cell('N' + str(count + 1))
            fifth_col.value = '80% Freq (Hz)'
            fifth_col.font = fifth_col.font.copy(bold=True)
            fifth_col.fill = fill

            first_row = sheet1.cell('J' + str(count + 2))
            first_row.value = 'COP X'
            first_row.font = first_row.font.copy(bold=True)
            first_row.fill = fill

            second_row = sheet1.cell('J' + str(count + 3))
            second_row.value = 'COP_Y'
            second_row.font = second_row.font.copy(bold=True)
            second_row.fill = fill



            data = [[peak_f_COP[i]["COP_X"], mean_f_COP[i]["COP_X"], median_f_COP[i]["COP_X"], f80_COP[i]["COP_X"]],
                    [peak_f_COP[i]["COP_Y"], mean_f_COP[i]["COP_Y"], median_f_COP[i]["COP_Y"], f80_COP[i]["COP_Y"]]]

            for row in data:
                sheet1.cell(row=count2, column=count_col, value=row[0])
                sheet1.cell(row=count2, column=count_col + 1, value=row[1])
                sheet1.cell(row=count2, column=count_col + 2, value=row[2])
                sheet1.cell(row=count2, column=count_col + 3, value=row[3])
                count2 = count2 + 1

            count = count + 12
            count2 = count2 + 10

    wb2.save(file_excel)


def mean_std(peak_cop):
    wb2 = load_workbook(file_excel)
    sheet1 = wb2.get_sheet_by_name(patient)

    freq_description = sheet1.cell('A252')
    freq_description.value = "**Frequency analysis of EMG and COP"

    fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    count = 276


    for i in peak_cop:
        if i != 'Reach_Ground' and i != 'Arms_extension':
            title = sheet1.cell('M' + str(count))
            title.value = "Correlation coeficient between each muscle and each COP direction - " + str(i)
            title.font = title.font.copy(bold=True)

            first_col = sheet1.cell('M' + str(count + 1))
            first_col.value = ''
            first_col.font = first_col.font.copy(bold=True)
            first_col.fill = fill

            first_col = sheet1.cell('M' + str(count + 2))
            first_col.value = ''
            first_col.font = first_col.font.copy(bold=True)
            first_col.fill = fill

            second_col = sheet1.cell('N' + str(count + 1))
            second_col.value = 'Peak Freq (Hz)'
            second_col.font = second_col.font.copy(bold=True)
            second_col.fill = fill

            mean1 = sheet1.cell('N' + str(count + 2))
            mean1.value = 'Mean Value'
            mean1.font = mean1.font.copy(bold=True)
            mean1.fill = fill

            std1 = sheet1.cell('O' + str(count + 2))
            std1.value = 'STD Value'
            std1.font = std1.font.copy(bold=True)
            std1.fill = fill

            third_col = sheet1.cell('P' + str(count + 1))
            third_col.value = 'Mean Freq (Hz)'
            third_col.font = third_col.font.copy(bold=True)
            third_col.fill = fill

            mean2 = sheet1.cell('P' + str(count + 2))
            mean2.value = 'Mean Value'
            mean2.font = mean2.font.copy(bold=True)
            mean2.fill = fill

            std2 = sheet1.cell('Q' + str(count + 2))
            std2.value = 'STD Value'
            std2.font = std2.font.copy(bold=True)
            std2.fill = fill

            fourth_col = sheet1.cell('R' + str(count + 1))
            fourth_col.value = 'Median Freq (Hz)'
            fourth_col.font = fourth_col.font.copy(bold=True)
            fourth_col.fill = fill

            mean3 = sheet1.cell('R' + str(count + 2))
            mean3.value = 'Mean Value'
            mean3.font = mean3.font.copy(bold=True)
            mean3.fill = fill

            std3 = sheet1.cell('S' + str(count + 2))
            std3.value = 'STD Value'
            std3.font = std3.font.copy(bold=True)
            std3.fill = fill

            fifth_col = sheet1.cell('T' + str(count + 1))
            fifth_col.value = '80% Freq (Hz)'
            fifth_col.font = fifth_col.font.copy(bold=True)
            fifth_col.fill = fill

            mean4 = sheet1.cell('T' + str(count + 2))
            mean4.value = 'Mean Value'
            mean4.font = mean4.font.copy(bold=True)
            mean4.fill = fill

            std4 = sheet1.cell('U' + str(count + 2))
            std4.value = 'STD Value'
            std4.font = std4.font.copy(bold=True)
            std4.fill = fill

            first_row = sheet1.cell('M' + str(count + 3))
            first_row.value = 'COP X'
            first_row.font = first_row.font.copy(bold=True)
            first_row.fill = fill

            second_row = sheet1.cell('M' + str(count + 4))
            second_row.value = 'COP Y'
            second_row.font = second_row.font.copy(bold=True)
            second_row.fill = fill





            count = count + 12

    wb2.save(file_excel)


def rest_parameters(rest_array, peak_f, mean_f, median_f, f80):
    wb2 = load_workbook(file_excel)
    sheet1 = wb2.get_sheet_by_name(patient)

    fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    count_max = 144
    count2 = 146
    count_col = 2

    count_freq = 390
    count2_freq = 392

    for i in rest_array:
        if i == "Relax":

            title = sheet1.cell('B' + str(count_max))
            title.value = "Max values of each muscle - " + str(i)
            title.font = title.font.copy(bold=True)

            first_col = sheet1.cell('A' + str(count_max + 1))
            first_col.value = ''
            first_col.font = first_col.font.copy(bold=True)

            second_col = sheet1.cell('B' + str(count_max + 1))
            second_col.value = 'Left'
            second_col.font = second_col.font.copy(bold=True)

            third_col = sheet1.cell('C' + str(count_max + 1))
            third_col.value = 'Right'
            third_col.font = third_col.font.copy(bold=True)

            first_row = sheet1.cell('A' + str(count_max + 2))
            first_row.value = 'Rectus_A'
            first_row.font = first_row.font.copy(bold=True)

            second_row = sheet1.cell('A' + str(count_max + 3))
            second_row.value = 'Obliques'
            second_row.font = second_row.font.copy(bold=True)

            third_row = sheet1.cell('A' + str(count_max + 4))
            third_row.value = 'Ilicostalis'
            third_row.font = third_row.font.copy(bold=True)

            fourth_row = sheet1.cell('A' + str(count_max + 5))
            fourth_row.value = 'Multifidus'
            fourth_row.font = fourth_row.font.copy(bold=True)

            data = [[rest_array[i][0], rest_array[i][1]],
                    [rest_array[i][2], rest_array[i][3]],
                    [rest_array[i][4], rest_array[i][5]],
                    [rest_array[i][6], rest_array[i][7]]
                    ]
            for row in data:
                sheet1.cell(row=count2, column=count_col, value=row[0])
                sheet1.cell(row=count2, column=count_col + 1, value=row[1])
                count2 = count2 + 1

            title = sheet1.cell('A' + str(count_freq))
            title.value = "Correlation coeficient between each muscle and each COP direction - " + str(i)
            title.font = title.font.copy(bold=True)

            first_col = sheet1.cell('A' + str(count_freq + 1))
            first_col.value = ''
            first_col.font = first_col.font.copy(bold=True)
            first_col.fill = fill

            second_col = sheet1.cell('B' + str(count_freq + 1))
            second_col.value = 'Peak Freq (Hz)'
            second_col.font = second_col.font.copy(bold=True)
            second_col.fill = fill

            third_col = sheet1.cell('C' + str(count_freq + 1))
            third_col.value = 'Mean Freq (Hz)'
            third_col.font = third_col.font.copy(bold=True)
            third_col.fill = fill

            fourth_col = sheet1.cell('D' + str(count_freq + 1))
            fourth_col.value = 'Median Freq (Hz)'
            fourth_col.font = fourth_col.font.copy(bold=True)
            fourth_col.fill = fill

            fifth_col = sheet1.cell('E' + str(count_freq + 1))
            fifth_col.value = '80% Freq (Hz)'
            fifth_col.font = fifth_col.font.copy(bold=True)
            fifth_col.fill = fill

            first_row = sheet1.cell('A' + str(count_freq + 2))
            first_row.value = 'Rectus_A_L'
            first_row.font = first_row.font.copy(bold=True)
            first_row.fill = fill

            second_row = sheet1.cell('A' + str(count_freq + 3))
            second_row.value = 'Rectus_A_R'
            second_row.font = second_row.font.copy(bold=True)
            second_row.fill = fill

            third_row = sheet1.cell('A' + str(count_freq + 4))
            third_row.value = 'Obliques_L'
            third_row.font = third_row.font.copy(bold=True)
            third_row.fill = fill

            fourth_row = sheet1.cell('A' + str(count_freq + 5))
            fourth_row.value = 'Obliques_R'
            fourth_row.font = fourth_row.font.copy(bold=True)
            fourth_row.fill = fill

            fifth_row = sheet1.cell('A' + str(count_freq + 6))
            fifth_row.value = 'Ilicostalis_L'
            fifth_row.font = fifth_row.font.copy(bold=True)
            fifth_row.fill = fill

            sixth_row = sheet1.cell('A' + str(count_freq + 7))
            sixth_row.value = 'Ilicostalis_R'
            sixth_row.font = sixth_row.font.copy(bold=True)
            sixth_row.fill = fill

            seventh_row = sheet1.cell('A' + str(count_freq + 8))
            seventh_row.value = 'Multifidus_L'
            seventh_row.font = seventh_row.font.copy(bold=True)
            seventh_row.fill = fill

            eigth_row = sheet1.cell('A' + str(count_freq + 9))
            eigth_row.value = 'Multifidus_R'
            eigth_row.font = eigth_row.font.copy(bold=True)
            eigth_row.fill = fill

            # data = [[peak_f[i][0, 0], mean_f[i][0, 0], median_f[i][0, 0], f80[i][0, 0]],
            #         [peak_f[i][0, 1], mean_f[i][0, 1], median_f[i][0, 1], f80[i][0, 1]],
            #         [peak_f[i][0, 2], mean_f[i][0, 2], median_f[i][0, 2], f80[i][0, 2]],
            #         [peak_f[i][0, 3], mean_f[i][0, 3], median_f[i][0, 3], f80[i][0, 3]],
            #         [peak_f[i][0, 4], mean_f[i][0, 4], median_f[i][0, 4], f80[i][0, 4]],
            #         [peak_f[i][0, 5], mean_f[i][0, 5], median_f[i][0, 5], f80[i][0, 5]],
            #         [peak_f[i][0, 6], mean_f[i][0, 6], median_f[i][0, 6], f80[i][0, 6]],
            #         [peak_f[i][0, 7], mean_f[i][0, 7], median_f[i][0, 7], f80[i][0, 7]]]
            # for row in data:
            #     sheet1.cell(row=count2_freq, column=count_col, value=row[0])
            #     sheet1.cell(row=count2_freq, column=count_col + 1, value=row[1])
            #     sheet1.cell(row=count2_freq, column=count_col + 2, value=row[2])
            #     sheet1.cell(row=count2_freq, column=count_col + 3, value=row[3])
            #     count2_freq = count2_freq + 1

    wb2.save(file_excel)


def evolution_parameters_COP(std, velocity, area, COP_array):
    wb2 = load_workbook(file_excel)
    sheet1 = wb2.get_sheet_by_name(patient)

    fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type='solid')

    count = 407
    time = 410
    time2 = 409

    for i in area:
        if i != 'Reach_Ground' and i != 'Arms_extension':

            title = sheet1.cell('A' + str(count))
            title.value = "Evolution of COP parameters - " + str(i)
            title.font = title.font.copy(bold=True)


            first_col = sheet1.cell('A' + str(count + 1))
            first_col.value = ''
            first_col.font = first_col.font.copy(bold=True)
            first_col.fill = fill

            first1_col = sheet1.cell('A' + str(count + 2))
            first1_col.value = ''
            first1_col.font = first1_col.font.copy(bold=True)
            first1_col.fill = fill

            second_col = sheet1.cell('B' + str(count + 1))
            second_col.value = 'COP X'
            second_col.font = second_col.font.copy(bold=True)
            second_col.fill = fill

            third_col = sheet1.cell('D' + str(count + 1))
            third_col.value = 'COP_Y'
            third_col.font = third_col.font.copy(bold=True)
            third_col.fill = fill

            second1_col = sheet1.cell('B' + str(count + 2))
            second1_col.value = 'STD (mm)'
            second1_col.font = second1_col.font.copy(bold=True)
            second1_col.fill = fill

            third1_col1 = sheet1.cell('C' + str(count + 2))
            third1_col1.value = 'Velocity (mm/s)'
            third1_col1.font = third1_col1.font.copy(bold=True)
            third1_col1.fill = fill

            fourth1_coll = sheet1.cell('D' + str(count + 2))
            fourth1_coll.value = 'STD (mm)'
            fourth1_coll.font = fourth1_coll.font.copy(bold=True)
            fourth1_coll.fill = fill

            fifht1_col1 = sheet1.cell('E' + str(count + 2))
            fifht1_col1.value = 'Velocity (mm/s)'
            fifht1_col1.font = fifht1_col1.font.copy(bold=True)
            fifht1_col1.fill = fill

            area = sheet1.cell('H' + str(count + 1))
            area.value = 'Area (mm2)'
            area.font = area.font.copy(bold=True)
            area.fill = fill

            area1 = sheet1.cell('G' + str(count + 1))
            area1.value = ''
            area1.font = area1.font.copy(bold=True)
            area1.fill = fill

            if i != "Reach_C" and i != "Reach_R" and i !="Reach_L":

                start = 0
                finish = 2.5
                start_array = 0
                time_array = 2500


                for n in range(0, len(velocity[i]["COP_X"])):
                    row = sheet1.cell('A' + str(time))
                    row.value = '[' + str(start) + '-' + str(finish) + '] s'
                    row.font = row.font.copy(bold=True)
                    row.fill = fill

                    row_area = sheet1.cell('G' + str(time2))
                    row_area.value = '[' + str(start) + '-' + str(finish) + '] s'
                    row_area.font = row.font.copy(bold=True)
                    row_area.fill = fill

                    data = [[std[i]["COP_X"][n], velocity[i]["COP_X"][n], std[i]["COP_Y"][n], velocity[i]["COP_Y"][n]]]
                    for row in data:
                        sheet1.cell(row=time2 + 1, column=2, value=row[0])
                        sheet1.cell(row=time2 + 1, column=3, value=row[1])
                        sheet1.cell(row=time2 + 1, column=4, value=row[2])
                        sheet1.cell(row=time2 + 1, column=5, value=row[3])

                    area_traj = convex_hull(COP_array[i]["COP_X"][start_array:time_array], COP_array[i]["COP_Y"][start_array:time_array])
                    area = area_calc(area_traj)

                    area_values = sheet1.cell('H' + str(time2))
                    area_values.value = area



                    time += 1
                    time2 += 1
                    start = finish
                    finish = finish + 2.5

                    start_array = time_array
                    time_array = time_array + 2500

            else:
                start = 0
                finish = 1
                start_array = 0
                time_array = 1000

                for n in range(0, len(velocity[i]["COP_X"])):
                    row = sheet1.cell('A' + str(time))
                    row.value = '[' + str(start) + '-' + str(finish) + '] s'
                    row.font = row.font.copy(bold=True)
                    row.fill = fill

                    row_area = sheet1.cell('G' + str(time2))
                    row_area.value = '[' + str(start) + '-' + str(finish) + '] s'
                    row_area.font = row.font.copy(bold=True)
                    row_area.fill = fill

                    data = [[std[i]["COP_X"][n], velocity[i]["COP_X"][n], std[i]["COP_Y"][n], velocity[i]["COP_Y"][n]]]
                    for row in data:
                        sheet1.cell(row=time2 + 1, column=2, value=row[0])
                        sheet1.cell(row=time2 + 1, column=3, value=row[1])
                        sheet1.cell(row=time2 + 1, column=4, value=row[2])
                        sheet1.cell(row=time2 + 1, column=5, value=row[3])

                    area_traj = convex_hull(COP_array[i]["COP_X"][start_array:time_array],
                                            COP_array[i]["COP_Y"][start_array:time_array])
                    area = area_calc(area_traj)

                    area_values = sheet1.cell('H' + str(time2))
                    area_values.value = area

                    time += 1
                    time2 += 1
                    start = finish
                    finish = finish + 1

                    start_array = time_array
                    time_array = time_array + 1000

            count = count + 23
            time = count + 3
            time2 = count + 2
    wb2.save(file_excel)


def EMG_evolution(EMG_array):
    wb2 = load_workbook(file_excel)
    sheet1 = wb2.get_sheet_by_name(patient)

    fill = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type='solid')

    count = 407
    time = 409
    time2 = 409

    for i in EMG_array:
        if i != 'Reach_Ground' and i != 'Arms_extension':
            title = sheet1.cell('L' + str(count))
            title.value = "Evolution of EMG values in percentage - " + str(i)
            title.font = title.font.copy(bold=True)

            first_col = sheet1.cell('L' + str(count + 1))
            first_col.value = ''
            first_col.font = first_col.font.copy(bold=True)
            first_col.fill = fill

            second_col = sheet1.cell('M' + str(count + 1))
            second_col.value = 'Rectus_A L'
            second_col.font = second_col.font.copy(bold=True)
            second_col.fill = fill

            third_col = sheet1.cell('N' + str(count + 1))
            third_col.value = 'Rectus_A R'
            third_col.font = third_col.font.copy(bold=True)
            third_col.fill = fill

            fourth_col = sheet1.cell('O' + str(count + 1))
            fourth_col.value = 'Obliques L'
            fourth_col.font = fourth_col.font.copy(bold=True)
            fourth_col.fill = fill

            fifth_col = sheet1.cell('P' + str(count + 1))
            fifth_col.value = 'Obliques R'
            fifth_col.font = fifth_col.font.copy(bold=True)
            fifth_col.fill = fill

            sixth_col = sheet1.cell('Q' + str(count + 1))
            sixth_col.value = 'Ilicostalis L'
            sixth_col.font = sixth_col.font.copy(bold=True)
            sixth_col.fill = fill

            seventh_col = sheet1.cell('R' + str(count + 1))
            seventh_col.value = 'Ilicostalis R'
            seventh_col.font = seventh_col.font.copy(bold=True)
            seventh_col.fill = fill

            eighth_col = sheet1.cell('S' + str(count + 1))
            eighth_col.value = 'Multifidus L'
            eighth_col.font = eighth_col.font.copy(bold=True)
            eighth_col.fill = fill

            nineth_col = sheet1.cell('T' + str(count + 1))
            nineth_col.value = 'Multifidus R'
            nineth_col.font = nineth_col.font.copy(bold=True)
            nineth_col.fill = fill


            if i != "Reach_C" and i != "Reach_R" and i !="Reach_L":
                start = 0
                finish = 2.5

                for n in range(0, len(EMG_array[i][:,0])):
                    row = sheet1.cell('L' + str(time))
                    row.value = '[' + str(start) + '-' + str(finish) + '] s'
                    row.font = row.font.copy(bold=True)
                    row.fill = fill

                    data = [[EMG_array[i][n,0], EMG_array[i][n,1], EMG_array[i][n,2], EMG_array[i][n,3], EMG_array[i][n,4],
                            EMG_array[i][n, 5], EMG_array[i][n,6], EMG_array[i][n,7]]]
                    for row in data:
                        sheet1.cell(row=time, column=13, value=row[0])
                        sheet1.cell(row=time, column=14, value=row[1])
                        sheet1.cell(row=time, column=15, value=row[2])
                        sheet1.cell(row=time, column=16, value=row[3])
                        sheet1.cell(row=time, column=17, value=row[4])
                        sheet1.cell(row=time, column=18, value=row[5])
                        sheet1.cell(row=time, column=19, value=row[6])
                        sheet1.cell(row=time, column=20, value=row[7])


                    time += 1
                    time2 += 1
                    start = finish
                    finish = finish + 2.5

            else:
                start = 0
                finish = 1

                for n in range(0, len(EMG_array[i][:,0])):
                    row = sheet1.cell('L' + str(time))
                    row.value = '[' + str(start) + '-' + str(finish) + '] s'
                    row.font = row.font.copy(bold=True)
                    row.fill = fill

                    data = [[EMG_array[i][n,0], EMG_array[i][n,1], EMG_array[i][n,2], EMG_array[i][n,3], EMG_array[i][n,4],
                            EMG_array[i][n, 5], EMG_array[i][n,6], EMG_array[i][n,7]]]
                    for row in data:
                        sheet1.cell(row=time, column=13, value=row[0])
                        sheet1.cell(row=time, column=14, value=row[1])
                        sheet1.cell(row=time, column=15, value=row[2])
                        sheet1.cell(row=time, column=16, value=row[3])
                        sheet1.cell(row=time, column=17, value=row[4])
                        sheet1.cell(row=time, column=18, value=row[5])
                        sheet1.cell(row=time, column=19, value=row[6])
                        sheet1.cell(row=time, column=20, value=row[7])


                    time += 1
                    time2 += 1
                    start = finish
                    finish = finish + 1

            count = count + 23
            time = count + 2
            time2 = count + 2
    wb2.save(file_excel)



def mean_std_COP_evolution(std, velocity, area):
    wb2 = load_workbook(file_excel)
    sheet1 = wb2.get_sheet_by_name(patient)

    fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type='solid')

    count = 6
    time = 10
    time2 = 9

    for i in area:
        if i != 'Reach_Ground' and i != 'Arms_extension':

            title = sheet1.cell('X' + str(count))
            title.value = "Evolution of COP parameters - " + str(i)
            title.font = title.font.copy(bold=True)


            first_col = sheet1.cell('X' + str(count + 1))
            first_col.value = ''
            first_col.font = first_col.font.copy(bold=True)
            first_col.fill = fill

            first1_col = sheet1.cell('X' + str(count + 2))
            first1_col.value = ''
            first1_col.font = first1_col.font.copy(bold=True)
            first1_col.fill = fill

            second_col = sheet1.cell('Y' + str(count + 1))
            second_col.value = 'COP X'
            second_col.font = second_col.font.copy(bold=True)
            second_col.fill = fill

            third_col = sheet1.cell('AC' + str(count + 1))
            third_col.value = 'COP_Y'
            third_col.font = third_col.font.copy(bold=True)
            third_col.fill = fill

            second1_col = sheet1.cell('Y' + str(count + 2))
            second1_col.value = 'STD (mm)'
            second1_col.font = second1_col.font.copy(bold=True)
            second1_col.fill = fill

            third1_col1 = sheet1.cell('AA' + str(count + 2))
            third1_col1.value = 'Velocity (mm/s)'
            third1_col1.font = third1_col1.font.copy(bold=True)
            third1_col1.fill = fill

            fourth1_coll = sheet1.cell('AC' + str(count + 2))
            fourth1_coll.value = 'STD (mm)'
            fourth1_coll.font = fourth1_coll.font.copy(bold=True)
            fourth1_coll.fill = fill

            fifht1_col1 = sheet1.cell('AE' + str(count + 2))
            fifht1_col1.value = 'Velocity (mm/s)'
            fifht1_col1.font = fifht1_col1.font.copy(bold=True)
            fifht1_col1.fill = fill

            mean_std_x = sheet1.cell('Y' + str(count + 3))
            mean_std_x.value = 'Mean'
            mean_std_x.font = mean_std_x.font.copy(bold=True)
            mean_std_x.fill = fill

            std_std_x = sheet1.cell('Z' + str(count + 3))
            std_std_x.value = 'STD'
            std_std_x.font = std_std_x.font.copy(bold=True)
            std_std_x.fill = fill

            mean_v_x = sheet1.cell('AA' + str(count + 3))
            mean_v_x.value = 'Mean'
            mean_v_x.font = mean_v_x.font.copy(bold=True)
            mean_v_x.fill = fill

            std_v_x = sheet1.cell('AB' + str(count + 3))
            std_v_x.value = 'STD'
            std_v_x.font = std_v_x.font.copy(bold=True)
            std_v_x.fill = fill

            mean_std_y = sheet1.cell('AC' + str(count + 3))
            mean_std_y.value = 'Mean'
            mean_std_y.font = mean_std_y.font.copy(bold=True)
            mean_std_y.fill = fill

            std_std_y = sheet1.cell('AD' + str(count + 3))
            std_std_y.value = 'STD'
            std_std_y.font = std_std_y.font.copy(bold=True)
            std_std_y.fill = fill

            mean_v_y = sheet1.cell('AE' + str(count + 3))
            mean_v_y.value = 'Mean'
            mean_v_y.font = mean_v_y.font.copy(bold=True)
            mean_v_y.fill = fill

            std_v_y = sheet1.cell('AF' + str(count + 3))
            std_v_y.value = 'STD'
            std_v_y.font = std_v_y.font.copy(bold=True)
            std_v_y.fill = fill




            area = sheet1.cell('AJ' + str(count + 1))
            area.value = 'Area (mm2)'
            area.font = area.font.copy(bold=True)
            area.fill = fill

            area1 = sheet1.cell('AI' + str(count + 1))
            area1.value = ''
            area1.font = area1.font.copy(bold=True)
            area1.fill = fill

            area_std = sheet1.cell('AJ' + str(count + 2))
            area_std.value = 'STD'
            area_std.font = area_std.font.copy(bold=True)
            area_std.fill = fill

            area_mean = sheet1.cell('AK' + str(count + 2))
            area_mean.value = 'Mean'
            area_mean.font = area_mean.font.copy(bold=True)
            area_mean.fill = fill



            start = 0
            finish = 2.5


            for n in range(0, len(velocity[i]["COP_X"]) - 1):
                row = sheet1.cell('X' + str(time))
                row.value = '[' + str(start) + '-' + str(finish) + '] s'
                row.font = row.font.copy(bold=True)
                row.fill = fill

                row_area = sheet1.cell('AI' + str(time2))
                row_area.value = '[' + str(start) + '-' + str(finish) + '] s'
                row_area.font = row.font.copy(bold=True)
                row_area.fill = fill




                time += 1
                time2 += 1
                start = finish
                finish = finish + 2.5

            count = count + 17
            time = count + 4
            time2 = count + 3
    wb2.save(file_excel)

def mean_values_EMG(EMG_values):
    wb2 = load_workbook(file_excel)
    sheet1 = wb2.get_sheet_by_name(patient)
    count = 614
    count2 = 616
    count_col = 2
    for i in EMG_values:
        if i != 'Reach_Ground' and i != 'Arms_extension':
            title = sheet1.cell('B' + str(count))
            title.value = "Mean values of each muscle - " + str(i)
            title.font = title.font.copy(bold=True)

            first_col = sheet1.cell('A' + str(count + 1))
            first_col.value = ''
            first_col.font = first_col.font.copy(bold=True)


            second_col = sheet1.cell('B' + str(count + 1))
            second_col.value = 'Left'
            second_col.font = second_col.font.copy(bold=True)


            third_col = sheet1.cell('C' + str(count + 1))
            third_col.value = 'Right'
            third_col.font = third_col.font.copy(bold=True)


            first_row = sheet1.cell('A' + str(count + 2))
            first_row.value = 'Rectus_A'
            first_row.font = first_row.font.copy(bold=True)


            second_row = sheet1.cell('A' + str(count + 3))
            second_row.value = 'Obliques'
            second_row.font = second_row.font.copy(bold=True)


            third_row = sheet1.cell('A' + str(count + 4))
            third_row.value = 'Ilicostalis'
            third_row.font = third_row.font.copy(bold=True)


            fourth_row = sheet1.cell('A' + str(count + 5))
            fourth_row.value = 'Multifidus'
            fourth_row.font = fourth_row.font.copy(bold=True)



            data = [[EMG_values[i][0], EMG_values[i][1]],
                    [EMG_values[i][2], EMG_values[i][3]],
                    [EMG_values[i][4], EMG_values[i][5]],
                    [EMG_values[i][6], EMG_values[i][7]]
                    ]
            for row in data:
                sheet1.cell(row=count2, column=count_col, value=row[0])
                sheet1.cell(row=count2, column=count_col + 1, value=row[1])
                count2 = count2 + 1

            count = count + 13
            count2 = count2 + 9
    wb2.save(file_excel)

def rest_mean_values(rest_array):
    wb2 = load_workbook(file_excel)
    sheet1 = wb2.get_sheet_by_name(patient)

    count_max = 734
    count2 = 736
    count_col = 2

    for i in rest_array:
        if i == "Relax":

            title = sheet1.cell('B' + str(count_max))
            title.value = "Max values of each muscle - " + str(i)
            title.font = title.font.copy(bold=True)

            first_col = sheet1.cell('A' + str(count_max + 1))
            first_col.value = ''
            first_col.font = first_col.font.copy(bold=True)

            second_col = sheet1.cell('B' + str(count_max + 1))
            second_col.value = 'Left'
            second_col.font = second_col.font.copy(bold=True)

            third_col = sheet1.cell('C' + str(count_max + 1))
            third_col.value = 'Right'
            third_col.font = third_col.font.copy(bold=True)

            first_row = sheet1.cell('A' + str(count_max + 2))
            first_row.value = 'Rectus_A'
            first_row.font = first_row.font.copy(bold=True)

            second_row = sheet1.cell('A' + str(count_max + 3))
            second_row.value = 'Obliques'
            second_row.font = second_row.font.copy(bold=True)

            third_row = sheet1.cell('A' + str(count_max + 4))
            third_row.value = 'Ilicostalis'
            third_row.font = third_row.font.copy(bold=True)

            fourth_row = sheet1.cell('A' + str(count_max + 5))
            fourth_row.value = 'Multifidus'
            fourth_row.font = fourth_row.font.copy(bold=True)

            data = [[rest_array[i][0], rest_array[i][1]],
                    [rest_array[i][2], rest_array[i][3]],
                    [rest_array[i][4], rest_array[i][5]],
                    [rest_array[i][6], rest_array[i][7]]
                    ]
            for row in data:
                sheet1.cell(row=count2, column=count_col, value=row[0])
                sheet1.cell(row=count2, column=count_col + 1, value=row[1])
                count2 = count2 + 1

    wb2.save(file_excel)




